from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file, g
from i18n import make_T, SUPPORTED, LANG_LABELS
import psycopg2
import psycopg2.extras
import psycopg2.pool
import os
import hashlib
import sqlite3
import re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from functools import wraps
from io import BytesIO
from datetime import datetime, timezone, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
import time as _time
# openpyxl은 다운로드 함수 내에서 지연 import (cold start 150ms 절감)

app = Flask(__name__, template_folder=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates'))

# ── 지점 이메일 시드 (branch code → email) ────────────────────────────────────
# 이미지에서 읽은 지점별 담당자 이메일. init_db에서 email이 비어있는 지점에만 적용.
_BRANCH_EMAIL_SEEDS = {
    'GMP': 'zegmp@eastarjet.com',        # 김포
    'CJU': 'zecju@eastarjet.com',        # 제주
    'CJJ': 'zecjj@eastarjet.com',        # 청주
    'KUV': 'zekuv@eastarjet.com',        # 군산
    'PUS': 'zepus@eastarjet.com',        # 부산/울산
    'ICN': 'zeicn@eastarjet.com',        # 인천
    'NRT': 'zenrt@eastarjet.com',        # 나리타(도쿄)
    'KIX': 'zekix@eastarjet.com',        # 간사이(오사카)
    'FUK': 'zefuk@eastarjet.com',        # 후쿠오카
    'TSA': 'zetsa@eastarjet.com',        # 송산(타이베이)
    'TPE': 'zetpe@eastarjet.com',        # 타오위안
    'BKK': 'zebkk@eastarjet.com',        # 방콕
    'DAD': 'zedadd2@eastarjet.com',      # 다낭
    'CXR': 'zecxr@eastarjet.com',        # 나트랑
    'PVG': 'zepvg@eastarjet.com',        # 상하이
    'PQC': 'zepqc@eastarjet.com',        # 푸꾸옥(무우쿡)
    'CNX': 'zecnx@eastarjet.com',        # 치앙마이
    'CTS': 'zects@eastarjet.com',        # 삿포로(지토세)
    'OKA': 'zeoka@eastarjet.com',        # 오키나와
    'KMJ': 'zekmi@eastarjet.com',        # 구마모토
    'YNJ': 'ls1012cv@eastarjet.com',     # 연길
    'CGO': 'rjl@eastarjet.com',          # 정저우
    'YNT': 'zeynt@eastarjet.com',        # 옌타이
    'MDC': 'zemdc@eastarjet.com',        # 마나도
    'TKS': 'zekcj@eastarjet.com',        # 도쿠시마
    'HKG': 'lsn.kim@eastarjet.com',      # 홍콩
    'ALA': 'zeala@eastarjet.com',        # 알마티
}

# ── 카탈로그 기본 카테고리 순서 ───────────────────────────────────────────────
CAT_ORDER = ['X-Banner', '저울', '스탠션', '아크릴 거치대', 'DESK TOP', '안내문', '카운터 종합 안내문', '간판·표찰', '스탬프·쿠폰']

# ── 운송아이템 카탈로그 기본 항목 (DB 시딩용) ─────────────────────────────────
CATALOG_ITEMS = [
    # X-Banner (15종)
    {'code':'xb01','img':'img_28','name':'승객 수하물 수취대 안내',         'cat':'X-Banner', 'sub_desc':'Baggage Claim Area<br>내용 추가 / 로고 변경',             'sort':10},
    {'code':'xb02','img':'img_31','name':'유모차 수취 안내',                 'cat':'X-Banner', 'sub_desc':'Baby Stroller Pick-up<br>로고 변경하여 진행',              'sort':20},
    {'code':'xb03','img':'img_25','name':'지연·결항·탑승구 변경 안내',       'cat':'X-Banner', 'sub_desc':'600×1800mm<br>중국어 / 대만어 / 일본어 3종',               'sort':30},
    {'code':'xb04','img':'img_21','name':'탑승수속 카운터',                  'cat':'X-Banner', 'sub_desc':'Check In Counter<br>チェックインカウンター / 办理手续柜台',   'sort':40},
    {'code':'xb05','img':'img_20','name':'탑승구',                           'cat':'X-Banner', 'sub_desc':'Boarding Gate',                                            'sort':50},
    {'code':'xb06','img':'img_07','name':'여권/탑승권 제시 안내',            'cat':'X-Banner', 'sub_desc':'600×1800mm / 2가지 시안<br>한·영·일·중 4개 언어',           'sort':60},
    {'code':'xb07','img':'img_44','name':'위탁수하물 금지물품 안내',         'cat':'X-Banner', 'sub_desc':'라이터 / 전자담배 / 배터리<br>3가지 시안',                   'sort':70},
    {'code':'xb12','img':'img_45','name':'위탁수하물 금지물품 안내 (대만)',  'cat':'X-Banner', 'sub_desc':'易斯達航空 / 대만 노선 전용<br>번체 중국어',                  'sort':80},
    {'code':'xb08','img':'img_12','name':'창문 닫이 안내 (군사공항용)',      'cat':'X-Banner', 'sub_desc':'연길 공항 전용<br>중국어(간/번체) 2종',                       'sort':90},
    {'code':'xb09','img':'img_36','name':'베트남용 ATC 안내',                'cat':'X-Banner', 'sub_desc':'지연 ATC 관련 안내문<br>베트남 노선 전용',                    'sort':100},
    {'code':'xb10','img':'img_34','name':'노선 안내',                        'cat':'X-Banner', 'sub_desc':'예) 인천 ↕ 팔라완<br>로고 변경하여 진행',                    'sort':110},
    {'code':'xb11','img':'img_39','name':'탑승 수속 안내 (운영시간)',        'cat':'X-Banner', 'sub_desc':'ZE[ ] COUNTER / 카운터 운영시간<br>チェックイン(Check-In)',   'sort':120},
    {'code':'sc01','img':'img_13','name':'기내 반입 수하물 안내 (저울 버전)','cat':'X-Banner', 'sub_desc':'55×40×20cm / 10kg<br>저울 버전 2개 + 저울 2개',              'sort':130},
    {'code':'rb03','img':'img_47','name':'INFORMATION',                      'cat':'X-Banner', 'sub_desc':'운영시간 안내 (한·중 2개 언어)',                              'sort':140},
    {'code':'rb04','img':'img_48','name':'국제선 대형수하물 위탁 카운터',    'cat':'X-Banner', 'sub_desc':'골프채 등 대형수하물 / 방향 표시',                            'sort':150},
    # 저울 (1종)
    {'code':'sc02','img':'img_38','name':'테스트 유닛',                      'cat':'저울',     'sub_desc':'기내반입 수하물 규격 확인용<br>내부 눈금(자) 포함',             'sort':10},
    # 스탠션 사인 꽂이 (4종)
    {'code':'st01','img':'img_02','name':'고정형 — 입구/출구 안내',          'cat':'스탠션',   'sub_desc':'300×300mm / 입구·출구·기다리는 곳',                           'sort':10},
    {'code':'st02','img':'img_29','name':'고정형 — 위탁수하물 금지물품',     'cat':'스탠션',   'sub_desc':'배터리 / 라이터 금지',                                         'sort':20},
    {'code':'st03','img':'img_04','name':'고정형 — 셀프체크인 승객 구분',   'cat':'스탠션',   'sub_desc':'CHECK-IN / BAG DROP / 2종 세트',                              'sort':30},
    {'code':'st04','img':'img_32','name':'고정형 — 셀프체크인 수하물 전용', 'cat':'스탠션',   'sub_desc':'KIOSK Baggage Check-in Counter',                              'sort':40},
    # 아크릴 거치대 (2종)
    {'code':'ac01','img':'img_01','name':'아크릴 거치대 (A4사이즈)',         'cat':'아크릴 거치대','sub_desc':'카운터용 A4 안내문 거치대',                              'sort':10},
    {'code':'ac02','img':'img_43','name':'약관 거치대 아크릴',               'cat':'아크릴 거치대','sub_desc':'290×320mm',                                             'sort':20},
    # DESK TOP SIGN (1종)
    {'code':'ds01','img':'img_46','name':'카운터 이석 및 수속마감 안내',     'cat':'DESK TOP', 'sub_desc':'360×170mm / 재질변경<br>한·영·일·중 4개 언어',               'sort':10},
    # 안내문 (2종)
    {'code':'nt01','img':'img_23','name':'탑승순서 안내문',                  'cat':'안내문',            'sub_desc':'580×1040mm<br>1~4단계 우선순서 표기',               'sort':10},
    {'code':'nt02','img':'img_11','name':'카운터 종합 안내문',               'cat':'안내문',            'sub_desc':'안내 내용 변경 가능<br>카운터 사이즈 확인 필요',      'sort':20},
    # 카운터 종합 안내문 (구 ROW BRD, 2종)
    {'code':'rb01','img':'img_05','name':'탑승순서 안내 (양면)',             'cat':'카운터 종합 안내문', 'sub_desc':'우선탑승 / BOARDING ZONE<br>사이즈·언어변경 가능', 'sort':10},
    {'code':'rb02','img':'img_06','name':'탑승중 안내',                      'cat':'카운터 종합 안내문', 'sub_desc':'NOW BOARDING',                                      'sort':20},
    # 간판·표찰 (3종)
    {'code':'sg01','img':'img_03','name':'사무실 간판 (부착형)',              'cat':'간판·표찰','sub_desc':'이스타항공 / イースタ-航空 / 易斯达航空公司<br>언어변경 가능', 'sort':10},
    {'code':'sg02','img':'img_42','name':'항공기피해 구제접수처 — 카운터형', 'cat':'간판·표찰','sub_desc':'200×155mm',                                                  'sort':20},
    {'code':'sg03','img':'img_17','name':'항공기피해 구제접수처 — 벽면부착형','cat':'간판·표찰','sub_desc':'190×100mm',                                                 'sort':30},
    # 스탬프·쿠폰 (4종)
    {'code':'sp01','img':'img_10','name':'GD 날인 스탬프 (중국노선)',        'cat':'스탬프·쿠폰','sub_desc':'지점명 변경하여 제작',                                    'sort':10},
    {'code':'sp02','img':'img_24','name':'AOC 도장 (인천)',                  'cat':'스탬프·쿠폰','sub_desc':'20mm 원형 / CORRECTION / AOC',                             'sort':20},
    {'code':'sp03','img':'img_27','name':'밀쿠폰 (Meal Coupon)',             'cat':'스탬프·쿠폰','sub_desc':'10,000원 / 사이즈·금액 변경 가능',                         'sort':30},
    {'code':'sp04','img':'img_26','name':'밀쿠폰 날짜 스탬프',               'cat':'스탬프·쿠폰','sub_desc':'38mm / EASTAR JET 로고 포함<br>날짜변경 가능',             'sort':40},
]
_secret = os.environ.get('SECRET_KEY', '')
if not _secret:
    import warnings
    warnings.warn('[SECURITY] SECRET_KEY 환경변수가 설정되지 않았습니다. 프로덕션에서는 반드시 설정하세요.')
    _secret = 'dev-secret-key-change-in-production'
app.secret_key = _secret
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

DATABASE_URL = os.environ.get('DATABASE_URL', '')

# ── 메일 발송 설정 ────────────────────────────────────────────────────────────
MAIL_HOST = os.environ.get('MAIL_HOST', '')
MAIL_PORT = int(os.environ.get('MAIL_PORT', '587'))
MAIL_USER = os.environ.get('MAIL_USER', '')
MAIL_PASS = os.environ.get('MAIL_PASS', '')
MAIL_FROM = os.environ.get('MAIL_FROM', MAIL_USER)
# URL → 키워드 인자 dict 변환 (psycopg2.connect(**_PG)로 사용)
# URL 파싱 대신 키워드 인자를 쓰면 비밀번호 특수문자([, ], @ 등) 인코딩 불필요
_PG = {}
if DATABASE_URL:
    try:
        # 1차 시도: urlparse (Python < 3.12 또는 안전한 비밀번호)
        from urllib.parse import urlparse, unquote
        _p = urlparse(DATABASE_URL)
        _PG = {
            'host':    _p.hostname,
            'port':    _p.port or 5432,
            'user':    _p.username,
            'password': unquote(_p.password) if _p.password else '',
            'dbname':  (_p.path or '/postgres').lstrip('/') or 'postgres',
            'sslmode': 'require',
        }
    except Exception:
        # 2차 시도: 수동 파싱 — Python 3.12에서 [ 포함 비밀번호 URL 파싱 실패 시
        # rfind('@')로 userinfo/host 경계를 찾고 raw 비밀번호를 그대로 사용
        try:
            _pi   = DATABASE_URL.index('://')
            _rest = DATABASE_URL[_pi + 3:]          # user:pass@host:port/db
            _sl   = _rest.find('/')
            _netloc = _rest[:_sl] if _sl >= 0 else _rest
            _path   = _rest[_sl + 1:] if _sl >= 0 else 'postgres'
            _at   = _netloc.rfind('@')
            _userinfo = _netloc[:_at]
            _hostport = _netloc[_at + 1:]
            _ci   = _userinfo.index(':')
            _user = _userinfo[:_ci]
            _pw   = _userinfo[_ci + 1:]             # raw 비밀번호 — 인코딩 불필요
            _hp   = _hostport.rsplit(':', 1)
            _PG = {
                'host':    _hp[0],
                'port':    int(_hp[1]) if len(_hp) > 1 else 5432,
                'user':    _user,
                'password': _pw,
                'dbname':  _path.split('?')[0] or 'postgres',
                'sslmode': 'require',
            }
        except Exception:
            _PG = {}
ALLOWED_IPS  = [ip.strip() for ip in os.environ.get('ALLOWED_IPS', '').split(',') if ip.strip()]
USE_SQLITE   = not _PG
_catalog_table_ready   = False  # warm 인스턴스에서 DDL 재실행 방지
_user_deleted_migrated = False  # user_deleted 컬럼 마이그레이션 완료 여부
_img_tmp_warmed        = False  # 이미지 /tmp 일괄 캐시 완료 여부
_DB_INITIALIZED        = False  # cold start당 init_db 1회만 실행
_PG_POOL               = None   # psycopg2 ThreadedConnectionPool (PG 전용)
# ── 성능 캐시 설정 ────────────────────────────────────────────────────────────
_NOTIF_TTL           = 30   # 알림 카운트 세션 캐시 TTL (초) — 매 요청 DB 조회 방지
# Vercel 환경에서는 /tmp만 쓰기 가능 — 로컬은 프로젝트 폴더 사용
_local_db    = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'inventory.db')
SQLITE_DB    = '/tmp/inventory.db' if not os.access(os.path.dirname(_local_db), os.W_OK) else _local_db


# ── DB 연결 ───────────────────────────────────────────────────────────────────

class _DB:
    """psycopg2를 sqlite3 인터페이스처럼 사용하기 위한 래퍼"""
    def __init__(self, conn, pool=None):
        self._conn = conn
        self._pool = pool       # ThreadedConnectionPool 참조 (있으면 putconn으로 반환)
        self._returned = False  # 이중 반환 방지

    def execute(self, sql, params=()):
        cur = self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params if params else None)
        return cur

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        if self._returned:
            return
        self._returned = True
        if self._pool and not self._conn.closed:
            try:
                self._pool.putconn(self._conn)
            except Exception:
                try:
                    self._conn.close()
                except Exception:
                    pass
        elif not self._conn.closed:
            try:
                self._conn.close()
            except Exception:
                pass


def _pg_to_sqlite(sql):
    """PostgreSQL SQL을 SQLite 호환 구문으로 변환"""
    # TO_CHAR 변환 (INTERVAL 포함 케이스 먼저)
    sql = re.sub(
        r"TO_CHAR\(CURRENT_DATE\s*-\s*INTERVAL\s*'(\d+)\s*months',\s*'YYYY-MM'\)",
        r"strftime('%Y-%m', date('now', '-\1 months'))", sql)
    sql = re.sub(r"TO_CHAR\(CURRENT_DATE,\s*'YYYY-MM'\)", "strftime('%Y-%m', 'now')", sql)
    sql = re.sub(r"TO_CHAR\(([^,]+),\s*'YYYY-MM'\)", r"strftime('%Y-%m', \1)", sql)
    # INTERVAL 날짜 산술
    sql = re.sub(r"CURRENT_DATE\s*-\s*INTERVAL\s*'(\d+)\s*months'", r"date('now', '-\1 months')", sql)
    sql = re.sub(r"CURRENT_DATE\s*-\s*INTERVAL\s*'(\d+)\s*days'", r"date('now', '-\1 days')", sql)
    # 타입 캐스트
    sql = re.sub(r'(\w+(?:\.\w+)*)::date\b', r'date(\1)', sql)
    sql = re.sub(r'(\w+(?:\.\w+)*)::text\b', r'\1', sql)
    # NOW() → CURRENT_TIMESTAMP
    sql = re.sub(r'\bNOW\(\)', 'CURRENT_TIMESTAMP', sql, flags=re.IGNORECASE)
    # 플레이스홀더
    sql = sql.replace('%s', '?')
    return sql


class _SQLiteDB:
    def __init__(self, conn):
        self._conn = conn

    def execute(self, sql, params=()):
        sql = _pg_to_sqlite(sql)
        cur = self._conn.cursor()
        cur.execute(sql, list(params) if params else [])
        return cur

    def commit(self):   self._conn.commit()
    def rollback(self): self._conn.rollback()
    def close(self):    self._conn.close()


def _get_pg_pool():
    """모듈 단위 ThreadedConnectionPool을 지연 생성해 반환."""
    global _PG_POOL
    if _PG_POOL is None and _PG:
        _PG_POOL = psycopg2.pool.ThreadedConnectionPool(1, 5, **_PG)
    return _PG_POOL


def get_db():
    if 'db' not in g:
        if USE_SQLITE:
            conn = sqlite3.connect(SQLITE_DB)
            conn.row_factory = sqlite3.Row
            g.db = _SQLiteDB(conn)
        else:
            pool = _get_pg_pool()
            g.db = _DB(pool.getconn(), pool=pool)
    else:
        db = g.db
        # 연결이 닫혔거나 이미 풀에 반환된 경우 새 연결 획득
        if not USE_SQLITE and (
            getattr(db, '_returned', False) or
            (hasattr(db, '_conn') and db._conn.closed)
        ):
            pool = _get_pg_pool()
            g.db = _DB(pool.getconn(), pool=pool)
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db is not None:
        try:
            db.close()
        except Exception:
            pass


def hash_pw(pw):
    return generate_password_hash(pw, method='pbkdf2:sha256:600000')


_MAIL_FOOTER_TXT = (
    '\n\n※ 본 메일은 발신전용입니다. '
    '문의사항은 해당 담당자에게 직접 연락하여 주시기 바랍니다.'
)
_MAIL_FOOTER_HTML = (
    '<p style="margin-top:20px;font-size:12px;color:#cc1625;">'
    '※ 본 메일은 발신전용입니다. '
    '문의사항은 해당 담당자에게 직접 연락하여 주시기 바랍니다.</p>'
)

def send_mail(to_list, subject, body):
    """to_list: 이메일 주소 리스트. MAIL_HOST 미설정 시 무시."""
    if not MAIL_HOST or not MAIL_USER or not to_list:
        return
    recipients = [addr for addr in to_list if addr and addr.strip()]
    if not recipients:
        return
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = MAIL_FROM
        msg['To']      = ', '.join(recipients)
        msg.attach(MIMEText(_MAIL_FOOTER_TXT.strip() + '\n\n' + body, 'plain', 'utf-8'))
        body_html = _MAIL_FOOTER_HTML + '<div style="font-family:sans-serif;font-size:14px;white-space:pre-wrap">' + body.replace('\n', '<br>') + '</div>'
        msg.attach(MIMEText(body_html, 'html', 'utf-8'))
        with smtplib.SMTP(MAIL_HOST, MAIL_PORT, timeout=10) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(MAIL_USER, MAIL_PASS)
            server.sendmail(MAIL_FROM, recipients, msg.as_string())
    except Exception as e:
        app.logger.warning(f'[send_mail] 발송 실패: {e}')


def verify_pw(pw, stored):
    if stored.startswith('pbkdf2:') or stored.startswith('scrypt:'):
        return check_password_hash(stored, pw)
    # 레거시 SHA256 (솔트 없음) — 검증 후 자동 재해시 처리됨
    return hashlib.sha256(pw.encode()).hexdigest() == stored


def validate_password(pw):
    if len(pw) < 8:
        return '비밀번호는 8자 이상이어야 합니다.'
    kinds = sum([
        bool(re.search(r'[A-Z]', pw)),
        bool(re.search(r'[a-z]', pw)),
        bool(re.search(r'[0-9]', pw)),
        bool(re.search(r'[^A-Za-z0-9]', pw)),
    ])
    if kinds < 3:
        return '영문 대/소문자, 숫자, 특수문자 중 3종류 이상 포함해야 합니다.'
    return None


def _to_dt(val):
    """str/datetime 모두 timezone-aware datetime으로 변환. 실패 시 None 반환."""
    if val is None:
        return None
    if isinstance(val, str):
        try:
            val = datetime.fromisoformat(val.replace(' ', 'T'))
        except Exception:
            return None
    if hasattr(val, 'tzinfo') and val.tzinfo is None:
        val = val.replace(tzinfo=timezone.utc)
    return val


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def _client_ip():
    forwarded = request.headers.get('X-Forwarded-For', '')
    return forwarded.split(',')[0].strip() if forwarded else (request.remote_addr or '')


def log_action(action, target_info=''):
    try:
        conn = get_db()
        conn.execute(
            'INSERT INTO access_logs (user_id, username, action, target_info, ip_address, created_at) '
            'VALUES (%s, %s, %s, %s, %s, NOW())',
            (session.get('user_id'), session.get('username'), action, target_info, _client_ip())
        )
        conn.commit()
        conn.close()
    except Exception:
        pass


# ── 이메일 전송 헬퍼 ──────────────────────────────────────────────────────────

def send_email(to_addr, subject, body_html):
    """환경변수: SMTP_HOST, SMTP_PORT(기본587), SMTP_USER, SMTP_PASS, SMTP_FROM"""
    smtp_host = os.environ.get('SMTP_HOST', '')
    smtp_user = os.environ.get('SMTP_USER', '')
    if not smtp_host or not smtp_user:
        app.logger.warning(f'SMTP 미설정, 메일 전송 생략: to={to_addr}')
        return False
    try:
        smtp_port = int(os.environ.get('SMTP_PORT', '587'))
        smtp_pass = os.environ.get('SMTP_PASS', '')
        smtp_from = os.environ.get('SMTP_FROM', smtp_user)
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = smtp_from
        msg['To']      = to_addr
        msg.attach(MIMEText(body_html, 'html', 'utf-8'))
        with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as srv:
            srv.ehlo(); srv.starttls(); srv.login(smtp_user, smtp_pass)
            srv.sendmail(smtp_from, [to_addr], msg.as_string())
        return True
    except Exception as e:
        app.logger.error(f'메일 전송 실패 ({to_addr}): {e}')
        return False


# ── 운송양식 신청 기간 헬퍼 ────────────────────────────────────────────────

def _get_active_supply_period():
    """현재 활성화된 신청 기간 반환. 없으면 None."""
    conn = get_db()
    ph = '%s' if not USE_SQLITE else '?'
    today = datetime.now().strftime('%Y-%m-%d')
    row = conn.execute(
        f"SELECT * FROM form_supply_settings WHERE is_enabled=1 AND period_start<={ph} AND period_end>={ph} ORDER BY id DESC LIMIT 1",
        (today, today)
    ).fetchone()
    conn.close()
    return row


# ── Jinja2 필터 ──────────────────────────────────────────────────────────────

@app.template_filter('dt')
def dt_filter(value):
    if value is None:
        return '—'
    return str(value)[:16]


# ── 오류 핸들러 (디버그용) ────────────────────────────────────────────────────

@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    tb = traceback.format_exc()
    print(f"[UNHANDLED ERROR] {tb}")
    return f"<pre style='color:red'>{tb}</pre>", 500


# ── 템플릿 전역 컨텍스트 ────────────────────────────────────────────────────────

@app.context_processor
def inject_globals():
    notif_count   = 0
    notif_list    = []
    pw_expire_days = None
    bid  = session.get('branch_id')
    role = session.get('role')

    if 'user_id' in session:
        # 비밀번호 만료일 계산 (DB 불필요)
        changed_ts = session.get('_pwd_changed_at', 0)
        if changed_ts:
            elapsed = datetime.now(timezone.utc).timestamp() - changed_ts
            pw_expire_days = int(180 - elapsed / 86400)

        now_ts = datetime.now(timezone.utc).timestamp()
        # 30초 캐시: 매 페이지 요청마다 DB 조회하지 않음
        if now_ts - session.get('_notif_ts', 0) < _NOTIF_TTL:
            notif_count = session.get('_notif_count', 0)
            # notif_list는 캐시 미적용(빈 목록) — 카운트 배지만 캐시
        else:
            try:
                conn = get_db()
                if role == 'admin':
                    r = conn.execute(
                        'SELECT COUNT(*) AS cnt FROM transfer_requests WHERE status=%s',
                        ('PENDING',)
                    ).fetchone()
                    notif_count = int(r['cnt']) if r else 0
                elif bid:
                    r = conn.execute(
                        'SELECT '
                        '  (SELECT COUNT(*) FROM transfer_requests WHERE from_branch_id=%s AND status=%s) AS pending,'
                        '  (SELECT COUNT(*) FROM notifications WHERE branch_id=%s AND is_read=0) AS unread',
                        (bid, 'PENDING', bid)
                    ).fetchone()
                    notif_count = (int(r['pending']) + int(r['unread'])) if r else 0
                    notif_list  = conn.execute(
                        'SELECT id, message, created_at FROM notifications '
                        'WHERE branch_id=%s AND is_read=0 ORDER BY id DESC LIMIT 10',
                        (bid,)
                    ).fetchall()
                conn.close()
            except Exception:
                pass
            session['_notif_count'] = notif_count
            session['_notif_ts']    = now_ts

    return {'notif_count': notif_count, 'notif_list': notif_list, 'endpoint': request.endpoint, 'pw_expire_days': pw_expire_days,
            'now_dt': datetime.now(timezone.utc).astimezone()}


# ── IP 화이트리스트 + 세션 타임아웃 ─────────────────────────────────────────

@app.before_request
def check_ip_and_session():
    # IP 화이트리스트 (/ping은 Vercel Cron이 호출하므로 제외)
    if ALLOWED_IPS and request.path != '/ping':
        client_ip = _client_ip()
        if client_ip not in ALLOWED_IPS:
            return render_template('403.html'), 403

    # 비밀번호 만료: 당일(<=0)만 강제 리디렉트, 7일 이하는 배너 경고만
    if 'user_id' in session and request.endpoint not in ('change_password', 'logout', 'set_lang', None):
        changed_ts = session.get('_pwd_changed_at', 0)
        if changed_ts:
            elapsed = datetime.now(timezone.utc).timestamp() - changed_ts
            days_left = int(180 - elapsed / 86400)
            if days_left <= 0:
                _t = make_T(session.get('lang', 'ko'))
                flash(_t('chpw.pw_expired_flash'), 'danger')
                return redirect(url_for('change_password'))

    # 30분 유휴 세션 만료 (API 엔드포인트 제외)
    if 'user_id' in session and not request.path.startswith('/api/'):
        last = session.get('_last_activity', 0)
        now_ts = datetime.now(timezone.utc).timestamp()
        if now_ts - last > 1800:
            uid = session.get('user_id')
            uname = session.get('username')
            session.clear()
            try:
                conn = get_db()
                conn.execute(
                    'INSERT INTO access_logs (user_id, username, action, ip_address, created_at) '
                    'VALUES (%s, %s, %s, %s, NOW())',
                    (uid, uname, '세션만료_자동로그아웃', _client_ip())
                )
                conn.commit()
                conn.close()
            except Exception:
                pass
            flash('30분간 활동이 없어 자동 로그아웃 되었습니다.', 'warning')
            return redirect(url_for('login'))
        session['_last_activity'] = now_ts
        session.permanent = True


# ── DB 초기화 ─────────────────────────────────────────────────────────────────

def init_db():
    global _DB_INITIALIZED
    if _DB_INITIALIZED:
        return
    # init_db 전용 직접 연결 — 풀을 거치지 않아 teardown과 충돌 없음
    if USE_SQLITE:
        _raw = sqlite3.connect(SQLITE_DB)
        _raw.row_factory = sqlite3.Row
        conn = _SQLiteDB(_raw)
    else:
        conn = _DB(psycopg2.connect(**_PG))
    try:
        if USE_SQLITE:
            conn.execute('''
                CREATE TABLE IF NOT EXISTS branches (
                    id   INTEGER PRIMARY KEY,
                    code TEXT NOT NULL UNIQUE,
                    name TEXT NOT NULL,
                    type TEXT NOT NULL
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS form_types (
                    id            INTEGER PRIMARY KEY,
                    name          TEXT NOT NULL UNIQUE,
                    unit          TEXT NOT NULL,
                    unit_detail   TEXT,
                    unit_price    INTEGER DEFAULT 0,
                    min_threshold INTEGER DEFAULT 2
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS inventory (
                    id            INTEGER PRIMARY KEY,
                    branch_id     INTEGER NOT NULL,
                    form_type_id  INTEGER NOT NULL,
                    quantity      INTEGER DEFAULT 0,
                    last_updated  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (branch_id)    REFERENCES branches(id),
                    FOREIGN KEY (form_type_id) REFERENCES form_types(id),
                    UNIQUE(branch_id, form_type_id)
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS transactions (
                    id               INTEGER PRIMARY KEY,
                    type             TEXT NOT NULL,
                    form_type_id     INTEGER NOT NULL,
                    from_branch_id   INTEGER,
                    to_branch_id     INTEGER,
                    quantity         INTEGER NOT NULL,
                    notes            TEXT,
                    created_by       TEXT,
                    created_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    period_month     TEXT,
                    transaction_date DATE DEFAULT (date('now')),
                    FOREIGN KEY (form_type_id)   REFERENCES form_types(id),
                    FOREIGN KEY (from_branch_id) REFERENCES branches(id),
                    FOREIGN KEY (to_branch_id)   REFERENCES branches(id)
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id                  INTEGER PRIMARY KEY,
                    username            TEXT NOT NULL UNIQUE,
                    password            TEXT NOT NULL,
                    branch_id           INTEGER,
                    role                TEXT DEFAULT 'staff',
                    failed_attempts     INTEGER NOT NULL DEFAULT 0,
                    locked_until        TIMESTAMP,
                    last_login_at       TIMESTAMP,
                    password_changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (branch_id) REFERENCES branches(id)
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS access_logs (
                    id          INTEGER PRIMARY KEY,
                    user_id     INTEGER,
                    username    TEXT,
                    action      TEXT NOT NULL,
                    target_info TEXT,
                    ip_address  TEXT,
                    created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS transfer_requests (
                    id              INTEGER PRIMARY KEY,
                    from_branch_id  INTEGER NOT NULL,
                    to_branch_id    INTEGER NOT NULL,
                    form_type_id    INTEGER NOT NULL,
                    quantity        INTEGER NOT NULL,
                    notes           TEXT,
                    status          TEXT DEFAULT 'PENDING',
                    reject_reason   TEXT,
                    requested_by    TEXT NOT NULL,
                    approved_by     TEXT,
                    created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (from_branch_id) REFERENCES branches(id),
                    FOREIGN KEY (to_branch_id)   REFERENCES branches(id),
                    FOREIGN KEY (form_type_id)   REFERENCES form_types(id)
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS notifications (
                    id                  INTEGER PRIMARY KEY,
                    branch_id           INTEGER NOT NULL,
                    message             TEXT NOT NULL,
                    transfer_request_id INTEGER,
                    is_read             INTEGER DEFAULT 0,
                    created_at          TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (branch_id) REFERENCES branches(id)
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS flight_schedule (
                    id           INTEGER PRIMARY KEY,
                    branch_id    INTEGER NOT NULL,
                    year_month   TEXT NOT NULL,
                    flight_count INTEGER NOT NULL DEFAULT 0,
                    updated_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(branch_id, year_month),
                    FOREIGN KEY (branch_id) REFERENCES branches(id)
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS catalog_branch_items (
                    id         INTEGER PRIMARY KEY,
                    branch_id  INTEGER NOT NULL,
                    item_code  TEXT NOT NULL,
                    quantity   INTEGER DEFAULT 1,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (branch_id) REFERENCES branches(id),
                    UNIQUE(branch_id, item_code)
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS catalog_defs (
                    code       TEXT PRIMARY KEY,
                    img        TEXT NOT NULL,
                    name       TEXT NOT NULL,
                    cat        TEXT NOT NULL,
                    sub_desc   TEXT NOT NULL DEFAULT '',
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    img_data   TEXT NOT NULL DEFAULT ''
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS catalog_requests (
                    id            INTEGER PRIMARY KEY,
                    branch_id     INTEGER NOT NULL REFERENCES branches(id),
                    status        TEXT NOT NULL DEFAULT 'draft',
                    notes         TEXT DEFAULT '',
                    reject_reason TEXT DEFAULT '',
                    created_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS catalog_request_items (
                    id              INTEGER PRIMARY KEY,
                    request_id      INTEGER NOT NULL REFERENCES catalog_requests(id),
                    item_code       TEXT NOT NULL,
                    quantity        INTEGER DEFAULT 1,
                    custom_img_data TEXT DEFAULT '',
                    custom_text     TEXT DEFAULT '',
                    UNIQUE(request_id, item_code)
                )
            ''')
            # ── 운송양식 공급 신청 시스템 ─────────────────────────────
            conn.execute('''
                CREATE TABLE IF NOT EXISTS form_supply_settings (
                    id           INTEGER PRIMARY KEY,
                    period_start TEXT NOT NULL,
                    period_end   TEXT NOT NULL,
                    is_enabled   INTEGER DEFAULT 1,
                    created_by   TEXT NOT NULL,
                    updated_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS form_supply_requests (
                    id            INTEGER PRIMARY KEY,
                    branch_id     INTEGER NOT NULL,
                    status        TEXT DEFAULT 'pending',
                    notes         TEXT DEFAULT '',
                    reject_reason TEXT DEFAULT '',
                    requested_by  TEXT NOT NULL,
                    processed_by  TEXT DEFAULT '',
                    processed_at  TIMESTAMP,
                    created_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (branch_id) REFERENCES branches(id)
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS form_supply_request_items (
                    id           INTEGER PRIMARY KEY,
                    request_id   INTEGER NOT NULL,
                    form_type_id INTEGER NOT NULL,
                    quantity     INTEGER NOT NULL DEFAULT 1,
                    FOREIGN KEY (request_id)   REFERENCES form_supply_requests(id),
                    FOREIGN KEY (form_type_id) REFERENCES form_types(id)
                )
            ''')
            # 기존 DB 마이그레이션 — 컬럼 추가
            for _alter in [
                "ALTER TABLE transfer_requests ADD COLUMN notify_email TEXT DEFAULT ''",
                "ALTER TABLE branches ADD COLUMN email TEXT DEFAULT ''",
            ]:
                try:
                    conn.execute(_alter)
                    conn.commit()
                except Exception:
                    pass
        else:
            # 기존 DB 여부 확인 — 존재하면 무거운 CREATE TABLE DDL 건너뜀 (cold start 단축)
            _r = conn.execute("SELECT to_regclass('public.users') AS t").fetchone()
            if not (_r and _r['t']):
                # 첫 배포만: 테이블 전체 생성 (1 round trip)
                conn.execute('''
                CREATE TABLE IF NOT EXISTS branches (
                    id   SERIAL PRIMARY KEY,
                    code TEXT NOT NULL UNIQUE,
                    name TEXT NOT NULL,
                    type TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS form_types (
                    id            SERIAL PRIMARY KEY,
                    name          TEXT NOT NULL UNIQUE,
                    unit          TEXT NOT NULL,
                    unit_detail   TEXT,
                    unit_price    INTEGER DEFAULT 0,
                    min_threshold INTEGER DEFAULT 2
                );
                CREATE TABLE IF NOT EXISTS inventory (
                    id            SERIAL PRIMARY KEY,
                    branch_id     INTEGER NOT NULL,
                    form_type_id  INTEGER NOT NULL,
                    quantity      INTEGER DEFAULT 0,
                    last_updated  TIMESTAMP DEFAULT NOW(),
                    FOREIGN KEY (branch_id)    REFERENCES branches(id),
                    FOREIGN KEY (form_type_id) REFERENCES form_types(id),
                    UNIQUE(branch_id, form_type_id)
                );
                CREATE TABLE IF NOT EXISTS transactions (
                    id               SERIAL PRIMARY KEY,
                    type             TEXT NOT NULL,
                    form_type_id     INTEGER NOT NULL,
                    from_branch_id   INTEGER,
                    to_branch_id     INTEGER,
                    quantity         INTEGER NOT NULL,
                    notes            TEXT,
                    created_by       TEXT,
                    created_at       TIMESTAMP DEFAULT NOW(),
                    period_month     TEXT,
                    transaction_date DATE DEFAULT CURRENT_DATE,
                    FOREIGN KEY (form_type_id)   REFERENCES form_types(id),
                    FOREIGN KEY (from_branch_id) REFERENCES branches(id),
                    FOREIGN KEY (to_branch_id)   REFERENCES branches(id)
                );
                CREATE TABLE IF NOT EXISTS users (
                    id        SERIAL PRIMARY KEY,
                    username  TEXT NOT NULL UNIQUE,
                    password  TEXT NOT NULL,
                    branch_id INTEGER,
                    role      TEXT DEFAULT 'staff',
                    FOREIGN KEY (branch_id) REFERENCES branches(id)
                );
                CREATE TABLE IF NOT EXISTS transfer_requests (
                    id              SERIAL PRIMARY KEY,
                    from_branch_id  INTEGER NOT NULL,
                    to_branch_id    INTEGER NOT NULL,
                    form_type_id    INTEGER NOT NULL,
                    quantity        INTEGER NOT NULL,
                    notes           TEXT,
                    status          TEXT DEFAULT 'PENDING',
                    reject_reason   TEXT,
                    requested_by    TEXT NOT NULL,
                    approved_by     TEXT,
                    created_at      TIMESTAMP DEFAULT NOW(),
                    updated_at      TIMESTAMP DEFAULT NOW(),
                    FOREIGN KEY (from_branch_id) REFERENCES branches(id),
                    FOREIGN KEY (to_branch_id)   REFERENCES branches(id),
                    FOREIGN KEY (form_type_id)   REFERENCES form_types(id)
                );
                CREATE TABLE IF NOT EXISTS notifications (
                    id                  SERIAL PRIMARY KEY,
                    branch_id           INTEGER NOT NULL,
                    message             TEXT NOT NULL,
                    transfer_request_id INTEGER,
                    is_read             INTEGER DEFAULT 0,
                    created_at          TIMESTAMP DEFAULT NOW(),
                    FOREIGN KEY (branch_id) REFERENCES branches(id)
                );
                CREATE TABLE IF NOT EXISTS flight_schedule (
                    id           SERIAL PRIMARY KEY,
                    branch_id    INTEGER NOT NULL REFERENCES branches(id),
                    year_month   TEXT NOT NULL,
                    flight_count INTEGER NOT NULL DEFAULT 0,
                    updated_at   TIMESTAMP DEFAULT NOW(),
                    UNIQUE(branch_id, year_month)
                );
                CREATE TABLE IF NOT EXISTS catalog_branch_items (
                    id         SERIAL PRIMARY KEY,
                    branch_id  INTEGER NOT NULL REFERENCES branches(id),
                    item_code  TEXT NOT NULL,
                    quantity   INTEGER DEFAULT 1,
                    updated_at TIMESTAMP DEFAULT NOW(),
                    UNIQUE(branch_id, item_code)
                );
                CREATE TABLE IF NOT EXISTS catalog_defs (
                    code       TEXT PRIMARY KEY,
                    img        TEXT NOT NULL,
                    name       TEXT NOT NULL,
                    cat        TEXT NOT NULL,
                    sub_desc   TEXT NOT NULL DEFAULT '',
                    sort_order INTEGER NOT NULL DEFAULT 0
                );
                CREATE TABLE IF NOT EXISTS catalog_requests (
                    id            SERIAL PRIMARY KEY,
                    branch_id     INTEGER NOT NULL REFERENCES branches(id),
                    status        TEXT NOT NULL DEFAULT 'draft',
                    notes         TEXT DEFAULT '',
                    reject_reason TEXT DEFAULT '',
                    created_at    TIMESTAMP DEFAULT NOW(),
                    updated_at    TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS catalog_request_items (
                    id              SERIAL PRIMARY KEY,
                    request_id      INTEGER NOT NULL REFERENCES catalog_requests(id),
                    item_code       TEXT NOT NULL,
                    quantity        INTEGER DEFAULT 1,
                    custom_img_data TEXT DEFAULT '',
                    custom_text     TEXT DEFAULT '',
                    UNIQUE(request_id, item_code)
                );
                CREATE TABLE IF NOT EXISTS form_supply_settings (
                    id           SERIAL PRIMARY KEY,
                    period_start TEXT NOT NULL,
                    period_end   TEXT NOT NULL,
                    is_enabled   INTEGER DEFAULT 1,
                    created_by   TEXT NOT NULL,
                    updated_at   TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS form_supply_requests (
                    id            SERIAL PRIMARY KEY,
                    branch_id     INTEGER NOT NULL REFERENCES branches(id),
                    status        TEXT DEFAULT 'pending',
                    notes         TEXT DEFAULT '',
                    reject_reason TEXT DEFAULT '',
                    requested_by  TEXT NOT NULL,
                    processed_by  TEXT DEFAULT '',
                    processed_at  TIMESTAMP,
                    created_at    TIMESTAMP DEFAULT NOW(),
                    updated_at    TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS form_supply_request_items (
                    id           SERIAL PRIMARY KEY,
                    request_id   INTEGER NOT NULL REFERENCES form_supply_requests(id),
                    form_type_id INTEGER NOT NULL REFERENCES form_types(id),
                    quantity     INTEGER NOT NULL DEFAULT 1
                );
            ''')
            # 콜드스타트마다 실행: 멱등 컬럼 마이그레이션 (IF NOT EXISTS — 빠름)
            for _mig in [
                "ALTER TABLE transfer_requests ADD COLUMN IF NOT EXISTS notify_email TEXT DEFAULT ''",
                "ALTER TABLE branches ADD COLUMN IF NOT EXISTS email TEXT DEFAULT ''",
                "ALTER TABLE users ADD COLUMN IF NOT EXISTS failed_attempts INTEGER NOT NULL DEFAULT 0",
                "ALTER TABLE users ADD COLUMN IF NOT EXISTS locked_until TIMESTAMP",
                "ALTER TABLE users ADD COLUMN IF NOT EXISTS last_login_at TIMESTAMP",
                "ALTER TABLE users ADD COLUMN IF NOT EXISTS password_changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
            ]:
                try:
                    conn.execute(_mig)
                    conn.commit()
                except Exception:
                    pass

        already_seeded = conn.execute('SELECT COUNT(*) AS cnt FROM branches').fetchone()['cnt'] > 0

        if not already_seeded:
            branches = [
                ('GMP', '김포', 'DOM'), ('CJU', '제주', 'DOM'), ('CJJ', '청주', 'DOM'),
                ('PUS', '부산', 'DOM'), ('TSA', '송산', 'INTL'), ('KMJ', '구마모토', 'INTL'),
                ('서울역', '서울역', 'DOM'), ('광명역', '광명역', 'DOM'), ('이지드랍', '이지드랍', 'DOM'),
                ('ICN', '인천', 'DOM'), ('TPE', '타오위안', 'INTL'), ('NRT', '나리타', 'INTL'),
                ('KIX', '간사이', 'INTL'), ('FUK', '후쿠오카', 'INTL'), ('CTS', '삿포로', 'INTL'),
                ('OKA', '오키나와', 'INTL'), ('TKS', '도쿠시마', 'INTL'), ('BKK', '방콕', 'INTL'),
                ('CNX', '치앙마이', 'INTL'), ('DAD', '다낭', 'INTL'), ('CXR', '나트랑', 'INTL'),
                ('PQC', '푸꾸옥', 'INTL'), ('MDC', '마나도', 'INTL'), ('PVG', '상하이', 'INTL'),
                ('YNJ', '연길', 'INTL'), ('CGO', '정저우', 'INTL'), ('YNT', '옌타이', 'INTL'),
                ('HKG', '홍콩', 'INTL'), ('ALA', '알마티', 'INTL'),
                ('KOJ', '가고시마', 'INTL'), ('HGH', '항저우', 'INTL'), ('XMN', '샤먼', 'INTL'),
                ('CARGO', '화물파트', 'CARGO'),
            ]
            for code, name, btype in branches:
                conn.execute(
                    'INSERT INTO branches (code, name, type) VALUES (%s,%s,%s) ON CONFLICT(code) DO NOTHING',
                    (code, name, btype)
                )

            form_types = [
                ('DOM BOARDING PASS (롤)',                                         'BOX', '50롤',     60000, 3,  1),
                ('INTL BOARDING PASS(QR)',                                         'BOX', '5,000장', 175000, 3,  2),
                ('INTL BOARDING PASS(QR, ICN)',                                    'BOX', '5,000장', 175000, 2,  3),
                ('AUTO BAG TAG',                                                   'BOX', '10롤',    118000, 5,  4),
                ('BAG TIPS',                                                       'BOX', '5,000장',  50000, 3,  5),
                ('BAG TIPS (SNOOPY, DOM)',                                         'BOX', '5,000장',  57000, 2,  6),
                ('MANUAL BAG TAG',                                                 'BOX', '5,000장', 180000, 2,  7),
                ('Carry on Bag TAG (INTL)',                                        'BOX', '5,000장', 145000, 2,  8),
                ('SRI 봉투(大)',                                                   'BOX', '500장',   330000, 1,  9),
                ('CO-MAIL 봉투(NEW)',                                              'BOX', '200장',   160000, 1, 10),
                ('유상비닐(小/PPS)',                                               '포대', '100개',  110000, 2, 11),
                ('유상비닐(大/PPL)',                                               '포대', '100개',  130000, 2, 12),
                ('BOX TAPE',                                                       'BOX', '50개',     55000, 3, 13),
                ('PREMIUM TAG(D/S)',                                               'BOX', '5,000장',  80000, 2, 14),
                ('FRAGILE TAG(NEW)',                                               'BOX', '5,000장',  80000, 2, 15),
                ('HEAVY TAG',                                                      'BOX', '5,000장',  80000, 2, 16),
                ('GTOG TAG',                                                       'BOX', '5,000장',  80000, 2, 17),
                ('Exit-Seat Sticker',                                              'BOX', '20,000장', 110000, 1, 18),
                ('COB LABEL',                                                      'BOX', '5,000장',  80000, 1, 19),
                ('UP SIDE LABEL',                                                  'BOX', '5,000장',  80000, 1, 20),
                ('WCHR Battery LABEL',                                             'BOX', '5,000장',  80000, 1, 21),
                ('CORROSIVE LABEL',                                                'BOX', '5,000장',  80000, 1, 22),
                ('Dry Ice LABEL',                                                  'BOX', '5,000장',  80000, 1, 23),
                ('한국 입국신고서 (ENG/CNA)',                                      'BOX', '5,000장', 150000, 1, 24),
                ('제주 E/D카드',                                                   'BOX', '5,000장', 150000, 1, 25),
                ('한국 세관신고서 (ENG/CNA)',                                      'BOX', '5,000장', 150000, 1, 26),
                ('한국 세관신고서 (ENG/JPN)',                                      'BOX', '5,000장', 150000, 1, 27),
                ('서약서 (DECLARATION OF INDEMNITY)',                              '권',  '100조',    8500, 3, 28),
                ('합의서 (Release And Indemnity Letter)',                           '권',  '100조',    8500, 2, 29),
                ('반려동물 서약서 (DECLARATION OF INDEMNITY,PET)',                 '권',  '100조',   10000, 2, 30),
                ('악기 서약서 (DECLARATION OF INDEMNITY,Musical Instrument)',      '권',  '100조',   10000, 1, 31),
                ('보호자 서약서 (DECLARATION OF PARENT GUARDIAN)',                 '권',  '100조',    8500, 1, 32),
                ('총기인수인계서 (Firearm handover form)',                          '권',  '100조',   10000, 1, 33),
                ('PIR',                                                            '권',  '100조',   10000, 2, 34),
                ('SHR',                                                            '권',  '100조',   14000, 1, 35),
                ('NOTOC',                                                          '권',  '100조',   10000, 1, 36),
                ('BAG(BINGO) CHART (양면)',                                        '권',  '100조',    3200, 5, 37),
                # 비활성 항목 (데이터 보존용)
                ('TRANSFER TAG',                                                   'BOX', '5,000장',  80000, 1, 999),
                ('AOC LABEL',                                                      'BOX', '5,000장',  80000, 1, 999),
                ('POB LABEL',                                                      'BOX', '5,000장',  80000, 1, 999),
            ]
            for name, unit, ud, price, thr, sort_order in form_types:
                conn.execute(
                    'INSERT INTO form_types (name, unit, unit_detail, unit_price, min_threshold, sort_order) '
                    'VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT(name) DO UPDATE SET sort_order=EXCLUDED.sort_order',
                    (name, unit, ud, price, thr, sort_order)
                )
        else:
            conn.execute("UPDATE branches SET type='DOM' WHERE code='ICN' AND type != 'DOM'")
            conn.execute("UPDATE branches SET name='타오위안' WHERE code='TPE' AND name='타이페이'")

        # 지점 이메일 시드 — email이 비어있는 지점에만 적용 (수동 설정 보호)
        ph = '%s' if not USE_SQLITE else '?'
        for _code, _email in _BRANCH_EMAIL_SEEDS.items():
            conn.execute(
                f"UPDATE branches SET email={ph} WHERE code={ph} AND (email IS NULL OR email='')",
                (_email, _code)
            )

        user_count = conn.execute('SELECT COUNT(*) AS cnt FROM users').fetchone()['cnt']
        if user_count == 0:
            admin_pw = hash_pw('admin1234')
            conn.execute(
                'INSERT INTO users (username, password, branch_id, role) VALUES (%s,%s,NULL,%s)',
                ('admin', admin_pw, 'admin')
            )
            staff_pw = hash_pw('staff1234')
            conn.execute(
                "INSERT INTO users (username, password, branch_id, role) "
                "SELECT %s, %s, id, 'staff' FROM branches WHERE code='GMP'",
                ('gmp', staff_pw)
            )
            conn.execute(
                "INSERT INTO users (username, password, branch_id, role) "
                "SELECT %s, %s, id, 'staff' FROM branches WHERE code='ICN'",
                ('icn', staff_pw)
            )
        conn.commit()

        # catalog_defs 시딩 (SQLite — DO UPDATE로 코드 변경사항 반영)
        if USE_SQLITE:
            try:
                _seed_catalog_defs(conn)
            except Exception as _se:
                print(f'[init_db/seed/sqlite] {_se}')
                try:
                    conn.rollback()
                except Exception:
                    pass

        if USE_SQLITE:
            cols = [r[1] for r in conn.execute('PRAGMA table_info(transactions)').fetchall()]
            if 'period_month' not in cols:
                conn.execute('ALTER TABLE transactions ADD COLUMN period_month TEXT')
                conn.execute(
                    "UPDATE transactions SET period_month = strftime('%Y-%m', created_at) WHERE type='OUT'"
                )
            if 'transaction_date' not in cols:
                conn.execute('ALTER TABLE transactions ADD COLUMN transaction_date TEXT')
                conn.execute("UPDATE transactions SET transaction_date = date(created_at)")
            cat_cols = [r[1] for r in conn.execute('PRAGMA table_info(catalog_defs)').fetchall()]
            if 'img_data' not in cat_cols:
                conn.execute("ALTER TABLE catalog_defs ADD COLUMN img_data TEXT NOT NULL DEFAULT ''")
            if 'user_deleted' not in cat_cols:
                conn.execute("ALTER TABLE catalog_defs ADD COLUMN user_deleted INTEGER NOT NULL DEFAULT 0")
            fsr_cols = [r[1] for r in conn.execute('PRAGMA table_info(form_supply_requests)').fetchall()]
            if 'approve_reason' not in fsr_cols:
                conn.execute("ALTER TABLE form_supply_requests ADD COLUMN approve_reason TEXT NOT NULL DEFAULT ''")
            if 'period_title' not in fsr_cols:
                conn.execute("ALTER TABLE form_supply_requests ADD COLUMN period_title TEXT NOT NULL DEFAULT ''")
            fss_cols = [r[1] for r in conn.execute('PRAGMA table_info(form_supply_settings)').fetchall()]
            if 'title' not in fss_cols:
                conn.execute("ALTER TABLE form_supply_settings ADD COLUMN title TEXT NOT NULL DEFAULT ''")
            ft_cols = [r[1] for r in conn.execute('PRAGMA table_info(form_types)').fetchall()]
            if 'sort_order' not in ft_cols:
                conn.execute("ALTER TABLE form_types ADD COLUMN sort_order INTEGER NOT NULL DEFAULT 999")
            if 'is_active' not in ft_cols:
                conn.execute("ALTER TABLE form_types ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
            if 'memo' not in ft_cols:
                conn.execute("ALTER TABLE form_types ADD COLUMN memo TEXT NOT NULL DEFAULT ''")
            # 이름 변경 마이그레이션
            _form_renames = [
                ('비상구열 스티커',          'Exit-Seat Sticker'),
                ('휠체어 배터리 분리 L/B',   'WCHR Battery LABEL'),
                ('서약서',                   '서약서 (DECLARATION OF INDEMNITY)'),
                ('합의서',                   '합의서 (Release And Indemnity Letter)'),
                ('반려동물 서약서',           '반려동물 서약서 (DECLARATION OF INDEMNITY,PET)'),
                ('악기서약서',               '악기 서약서 (DECLARATION OF INDEMNITY,Musical Instrument)'),
                ('보호자 서약서',             '보호자 서약서 (DECLARATION OF PARENT GUARDIAN)'),
                ('총기인수인계서',             '총기인수인계서 (Firearm handover form)'),
                ('BAG BINGO CHART(양면)',     'BAG(BINGO) CHART (양면)'),
            ]
            for old, new in _form_renames:
                conn.execute('UPDATE form_types SET name=? WHERE name=?', (new, old))
            # sort_order 설정
            _form_orders = [
                ('DOM BOARDING PASS (롤)', 1), ('INTL BOARDING PASS(QR)', 2),
                ('INTL BOARDING PASS(QR, ICN)', 3), ('AUTO BAG TAG', 4),
                ('BAG TIPS', 5), ('BAG TIPS (SNOOPY, DOM)', 6),
                ('MANUAL BAG TAG', 7), ('Carry on Bag TAG (INTL)', 8),
                ('SRI 봉투(大)', 9), ('CO-MAIL 봉투(NEW)', 10),
                ('유상비닐(小/PPS)', 11), ('유상비닐(大/PPL)', 12),
                ('BOX TAPE', 13), ('PREMIUM TAG(D/S)', 14),
                ('FRAGILE TAG(NEW)', 15), ('HEAVY TAG', 16),
                ('GTOG TAG', 17), ('TRANSFER TAG', 18),
                ('Exit-Seat Sticker', 19),
                ('AOC LABEL', 20), ('POB LABEL', 21),
                ('COB LABEL', 22), ('UP SIDE LABEL', 23),
                ('WCHR Battery LABEL', 24), ('CORROSIVE LABEL', 25),
                ('Dry Ice LABEL', 26), ('한국 입국신고서 (ENG/CNA)', 27),
                ('제주 E/D카드', 28), ('한국 세관신고서 (ENG/CNA)', 29),
                ('한국 세관신고서 (ENG/JPN)', 30),
                ('서약서 (DECLARATION OF INDEMNITY)', 31),
                ('합의서 (Release And Indemnity Letter)', 32),
                ('반려동물 서약서 (DECLARATION OF INDEMNITY,PET)', 33),
                ('악기 서약서 (DECLARATION OF INDEMNITY,Musical Instrument)', 34),
                ('보호자 서약서 (DECLARATION OF PARENT GUARDIAN)', 35),
                ('총기인수인계서 (Firearm handover form)', 36),
                ('PIR', 37), ('SHR', 38), ('NOTOC', 39),
                ('BAG(BINGO) CHART (양면)', 40),
            ]
            for name, order in _form_orders:
                conn.execute('UPDATE form_types SET sort_order=? WHERE name=?', (order, name))
            for _reactivate in ('TRANSFER TAG', 'AOC LABEL', 'POB LABEL'):
                conn.execute('UPDATE form_types SET is_active=1 WHERE name=?', (_reactivate,))
        else:
            conn.execute('''
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='transactions' AND column_name='period_month'
                    ) THEN
                        ALTER TABLE transactions ADD COLUMN period_month TEXT;
                        UPDATE transactions
                           SET period_month = TO_CHAR(created_at, 'YYYY-MM')
                         WHERE type = 'OUT';
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='transactions' AND column_name='transaction_date'
                    ) THEN
                        ALTER TABLE transactions ADD COLUMN transaction_date DATE DEFAULT CURRENT_DATE;
                        UPDATE transactions SET transaction_date = created_at::date;
                    END IF;
                    CREATE INDEX IF NOT EXISTS idx_inventory_branch    ON inventory(branch_id);
                    CREATE INDEX IF NOT EXISTS idx_inventory_form      ON inventory(form_type_id);
                    CREATE INDEX IF NOT EXISTS idx_tx_created_at       ON transactions(created_at);
                    CREATE INDEX IF NOT EXISTS idx_tx_from_branch      ON transactions(from_branch_id);
                    CREATE INDEX IF NOT EXISTS idx_tx_to_branch        ON transactions(to_branch_id);
                    CREATE INDEX IF NOT EXISTS idx_tx_period_month     ON transactions(period_month);
                    CREATE INDEX IF NOT EXISTS idx_tx_transaction_date ON transactions(transaction_date);
                    -- flight_schedule
                    CREATE TABLE IF NOT EXISTS flight_schedule (
                        id           SERIAL PRIMARY KEY,
                        branch_id    INTEGER NOT NULL REFERENCES branches(id),
                        year_month   TEXT NOT NULL,
                        flight_count INTEGER NOT NULL DEFAULT 0,
                        updated_at   TIMESTAMP DEFAULT NOW(),
                        UNIQUE(branch_id, year_month)
                    );
                    -- catalog_branch_items
                    CREATE TABLE IF NOT EXISTS catalog_branch_items (
                        id         SERIAL PRIMARY KEY,
                        branch_id  INTEGER NOT NULL REFERENCES branches(id),
                        item_code  TEXT NOT NULL,
                        quantity   INTEGER DEFAULT 1,
                        updated_at TIMESTAMP DEFAULT NOW(),
                        UNIQUE(branch_id, item_code)
                    );
                    CREATE TABLE IF NOT EXISTS catalog_defs (
                        code       TEXT PRIMARY KEY,
                        img        TEXT NOT NULL,
                        name       TEXT NOT NULL,
                        cat        TEXT NOT NULL,
                        sub_desc   TEXT NOT NULL DEFAULT '',
                        sort_order INTEGER NOT NULL DEFAULT 0,
                        img_data   TEXT NOT NULL DEFAULT ''
                    );
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='catalog_defs' AND column_name='img_data'
                    ) THEN
                        ALTER TABLE catalog_defs ADD COLUMN img_data TEXT NOT NULL DEFAULT '';
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='catalog_defs' AND column_name='user_deleted'
                    ) THEN
                        ALTER TABLE catalog_defs ADD COLUMN user_deleted BOOLEAN NOT NULL DEFAULT FALSE;
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='form_types' AND column_name='sort_order'
                    ) THEN
                        ALTER TABLE form_types ADD COLUMN sort_order INTEGER NOT NULL DEFAULT 999;
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='form_types' AND column_name='is_active'
                    ) THEN
                        ALTER TABLE form_types ADD COLUMN is_active BOOLEAN NOT NULL DEFAULT TRUE;
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='form_types' AND column_name='memo'
                    ) THEN
                        ALTER TABLE form_types ADD COLUMN memo TEXT NOT NULL DEFAULT '';
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='form_supply_requests' AND column_name='approve_reason'
                    ) THEN
                        ALTER TABLE form_supply_requests ADD COLUMN approve_reason TEXT NOT NULL DEFAULT '';
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='form_supply_requests' AND column_name='period_title'
                    ) THEN
                        ALTER TABLE form_supply_requests ADD COLUMN period_title TEXT NOT NULL DEFAULT '';
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='form_supply_settings' AND column_name='title'
                    ) THEN
                        ALTER TABLE form_supply_settings ADD COLUMN title TEXT NOT NULL DEFAULT '';
                    END IF;
                END $$
            ''')
            # DO $$ 블록 외부에서 개별 실행 — IF NOT EXISTS 지원으로 멱등성 보장
            conn.execute("ALTER TABLE form_supply_settings ADD COLUMN IF NOT EXISTS title TEXT NOT NULL DEFAULT ''")
            conn.execute("ALTER TABLE form_supply_requests ADD COLUMN IF NOT EXISTS period_title TEXT NOT NULL DEFAULT ''")
            conn.execute("ALTER TABLE form_supply_requests ADD COLUMN IF NOT EXISTS approve_reason TEXT NOT NULL DEFAULT ''")
            # form_types 이름 변경 + sort_order + 비활성 마이그레이션 (PostgreSQL)
            _form_renames = [
                ('비상구열 스티커',          'Exit-Seat Sticker'),
                ('휠체어 배터리 분리 L/B',   'WCHR Battery LABEL'),
                ('서약서',                   '서약서 (DECLARATION OF INDEMNITY)'),
                ('합의서',                   '합의서 (Release And Indemnity Letter)'),
                ('반려동물 서약서',           '반려동물 서약서 (DECLARATION OF INDEMNITY,PET)'),
                ('악기서약서',               '악기 서약서 (DECLARATION OF INDEMNITY,Musical Instrument)'),
                ('보호자 서약서',             '보호자 서약서 (DECLARATION OF PARENT GUARDIAN)'),
                ('총기인수인계서',             '총기인수인계서 (Firearm handover form)'),
                ('BAG BINGO CHART(양면)',     'BAG(BINGO) CHART (양면)'),
            ]
            for old, new in _form_renames:
                conn.execute('UPDATE form_types SET name=%s WHERE name=%s', (new, old))
            _form_orders = [
                ('DOM BOARDING PASS (롤)', 1), ('INTL BOARDING PASS(QR)', 2),
                ('INTL BOARDING PASS(QR, ICN)', 3), ('AUTO BAG TAG', 4),
                ('BAG TIPS', 5), ('BAG TIPS (SNOOPY, DOM)', 6),
                ('MANUAL BAG TAG', 7), ('Carry on Bag TAG (INTL)', 8),
                ('SRI 봉투(大)', 9), ('CO-MAIL 봉투(NEW)', 10),
                ('유상비닐(小/PPS)', 11), ('유상비닐(大/PPL)', 12),
                ('BOX TAPE', 13), ('PREMIUM TAG(D/S)', 14),
                ('FRAGILE TAG(NEW)', 15), ('HEAVY TAG', 16),
                ('GTOG TAG', 17), ('TRANSFER TAG', 18),
                ('Exit-Seat Sticker', 19),
                ('AOC LABEL', 20), ('POB LABEL', 21),
                ('COB LABEL', 22), ('UP SIDE LABEL', 23),
                ('WCHR Battery LABEL', 24), ('CORROSIVE LABEL', 25),
                ('Dry Ice LABEL', 26), ('한국 입국신고서 (ENG/CNA)', 27),
                ('제주 E/D카드', 28), ('한국 세관신고서 (ENG/CNA)', 29),
                ('한국 세관신고서 (ENG/JPN)', 30),
                ('서약서 (DECLARATION OF INDEMNITY)', 31),
                ('합의서 (Release And Indemnity Letter)', 32),
                ('반려동물 서약서 (DECLARATION OF INDEMNITY,PET)', 33),
                ('악기 서약서 (DECLARATION OF INDEMNITY,Musical Instrument)', 34),
                ('보호자 서약서 (DECLARATION OF PARENT GUARDIAN)', 35),
                ('총기인수인계서 (Firearm handover form)', 36),
                ('PIR', 37), ('SHR', 38), ('NOTOC', 39),
                ('BAG(BINGO) CHART (양면)', 40),
            ]
            for name, order in _form_orders:
                conn.execute('UPDATE form_types SET sort_order=%s WHERE name=%s', (order, name))
            for _reactivate in ('TRANSFER TAG', 'AOC LABEL', 'POB LABEL'):
                conn.execute('UPDATE form_types SET is_active=TRUE WHERE name=%s', (_reactivate,))
            # 신규 지점 추가 (이미 있으면 무시)
            for _bc, _bn, _bt in [('KOJ', '가고시마', 'INTL'), ('HGH', '항저우', 'INTL'), ('XMN', '샤먼', 'INTL')]:
                conn.execute(
                    'INSERT INTO branches (code, name, type) VALUES (%s,%s,%s) ON CONFLICT(code) DO NOTHING',
                    (_bc, _bn, _bt)
                )
            # KOR → KOJ 코드 오류 수정
            conn.execute("UPDATE branches SET code='KOJ' WHERE code='KOR' AND name='가고시마'")
        conn.commit()

        # catalog_defs 시딩 (DO UPDATE — 코드 변경사항 배포 시 자동 반영)
        # 별도 try/except: 시딩 실패가 전체 init_db를 막지 않도록
        try:
            _seed_catalog_defs(conn)
        except Exception as _se:
            print(f'[init_db/seed] {_se}')
            try:
                conn.rollback()
            except Exception:
                pass

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    _DB_INITIALIZED = True


# ── i18n ──────────────────────────────────────────────────────────────────────

@app.context_processor
def inject_i18n():
    lang = session.get('lang', 'ko')
    return {'T': make_T(lang), 'cur_lang': lang, 'LANG_LABELS': LANG_LABELS, 'SUPPORTED_LANGS': SUPPORTED}

@app.route('/lang/<code>')
def set_lang(code):
    if code in SUPPORTED:
        session['lang'] = code
    return redirect(request.referrer or url_for('dashboard'))


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/ping')
def ping():
    """Vercel Cron이 5분마다 호출 — cold start 방지용."""
    return 'ok', 200


@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        ip = _client_ip()
        conn = get_db()

        row = conn.execute(
            'SELECT u.*, b.name branch_name FROM users u LEFT JOIN branches b ON u.branch_id=b.id '
            'WHERE u.username=%s',
            (username,)
        ).fetchone()
        user = dict(row) if row else None

        # 잠금 확인
        if user and user.get('locked_until'):
            locked_until = _to_dt(user['locked_until'])
            if locked_until and datetime.now(timezone.utc) < locked_until:
                remaining = int((locked_until - datetime.now(timezone.utc)).total_seconds() / 60) + 1
                flash(f'로그인 5회 실패로 계정이 잠겼습니다. {remaining}분 후 다시 시도하세요.', 'danger')
                conn.close()
                return render_template('login.html')

        if user and verify_pw(password, user['password']):
            # 레거시 SHA256 → PBKDF2 자동 재해시
            if not user['password'].startswith('pbkdf2:') and not user['password'].startswith('scrypt:'):
                conn.execute('UPDATE users SET password=%s WHERE id=%s', (hash_pw(password), user['id']))

            conn.execute(
                'UPDATE users SET failed_attempts=0, locked_until=NULL, last_login_at=NOW() WHERE id=%s',
                (user['id'],)
            )
            conn.commit()

            # 접속 로그
            conn.execute(
                'INSERT INTO access_logs (user_id, username, action, ip_address, created_at) '
                'VALUES (%s, %s, %s, %s, NOW())',
                (user['id'], user['username'], '로그인', ip)
            )
            conn.commit()
            conn.close()

            # 세션 고정 방어: 기존 세션 데이터 완전 초기화 후 재발급
            now_ts = datetime.now(timezone.utc).timestamp()
            pwd_changed_at = _to_dt(user.get('password_changed_at'))
            pwd_changed_ts = pwd_changed_at.timestamp() if pwd_changed_at else now_ts

            session.clear()
            session.permanent = True
            session['_last_activity'] = now_ts
            session['_pwd_changed_at'] = pwd_changed_ts
            session.update(
                user_id=user['id'], username=user['username'],
                role=user['role'], branch_id=user['branch_id'],
                branch_name=user['branch_name']
            )
            return redirect(url_for('dashboard'))

        # 로그인 실패
        if user:
            new_attempts = (user.get('failed_attempts') or 0) + 1
            if new_attempts >= 5:
                conn.execute(
                    "UPDATE users SET failed_attempts=%s, locked_until=NOW() + INTERVAL '10 minutes' WHERE id=%s",
                    (new_attempts, user['id'])
                )
                flash('로그인 5회 실패로 계정이 10분간 잠겼습니다.', 'danger')
            else:
                conn.execute(
                    'UPDATE users SET failed_attempts=%s WHERE id=%s',
                    (new_attempts, user['id'])
                )
                flash(f'아이디 또는 비밀번호가 올바르지 않습니다. ({new_attempts}/5회)', 'danger')
            conn.execute(
                'INSERT INTO access_logs (user_id, username, action, ip_address, created_at) '
                'VALUES (%s, %s, %s, %s, NOW())',
                (user['id'], username, f'로그인실패({new_attempts}회)', ip)
            )
            conn.commit()
        else:
            flash('아이디 또는 비밀번호가 올바르지 않습니다.', 'danger')

        conn.close()
    return render_template('login.html')


@app.route('/logout')
def logout():
    log_action('로그아웃')
    session.clear()
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    bid = session.get('branch_id')
    role = session.get('role')

    branch_cond = f'AND i.branch_id = {bid}' if role != 'admin' and bid else ''

    low_stock = conn.execute(f'''
        SELECT i.quantity, f.name form_name, f.min_threshold,
               b.name branch_name, b.code branch_code
        FROM inventory i
        JOIN form_types f ON i.form_type_id = f.id
        JOIN branches b ON i.branch_id = b.id
        WHERE i.quantity > 0 AND i.quantity <= f.min_threshold {branch_cond}
        ORDER BY i.quantity ASC LIMIT 20
    ''').fetchall()

    empty_stock = conn.execute(f'''
        SELECT f.name form_name, b.name branch_name, b.code branch_code
        FROM inventory i
        JOIN form_types f ON i.form_type_id = f.id
        JOIN branches b ON i.branch_id = b.id
        WHERE i.quantity = 0 {branch_cond}
        ORDER BY b.code, f.name LIMIT 20
    ''').fetchall()

    tx_cond = f'WHERE (t.from_branch_id={bid} OR t.to_branch_id={bid})' if role != 'admin' and bid else ''
    recent_tx = conn.execute(f'''
        SELECT t.*, f.name form_name,
               fb.name from_branch, tb.name to_branch
        FROM transactions t
        JOIN form_types f ON t.form_type_id = f.id
        LEFT JOIN branches fb ON t.from_branch_id = fb.id
        LEFT JOIN branches tb ON t.to_branch_id = tb.id
        {tx_cond}
        ORDER BY t.created_at DESC LIMIT 10
    ''').fetchall()

    _s = conn.execute('''
        SELECT
            (SELECT COUNT(*) FROM branches)   AS branch_cnt,
            (SELECT COUNT(*) FROM form_types) AS form_cnt,
            (SELECT COUNT(*) FROM transactions WHERE transaction_date = CURRENT_DATE) AS today_cnt
    ''').fetchone()
    stats = {
        'branches': _s['branch_cnt'],
        'forms':    _s['form_cnt'],
        'today_tx': _s['today_cnt'],
        'alerts':   len(low_stock) + len(empty_stock),
    }
    # 6개월 출고 트렌드 (차트용)
    import json as _json
    chart_data = {'labels': [], 'datasets': []}
    try:
        months = conn.execute('''
            SELECT TO_CHAR(CURRENT_DATE - INTERVAL '5 months', 'YYYY-MM') AS m0,
                   TO_CHAR(CURRENT_DATE - INTERVAL '4 months', 'YYYY-MM') AS m1,
                   TO_CHAR(CURRENT_DATE - INTERVAL '3 months', 'YYYY-MM') AS m2,
                   TO_CHAR(CURRENT_DATE - INTERVAL '2 months', 'YYYY-MM') AS m3,
                   TO_CHAR(CURRENT_DATE - INTERVAL '1 months', 'YYYY-MM') AS m4,
                   TO_CHAR(CURRENT_DATE, 'YYYY-MM') AS m5
        ''').fetchone()
        labels = [months[f'm{i}'] for i in range(6)]
        chart_cond = f'AND from_branch_id={bid}' if role != 'admin' and bid else ''
        rows_chart = conn.execute(f'''
            SELECT TO_CHAR(transaction_date, 'YYYY-MM') AS ym,
                   SUM(quantity) AS total
            FROM transactions
            WHERE type='OUT' AND transaction_date >= CURRENT_DATE - INTERVAL '5 months' {chart_cond}
            GROUP BY ym ORDER BY ym
        ''').fetchall()
        monthly_map = {r['ym']: int(r['total']) for r in rows_chart}
        data_vals = [monthly_map.get(lb, 0) for lb in labels]
        chart_data = {'labels': labels, 'datasets': [{'label': '전체 출고량', 'data': data_vals}]}
    except Exception:
        pass
    conn.close()
    return render_template('dashboard.html', low_stock=low_stock, empty_stock=empty_stock,
                           recent_tx=recent_tx, stats=stats,
                           chart_data=_json.dumps(chart_data))


@app.route('/inventory')
@login_required
def inventory():
    conn = get_db()
    branches = conn.execute('SELECT * FROM branches ORDER BY type, code').fetchall()
    form_types = conn.execute('SELECT * FROM form_types WHERE is_active ORDER BY sort_order').fetchall()

    bf = request.args.get('branch_id', '')
    ff = request.args.get('form_type_id', '')

    conditions, params = ['1=1'], []
    if session.get('role') != 'admin' and session.get('branch_id'):
        conditions.append('i.branch_id=%s'); params.append(session['branch_id'])
    elif bf:
        conditions.append('i.branch_id=%s'); params.append(bf)
    if ff:
        conditions.append('i.form_type_id=%s'); params.append(ff)

    where_clause = " AND ".join(conditions)

    rows = conn.execute(f'''
        SELECT i.*, f.name form_name, f.unit, f.unit_detail, f.unit_price, f.min_threshold,
               b.name branch_name, b.code branch_code, b.type branch_type
        FROM inventory i
        JOIN form_types f ON i.form_type_id = f.id
        JOIN branches b ON i.branch_id = b.id
        WHERE {where_clause}
        ORDER BY b.type, b.code, f.name
    ''', params).fetchall()

    summary_rows = conn.execute(f'''
        SELECT f.id form_type_id,
               f.name form_name, f.unit, f.unit_detail, f.unit_price, f.min_threshold,
               SUM(i.quantity)              total_qty,
               COUNT(DISTINCT i.branch_id) branch_cnt,
               SUM(CASE WHEN i.quantity = 0 THEN 1 ELSE 0 END)               empty_cnt,
               SUM(CASE WHEN i.quantity > 0 AND i.quantity <= f.min_threshold
                        THEN 1 ELSE 0 END)                                    low_cnt,
               SUM(i.quantity * f.unit_price)                                 total_value
        FROM inventory i
        JOIN form_types f ON i.form_type_id = f.id
        JOIN branches b ON i.branch_id = b.id
        WHERE {where_clause}
        GROUP BY f.id
        ORDER BY f.name
    ''', params).fetchall()

    grand_total_qty   = sum(r['total_qty']   or 0 for r in summary_rows)
    grand_total_value = sum(r['total_value'] or 0 for r in summary_rows)

    from datetime import date
    from collections import defaultdict
    today = date.today()
    month_labels = []
    y, m = today.year, today.month
    for _ in range(6):
        month_labels.insert(0, f'{y}-{m:02d}')
        m -= 1
        if m == 0: m, y = 12, y - 1

    if USE_SQLITE:
        _ph = ','.join(['?' for _ in month_labels])
        mo_conditions = [
            "t.type = 'OUT'",
            f"COALESCE(t.period_month, strftime('%Y-%m', t.created_at)) IN ({_ph})"
        ]
        mo_params = list(month_labels)
    else:
        mo_conditions = [
            "t.type = 'OUT'",
            f"COALESCE(t.period_month, TO_CHAR(t.created_at, 'YYYY-MM')) = ANY(%s)"
        ]
        mo_params = [month_labels]
    if session.get('role') != 'admin' and session.get('branch_id'):
        mo_conditions.append('t.from_branch_id = %s')
        mo_params.append(session['branch_id'])
    elif bf:
        mo_conditions.append('t.from_branch_id = %s')
        mo_params.append(bf)
    if ff:
        mo_conditions.append('t.form_type_id = %s')
        mo_params.append(ff)

    monthly_raw = conn.execute(f'''
        SELECT f.id form_type_id, f.name form_name, f.unit,
               COALESCE(t.period_month, TO_CHAR(t.created_at, 'YYYY-MM')) mo,
               SUM(t.quantity) qty
        FROM transactions t
        JOIN form_types f ON t.form_type_id = f.id
        WHERE {" AND ".join(mo_conditions)}
        GROUP BY f.id, COALESCE(t.period_month, TO_CHAR(t.created_at, 'YYYY-MM'))
        ORDER BY f.name, mo
    ''', mo_params).fetchall()

    monthly_pivot = defaultdict(lambda: defaultdict(int))
    monthly_meta = {}
    for r in monthly_raw:
        monthly_pivot[r['form_type_id']][r['mo']] += r['qty']
        monthly_meta[r['form_type_id']] = (r['form_name'], r['unit'])

    monthly_rows = []
    for fid_key in sorted(monthly_pivot, key=lambda x: monthly_meta[x][0]):
        fname, unit = monthly_meta[fid_key]
        vals = [monthly_pivot[fid_key].get(mo, 0) for mo in month_labels]
        monthly_rows.append({'form_name': fname, 'unit': unit,
                             'vals': vals, 'total': sum(vals)})
    monthly_totals = [sum(r['vals'][i] for r in monthly_rows) for i in range(6)]

    conn.close()
    return render_template('inventory.html',
                           rows=rows, summary_rows=summary_rows,
                           grand_total_qty=grand_total_qty,
                           grand_total_value=grand_total_value,
                           monthly_rows=monthly_rows,
                           monthly_totals=monthly_totals,
                           month_labels=month_labels,
                           branches=branches, form_types=form_types,
                           bf=bf, ff=ff)


@app.route('/inbound', methods=['GET', 'POST'])
@login_required
def inbound():
    conn = get_db()
    if request.method == 'POST':
        if session.get('role') != 'admin':
            if not session.get('branch_id'):
                flash('소속 지점이 없습니다. 관리자에게 문의하세요.', 'danger')
                conn.close()
                return redirect(url_for('dashboard'))
            bid = str(session['branch_id'])
        else:
            bid = request.form['branch_id']
        fid = request.form['form_type_id']
        qty = int(request.form['quantity'])
        notes = request.form.get('notes', '')
        from datetime import date as _date
        tx_date = request.form.get('transaction_date') or _date.today().isoformat()

        conn.execute('''
            INSERT INTO inventory (branch_id, form_type_id, quantity, last_updated)
            VALUES (%s,%s,%s,NOW())
            ON CONFLICT(branch_id, form_type_id) DO UPDATE SET
              quantity = inventory.quantity + EXCLUDED.quantity,
              last_updated = NOW()
        ''', (bid, fid, qty))
        conn.execute(
            "INSERT INTO transactions "
            "(type, form_type_id, to_branch_id, quantity, notes, created_by, transaction_date) "
            "VALUES ('IN',%s,%s,%s,%s,%s,%s)",
            (fid, bid, qty, notes, session['username'], tx_date)
        )
        conn.commit()

        b = conn.execute('SELECT name FROM branches WHERE id=%s', (bid,)).fetchone()
        f = conn.execute('SELECT name, min_threshold FROM form_types WHERE id=%s', (fid,)).fetchone()
        # 입고 후에도 여전히 부족한 경우 알림 (기존 미읽음 없을 때만)
        inv_after = conn.execute(
            'SELECT quantity FROM inventory WHERE branch_id=%s AND form_type_id=%s', (bid, fid)
        ).fetchone()
        if inv_after and f and inv_after['quantity'] <= f['min_threshold']:
            dup_notif = conn.execute(
                "SELECT id FROM notifications WHERE branch_id=%s AND is_read=0 "
                "AND message LIKE %s LIMIT 1",
                (bid, f'%{f["name"]}%')
            ).fetchone()
            if not dup_notif:
                status_label = '소진' if inv_after['quantity'] == 0 else '부족'
                conn.execute(
                    "INSERT INTO notifications (branch_id, message) VALUES (%s,%s)",
                    (bid, f'[재고{status_label}] {b["name"]} — {f["name"]} 잔여 {inv_after["quantity"]}개')
                )
                conn.commit()
        flash(f'입고 완료 ✔ {b["name"]} — {f["name"]} {qty}개 ({tx_date})', 'success')
        conn.close()
        return redirect(url_for('inbound'))

    branches = conn.execute('SELECT * FROM branches ORDER BY type, code').fetchall()
    form_types = conn.execute('SELECT * FROM form_types WHERE is_active ORDER BY sort_order').fetchall()
    conn.close()
    from datetime import date
    return render_template('inbound.html', branches=branches, form_types=form_types,
                           selected_branch=session.get('branch_id'),
                           today=date.today().isoformat())


@app.route('/outbound', methods=['GET', 'POST'])
@login_required
def outbound():
    conn = get_db()
    if request.method == 'POST':
        if session.get('role') != 'admin':
            if not session.get('branch_id'):
                flash('소속 지점이 없습니다. 관리자에게 문의하세요.', 'danger')
                conn.close()
                return redirect(url_for('dashboard'))
            bid = str(session['branch_id'])
        else:
            bid = request.form['branch_id']
        fid = request.form['form_type_id']
        qty = int(request.form['quantity'])
        notes = request.form.get('notes', '')
        from datetime import date as _date
        tx_date = request.form.get('transaction_date') or _date.today().isoformat()
        period_month = tx_date[:7]  # YYYY-MM

        cur = conn.execute(
            'SELECT quantity FROM inventory WHERE branch_id=%s AND form_type_id=%s',
            (bid, fid)
        ).fetchone()
        if not cur or cur['quantity'] < qty:
            flash('재고가 부족합니다.', 'danger')
            conn.close()
            return redirect(url_for('outbound'))

        conn.execute(
            'UPDATE inventory SET quantity=quantity-%s, last_updated=NOW() WHERE branch_id=%s AND form_type_id=%s',
            (qty, bid, fid)
        )
        conn.execute(
            "INSERT INTO transactions "
            "(type, form_type_id, from_branch_id, quantity, notes, created_by, period_month, transaction_date) "
            "VALUES ('OUT',%s,%s,%s,%s,%s,%s,%s)",
            (fid, bid, qty, notes, session['username'], period_month, tx_date)
        )
        conn.commit()
        # 출고 후 재고 부족/소진 즉시 알림
        inv_after = conn.execute(
            'SELECT i.quantity, f.min_threshold FROM inventory i '
            'JOIN form_types f ON i.form_type_id=f.id '
            'WHERE i.branch_id=%s AND i.form_type_id=%s', (bid, fid)
        ).fetchone()
        if inv_after and inv_after['quantity'] <= inv_after['min_threshold']:
            b_name = conn.execute('SELECT name FROM branches WHERE id=%s', (bid,)).fetchone()
            f_name = conn.execute('SELECT name FROM form_types WHERE id=%s', (fid,)).fetchone()
            if b_name and f_name:
                dup_notif = conn.execute(
                    "SELECT id FROM notifications WHERE branch_id=%s AND is_read=0 "
                    "AND message LIKE %s LIMIT 1",
                    (bid, f'%{f_name["name"]}%')
                ).fetchone()
                if not dup_notif:
                    status_label = '소진' if inv_after['quantity'] == 0 else '부족'
                    conn.execute(
                        "INSERT INTO notifications (branch_id, message) VALUES (%s,%s)",
                        (bid, f'[재고{status_label}] {b_name["name"]} — {f_name["name"]} 잔여 {inv_after["quantity"]}개')
                    )
                    conn.commit()
        flash(f'출고 처리 완료 ✔ ({tx_date})', 'success')
        conn.close()
        return redirect(url_for('outbound'))

    branches = conn.execute('SELECT * FROM branches ORDER BY type, code').fetchall()
    form_types = conn.execute('SELECT * FROM form_types WHERE is_active ORDER BY sort_order').fetchall()
    conn.close()
    from datetime import date
    return render_template('outbound.html', branches=branches, form_types=form_types,
                           selected_branch=session.get('branch_id'),
                           today=date.today().isoformat())


_TR_SELECT = '''
    SELECT tr.id, tr.from_branch_id, tr.to_branch_id, tr.form_type_id,
           tr.quantity, tr.notes, tr.status, tr.reject_reason,
           tr.requested_by, tr.approved_by, tr.created_at, tr.updated_at,
           fb.name from_branch_name, fb.code from_branch_code,
           tb.name to_branch_name,  tb.code to_branch_code,
           f.name form_name, f.unit,
           COALESCE(inv.quantity, 0) stock_qty
    FROM transfer_requests tr
    JOIN branches  fb  ON tr.from_branch_id = fb.id
    JOIN branches  tb  ON tr.to_branch_id   = tb.id
    JOIN form_types f  ON tr.form_type_id   = f.id
    LEFT JOIN inventory inv ON inv.branch_id = tr.from_branch_id
                            AND inv.form_type_id = tr.form_type_id
'''


@app.route('/transfer', methods=['GET', 'POST'])
@login_required
def transfer():
    conn = get_db()
    role = session.get('role')
    bid  = session.get('branch_id')

    if request.method == 'POST':
        if role != 'admin':
            if not bid:
                flash('소속 지점이 없습니다. 관리자에게 문의하세요.', 'danger')
                conn.close()
                return redirect(url_for('dashboard'))
            to_bid = str(bid)
        else:
            to_bid = request.form.get('to_branch_id', '')

        from_bid = request.form.get('from_branch_id', '')
        fid      = request.form.get('form_type_id', '')
        notes    = request.form.get('notes', '')
        notify_email = request.form.get('notify_email', '').strip()
        try:
            qty = int(request.form['quantity'])
        except (ValueError, KeyError):
            flash('수량을 올바르게 입력해주세요.', 'danger')
            conn.close()
            return redirect(url_for('transfer'))

        if not from_bid or not fid or not to_bid:
            flash('모든 필수 항목을 입력해주세요.', 'danger')
            conn.close()
            return redirect(url_for('transfer'))
        if from_bid == to_bid:
            flash('출발·도착 지점이 같습니다.', 'danger')
            conn.close()
            return redirect(url_for('transfer'))
        if qty <= 0:
            flash('수량은 1 이상이어야 합니다.', 'danger')
            conn.close()
            return redirect(url_for('transfer'))

        from_b = conn.execute('SELECT name, email FROM branches WHERE id=%s', (from_bid,)).fetchone()
        to_b   = conn.execute('SELECT name, email FROM branches WHERE id=%s', (to_bid,)).fetchone()
        ft     = conn.execute('SELECT name, unit FROM form_types WHERE id=%s', (fid,)).fetchone()
        if not from_b or not to_b or not ft:
            flash('잘못된 요청입니다.', 'danger')
            conn.close()
            return redirect(url_for('transfer'))

        # 동일 조합 PENDING 중복 신청 방지
        dup = conn.execute(
            "SELECT id FROM transfer_requests "
            "WHERE from_branch_id=%s AND to_branch_id=%s AND form_type_id=%s AND status='PENDING'",
            (from_bid, to_bid, fid)
        ).fetchone()
        if dup:
            flash(f'동일한 이전 신청이 이미 처리 대기 중입니다. (신청 #{dup["id"]})', 'warning')
            conn.close()
            return redirect(url_for('transfer'))

        conn.execute(
            "INSERT INTO transfer_requests "
            "(from_branch_id, to_branch_id, form_type_id, quantity, notes, requested_by, notify_email) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (from_bid, to_bid, fid, qty, notes, session['username'], notify_email)
        )
        msg_from = (f"[이전 신청] {to_b['name']}에서 "
                    f"{ft['name']} {qty}{ft['unit']} 이전을 요청했습니다.")
        conn.execute(
            "INSERT INTO notifications (branch_id, message) VALUES (%s,%s)",
            (from_bid, msg_from)
        )
        msg_to = (f"[이전 신청] {from_b['name']}에 "
                  f"{ft['name']} {qty}{ft['unit']} 이전 신청이 완료되었습니다.")
        conn.execute(
            "INSERT INTO notifications (branch_id, message) VALUES (%s,%s)",
            (to_bid, msg_to)
        )
        conn.commit()
        # 이메일 발송 — 요청 받은 지점(from_bid)에만 발송
        send_mail(
            [from_b['email']],
            f"[ZEGO 이전 신청] {to_b['name']} → {from_b['name']} {ft['name']} {qty}{ft['unit']}",
            f"{to_b['name']} 지점에서 {ft['name']} {qty}{ft['unit']} 이전을 요청했습니다.\n\n"
            f"ZEGO에 로그인하여 받은 신청 탭에서 승인 또는 반려해 주세요."
        )
        flash(f'{from_b["name"]}에 이전 신청 완료. 승인을 기다려 주세요.', 'success')
        conn.close()
        return redirect(url_for('transfer') + '?tab=outbox')

    # GET — mark notifications read for this branch
    if bid:
        try:
            conn.execute(
                'UPDATE notifications SET is_read=1 WHERE branch_id=%s AND is_read=0', (bid,)
            )
            conn.commit()
        except Exception:
            pass

    branches   = conn.execute('SELECT * FROM branches ORDER BY type, code').fetchall()
    form_types = conn.execute('SELECT * FROM form_types WHERE is_active ORDER BY sort_order').fetchall()

    if role == 'admin':
        inbox = []   # 관리자는 받은신청(inbox) 미사용
    elif bid:
        inbox = conn.execute(
            _TR_SELECT + ' WHERE tr.from_branch_id=%s AND tr.status=%s ORDER BY tr.created_at DESC',
            (bid, 'PENDING')
        ).fetchall()
    else:
        inbox = []

    if role == 'admin':
        outbox = conn.execute(
            _TR_SELECT + ' ORDER BY tr.created_at DESC LIMIT 200'
        ).fetchall()
    elif bid:
        outbox = conn.execute(
            _TR_SELECT + ' WHERE tr.to_branch_id=%s ORDER BY tr.created_at DESC LIMIT 100',
            (bid,)
        ).fetchall()
    else:
        outbox = []

    conn.close()
    default_tab = 'history' if role == 'admin' else 'request'
    active_tab  = request.args.get('tab', default_tab)
    return render_template('transfer.html',
                           branches=branches, form_types=form_types,
                           inbox=inbox, outbox=outbox,
                           selected_branch=bid,
                           active_tab=active_tab,
                           inbox_count=len(inbox))


@app.route('/transfer/requests/<int:req_id>/approve', methods=['POST'])
@login_required
def approve_transfer(req_id):
    conn = get_db()
    req = conn.execute(
        _TR_SELECT + ' WHERE tr.id=%s', (req_id,)
    ).fetchone()
    if not req:
        flash('요청을 찾을 수 없습니다.', 'danger')
        conn.close()
        return redirect(url_for('transfer') + '?tab=inbox')

    role = session.get('role')
    bid  = session.get('branch_id')
    if role != 'admin' and bid != req['from_branch_id']:
        flash('승인 권한이 없습니다.', 'danger')
        conn.close()
        return redirect(url_for('transfer') + '?tab=inbox')
    if req['status'] != 'PENDING':
        flash('이미 처리된 요청입니다.', 'warning')
        conn.close()
        return redirect(url_for('transfer') + '?tab=inbox')

    stock = conn.execute(
        'SELECT quantity FROM inventory WHERE branch_id=%s AND form_type_id=%s',
        (req['from_branch_id'], req['form_type_id'])
    ).fetchone()
    avail = stock['quantity'] if stock else 0
    if avail < req['quantity']:
        flash(f'재고가 부족하여 승인할 수 없습니다. (현재 재고: {avail})', 'danger')
        conn.close()
        return redirect(url_for('transfer') + '?tab=inbox')

    notify_email = request.form.get('notify_email', '').strip()

    conn.execute(
        "UPDATE transfer_requests SET status='APPROVED', approved_by=%s, updated_at=NOW() WHERE id=%s",
        (session['username'], req_id)
    )
    msg = (f"[이전 승인] {req['from_branch_name']}에서 "
           f"{req['form_name']} {req['quantity']}{req['unit']} 이전이 승인되었습니다. "
           f"실물 수령 후 확인을 눌러주세요.")
    conn.execute(
        "INSERT INTO notifications (branch_id, message) VALUES (%s,%s)",
        (req['to_branch_id'], msg)
    )
    to_email = conn.execute('SELECT email FROM branches WHERE id=%s', (req['to_branch_id'],)).fetchone()
    conn.commit()

    # 이메일 통보 (옵션)
    if notify_email:
        email_subject = f"[ZEGO] 이전 요청 승인 — {req['form_name']}"
        email_body = f"""
        <div style="font-family:sans-serif;max-width:500px">
          <h3 style="color:#16A34A">이전 요청이 승인되었습니다</h3>
          <table style="border-collapse:collapse;width:100%">
            <tr><td style="padding:6px;font-weight:bold">양식명</td><td style="padding:6px">{req['form_name']}</td></tr>
            <tr style="background:#f9f9f9"><td style="padding:6px;font-weight:bold">수량</td><td style="padding:6px">{req['quantity']} {req['unit']}</td></tr>
            <tr><td style="padding:6px;font-weight:bold">발신 지점</td><td style="padding:6px">{req['from_branch_name']}</td></tr>
            <tr style="background:#f9f9f9"><td style="padding:6px;font-weight:bold">승인자</td><td style="padding:6px">{session['username']}</td></tr>
          </table>
          <p style="color:#666;margin-top:12px">실물 수령 후 ZEGO에서 수령 확인을 눌러주세요.</p>
          <p style="color:#666;font-size:0.85em">ZEGO 재고관리 시스템 자동 발송</p>
        </div>
        """
        send_email(notify_email, email_subject, email_body)

    flash('승인 완료. 요청 지점에 알림을 보냈습니다.', 'success')
    conn.close()
    send_mail(
        [to_email['email'] if to_email else None],
        f"[ZEGO 이전 승인] {req['from_branch_name']} → {req['form_name']} {req['quantity']}{req['unit']}",
        f"{req['from_branch_name']} 지점에서 {req['form_name']} {req['quantity']}{req['unit']} 이전이 승인되었습니다.\n\n"
        f"실물 수령 후 ZEGO에 로그인하여 수령 확인을 눌러주세요."
    )
    return redirect(url_for('transfer') + '?tab=inbox')


@app.route('/transfer/requests/<int:req_id>/reject', methods=['POST'])
@login_required
def reject_transfer(req_id):
    conn = get_db()
    req = conn.execute(
        _TR_SELECT + ' WHERE tr.id=%s', (req_id,)
    ).fetchone()
    if not req:
        flash('요청을 찾을 수 없습니다.', 'danger')
        conn.close()
        return redirect(url_for('transfer') + '?tab=inbox')

    role = session.get('role')
    bid  = session.get('branch_id')
    if role != 'admin' and bid != req['from_branch_id']:
        flash('반려 권한이 없습니다.', 'danger')
        conn.close()
        return redirect(url_for('transfer') + '?tab=inbox')
    if req['status'] != 'PENDING':
        flash('이미 처리된 요청입니다.', 'warning')
        conn.close()
        return redirect(url_for('transfer') + '?tab=inbox')

    reason = request.form.get('reject_reason', '').strip()
    if not reason:
        flash('반려 사유를 입력해주세요.', 'danger')
        conn.close()
        return redirect(url_for('transfer') + '?tab=inbox')

    notify_email = request.form.get('notify_email', '').strip()

    conn.execute(
        "UPDATE transfer_requests SET status='REJECTED', reject_reason=%s, "
        "approved_by=%s, updated_at=NOW() WHERE id=%s",
        (reason, session['username'], req_id)
    )
    msg = (f"[이전 반려] {req['from_branch_name']}에서 "
           f"{req['form_name']} {req['quantity']}{req['unit']} 이전이 반려되었습니다. "
           f"사유: {reason}")
    conn.execute(
        "INSERT INTO notifications (branch_id, message) VALUES (%s,%s)",
        (req['to_branch_id'], msg)
    )
    to_email = conn.execute('SELECT email FROM branches WHERE id=%s', (req['to_branch_id'],)).fetchone()
    conn.commit()

    # 이메일 통보 (옵션)
    if notify_email:
        email_subject = f"[ZEGO] 이전 요청 반려 — {req['form_name']}"
        email_body = f"""
        <div style="font-family:sans-serif;max-width:500px">
          <h3 style="color:#DC2626">이전 요청이 반려되었습니다</h3>
          <table style="border-collapse:collapse;width:100%">
            <tr><td style="padding:6px;font-weight:bold">양식명</td><td style="padding:6px">{req['form_name']}</td></tr>
            <tr style="background:#f9f9f9"><td style="padding:6px;font-weight:bold">수량</td><td style="padding:6px">{req['quantity']} {req['unit']}</td></tr>
            <tr><td style="padding:6px;font-weight:bold">발신 지점</td><td style="padding:6px">{req['from_branch_name']}</td></tr>
            <tr style="background:#f9f9f9"><td style="padding:6px;font-weight:bold">반려 사유</td><td style="padding:6px">{reason}</td></tr>
          </table>
          <p style="color:#666;font-size:0.85em;margin-top:16px">ZEGO 재고관리 시스템 자동 발송</p>
        </div>
        """
        send_email(notify_email, email_subject, email_body)

    flash('반려 처리 완료. 요청 지점에 알림을 보냈습니다.', 'success')
    conn.close()
    send_mail(
        [to_email['email'] if to_email else None],
        f"[ZEGO 이전 반려] {req['from_branch_name']} → {req['form_name']} {req['quantity']}{req['unit']}",
        f"{req['from_branch_name']} 지점에서 {req['form_name']} {req['quantity']}{req['unit']} 이전이 반려되었습니다.\n\n"
        f"반려 사유: {reason}"
    )
    return redirect(url_for('transfer') + '?tab=inbox')


@app.route('/transfer/requests/<int:req_id>/confirm', methods=['POST'])
@login_required
def confirm_transfer(req_id):
    conn = get_db()
    req = conn.execute(
        _TR_SELECT + ' WHERE tr.id=%s', (req_id,)
    ).fetchone()
    if not req:
        flash('요청을 찾을 수 없습니다.', 'danger')
        conn.close()
        return redirect(url_for('transfer') + '?tab=outbox')

    role = session.get('role')
    bid  = session.get('branch_id')
    if role != 'admin' and bid != req['to_branch_id']:
        flash('확인 권한이 없습니다.', 'danger')
        conn.close()
        return redirect(url_for('transfer') + '?tab=outbox')
    if req['status'] != 'APPROVED':
        flash('승인된 요청만 확인할 수 있습니다.', 'warning')
        conn.close()
        return redirect(url_for('transfer') + '?tab=outbox')

    qty      = req['quantity']
    from_bid = req['from_branch_id']
    to_bid   = req['to_branch_id']
    fid      = req['form_type_id']

    stock = conn.execute(
        'SELECT quantity FROM inventory WHERE branch_id=%s AND form_type_id=%s',
        (from_bid, fid)
    ).fetchone()
    if not stock or stock['quantity'] < qty:
        flash('출발 지점 재고가 부족합니다. 관리자에게 문의하세요.', 'danger')
        conn.close()
        return redirect(url_for('transfer') + '?tab=outbox')

    conn.execute(
        'UPDATE inventory SET quantity=quantity-%s, last_updated=NOW() '
        'WHERE branch_id=%s AND form_type_id=%s',
        (qty, from_bid, fid)
    )
    conn.execute('''
        INSERT INTO inventory (branch_id, form_type_id, quantity, last_updated)
        VALUES (%s,%s,%s,NOW())
        ON CONFLICT(branch_id, form_type_id) DO UPDATE SET
          quantity     = inventory.quantity + EXCLUDED.quantity,
          last_updated = NOW()
    ''', (to_bid, fid, qty))
    conn.execute(
        "INSERT INTO transactions "
        "(type, form_type_id, from_branch_id, to_branch_id, quantity, notes, created_by) "
        "VALUES ('TRANSFER',%s,%s,%s,%s,%s,%s)",
        (fid, from_bid, to_bid, qty, req['notes'] or '', session['username'])
    )
    conn.execute(
        "UPDATE transfer_requests SET status='CONFIRMED', updated_at=NOW() WHERE id=%s",
        (req_id,)
    )
    conn.commit()
    flash(f'수령 확인 완료 ✔ {req["form_name"]} {qty}{req["unit"]} 재고가 반영되었습니다.', 'success')
    conn.close()
    return redirect(url_for('transfer') + '?tab=outbox')


@app.route('/transactions')
@login_required
def transactions():
    conn = get_db()
    branches   = conn.execute('SELECT * FROM branches ORDER BY type, code').fetchall()
    form_types = conn.execute('SELECT * FROM form_types WHERE is_active ORDER BY sort_order').fetchall()

    bf        = request.args.get('branch_id', '')
    ff        = request.args.get('form_type_id', '')
    tf        = request.args.get('type', '')
    month_f   = request.args.get('month_f', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    conditions, params = ['1=1'], []
    role = session.get('role')
    bid  = session.get('branch_id')
    if role != 'admin' and bid:
        conditions.append('(t.from_branch_id=%s OR t.to_branch_id=%s)')
        params.extend([bid, bid])
    elif bf:
        conditions.append('(t.from_branch_id=%s OR t.to_branch_id=%s)')
        params.extend([bf, bf])
    if ff:
        conditions.append('t.form_type_id=%s'); params.append(ff)
    if tf:
        conditions.append('t.type=%s'); params.append(tf)
    if month_f:
        conditions.append("TO_CHAR(t.transaction_date, 'YYYY-MM') = %s"); params.append(month_f)
    if date_from:
        conditions.append('t.transaction_date >= %s'); params.append(date_from)
    if date_to:
        conditions.append('t.transaction_date <= %s'); params.append(date_to)

    rows = conn.execute(f'''
        SELECT t.*, f.name form_name, f.unit,
               fb.name from_branch_name, fb.code from_branch_code,
               tb.name to_branch_name,   tb.code to_branch_code
        FROM transactions t
        JOIN form_types f ON t.form_type_id = f.id
        LEFT JOIN branches fb ON t.from_branch_id = fb.id
        LEFT JOIN branches tb ON t.to_branch_id   = tb.id
        WHERE {" AND ".join(conditions)}
        ORDER BY t.transaction_date DESC, t.created_at DESC LIMIT 300
    ''', params).fetchall()
    conn.close()
    return render_template('transactions.html', rows=rows, branches=branches, form_types=form_types,
                           bf=bf, ff=ff, tf=tf, month_f=month_f,
                           date_from=date_from, date_to=date_to)


@app.route('/transactions/<int:tx_id>/edit', methods=['POST'])
@login_required
def transaction_edit(tx_id):
    conn = get_db()
    ph   = '%s' if not USE_SQLITE else '?'
    role = session.get('role')
    bid  = session.get('branch_id')
    data = request.get_json(silent=True) or {}

    tx = conn.execute(
        f'SELECT * FROM transactions WHERE id={ph}', (tx_id,)
    ).fetchone()
    if not tx:
        conn.close()
        return jsonify({'ok': False, 'msg': '기록을 찾을 수 없습니다.'}), 404
    if tx['type'] != 'OUT':
        conn.close()
        return jsonify({'ok': False, 'msg': '출고 기록만 수정할 수 있습니다.'}), 400
    if role != 'admin' and tx['from_branch_id'] != bid:
        conn.close()
        return jsonify({'ok': False, 'msg': '권한이 없습니다.'}), 403

    try:
        new_qty = int(data.get('quantity', tx['quantity']))
    except (TypeError, ValueError):
        conn.close()
        return jsonify({'ok': False, 'msg': '수량이 올바르지 않습니다.'}), 400
    if new_qty <= 0:
        conn.close()
        return jsonify({'ok': False, 'msg': '수량은 1 이상이어야 합니다.'}), 400

    from datetime import date as _date
    new_date  = (data.get('transaction_date') or str(tx['transaction_date'] or _date.today().isoformat()))[:10]
    new_notes = (data.get('notes') or '').strip()

    old_qty  = tx['quantity']
    qty_diff = new_qty - old_qty   # 양수: 더 차감, 음수: 재고 복원

    if qty_diff != 0:
        inv = conn.execute(
            f'SELECT quantity FROM inventory WHERE branch_id={ph} AND form_type_id={ph}',
            (tx['from_branch_id'], tx['form_type_id'])
        ).fetchone()
        cur_inv = inv['quantity'] if inv else 0
        new_inv = cur_inv - qty_diff
        if new_inv < 0:
            conn.close()
            return jsonify({'ok': False, 'msg': f'재고가 부족합니다. (현재 재고: {cur_inv}, 추가 차감 필요: {qty_diff})'}), 400
        conn.execute(
            f'UPDATE inventory SET quantity={ph}, last_updated=NOW() '
            f'WHERE branch_id={ph} AND form_type_id={ph}',
            (new_inv, tx['from_branch_id'], tx['form_type_id'])
        )

    conn.execute(
        f'UPDATE transactions SET quantity={ph}, transaction_date={ph}, period_month={ph}, notes={ph} '
        f'WHERE id={ph}',
        (new_qty, new_date, new_date[:7], new_notes, tx_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'new_qty': new_qty, 'new_date': new_date, 'new_notes': new_notes})


@app.route('/report')
@login_required
def report():
    conn = get_db()
    bid  = session.get('branch_id')
    role = session.get('role')

    b_cond = f'AND b.id={bid}'          if role != 'admin' and bid else ''
    i_cond = f'AND i.branch_id={bid}'   if role != 'admin' and bid else ''
    t_cond = f'AND (t.from_branch_id={bid} OR t.to_branch_id={bid})' if role != 'admin' and bid else ''

    summary = conn.execute(f'''
        SELECT b.code, b.name, b.type,
               COUNT(CASE WHEN i.quantity=0 THEN 1 END) empty_cnt,
               COUNT(CASE WHEN i.quantity>0 AND i.quantity<=f.min_threshold THEN 1 END) low_cnt,
               COUNT(CASE WHEN i.quantity>f.min_threshold THEN 1 END) ok_cnt,
               COALESCE(SUM(i.quantity*f.unit_price),0) total_value
        FROM branches b
        LEFT JOIN inventory i  ON b.id=i.branch_id
        LEFT JOIN form_types f ON i.form_type_id=f.id
        WHERE 1=1 {b_cond}
        GROUP BY b.id, b.code, b.name, b.type ORDER BY b.type, b.code
    ''').fetchall()

    monthly_raw = conn.execute(f'''
        SELECT TO_CHAR(t.created_at, 'YYYY-MM') mo, t.type,
               COUNT(*) tx_count, COALESCE(SUM(t.quantity),0) total_qty
        FROM transactions t
        WHERE t.created_at >= CURRENT_DATE - INTERVAL '6 months' {t_cond}
        GROUP BY TO_CHAR(t.created_at, 'YYYY-MM'), t.type ORDER BY TO_CHAR(t.created_at, 'YYYY-MM') ASC
    ''').fetchall()

    months_set = sorted({r['mo'] for r in monthly_raw})
    def _monthly_vals(tx_type):
        m = {r['mo']: r['total_qty'] for r in monthly_raw if r['type'] == tx_type}
        return [m.get(mo, 0) for mo in months_set]
    chart_monthly = {
        'labels':   months_set,
        'in':       _monthly_vals('IN'),
        'out':      _monthly_vals('OUT'),
        'transfer': _monthly_vals('TRANSFER'),
    }

    status_row = conn.execute(f'''
        SELECT COUNT(CASE WHEN i.quantity=0 THEN 1 END) empty_cnt,
               COUNT(CASE WHEN i.quantity>0 AND i.quantity<=f.min_threshold THEN 1 END) low_cnt,
               COUNT(CASE WHEN i.quantity>f.min_threshold THEN 1 END) ok_cnt
        FROM inventory i
        JOIN form_types f ON i.form_type_id=f.id
        WHERE 1=1 {i_cond}
    ''').fetchone()
    chart_status = {
        'ok':    status_row['ok_cnt']    or 0,
        'low':   status_row['low_cnt']   or 0,
        'empty': status_row['empty_cnt'] or 0,
    }

    chart_branch_qty = conn.execute(f'''
        SELECT b.code label, b.type btype, COALESCE(SUM(i.quantity),0) qty
        FROM branches b
        LEFT JOIN inventory i ON b.id=i.branch_id
        WHERE 1=1 {b_cond}
        GROUP BY b.id, b.code, b.type HAVING COALESCE(SUM(i.quantity),0)>0 ORDER BY qty DESC LIMIT 20
    ''').fetchall()

    chart_branch_status = conn.execute(f'''
        SELECT b.code label,
               COALESCE(SUM(CASE WHEN i.quantity=0 THEN 1 ELSE 0 END),0) empty_cnt,
               COALESCE(SUM(CASE WHEN i.quantity>0 AND i.quantity<=f.min_threshold THEN 1 ELSE 0 END),0) low_cnt,
               COALESCE(SUM(CASE WHEN i.quantity>f.min_threshold THEN 1 ELSE 0 END),0) ok_cnt
        FROM branches b
        LEFT JOIN inventory i  ON b.id=i.branch_id
        LEFT JOIN form_types f ON i.form_type_id=f.id
        WHERE 1=1 {b_cond}
        GROUP BY b.id, b.code, b.type
        HAVING (COALESCE(SUM(CASE WHEN i.quantity=0 THEN 1 ELSE 0 END),0) +
                COALESCE(SUM(CASE WHEN i.quantity>0 AND i.quantity<=f.min_threshold THEN 1 ELSE 0 END),0) +
                COALESCE(SUM(CASE WHEN i.quantity>f.min_threshold THEN 1 ELSE 0 END),0)) > 0
        ORDER BY b.type, b.code
    ''').fetchall()

    chart_top_out = conn.execute(f'''
        SELECT f.name label, COALESCE(SUM(t.quantity),0) qty
        FROM transactions t
        JOIN form_types f ON t.form_type_id=f.id
        WHERE t.type='OUT' AND t.created_at >= CURRENT_DATE - INTERVAL '30 days' {t_cond}
        GROUP BY f.id, f.name ORDER BY qty DESC LIMIT 10
    ''').fetchall()

    chart_form_qty = conn.execute(f'''
        SELECT f.name label, COALESCE(SUM(i.quantity),0) qty
        FROM form_types f
        LEFT JOIN inventory i ON f.id=i.form_type_id
        WHERE 1=1 {i_cond}
        GROUP BY f.id, f.name HAVING COALESCE(SUM(i.quantity),0)>0 ORDER BY qty DESC LIMIT 20
    ''').fetchall()

    flight_cond = f'AND fs.branch_id={bid}' if role != 'admin' and bid else ''
    flight_raw = conn.execute(f'''
        SELECT fs.year_month, SUM(fs.flight_count) total_flights
        FROM flight_schedule fs
        WHERE fs.year_month >= TO_CHAR(CURRENT_DATE - INTERVAL '6 months', 'YYYY-MM')
        {flight_cond}
        GROUP BY fs.year_month ORDER BY fs.year_month
    ''').fetchall()
    chart_flight = {
        'labels': [r['year_month'] for r in flight_raw],
        'data':   [r['total_flights'] for r in flight_raw],
    }

    conn.close()

    def _rows_to_chart(rows):
        return {'labels': [r['label'] for r in rows],
                'data':   [r['qty']   for r in rows]}

    return render_template('report.html',
        summary=summary, monthly=monthly_raw,
        chart_monthly      = chart_monthly,
        chart_status       = chart_status,
        chart_branch_qty   = _rows_to_chart(chart_branch_qty),
        chart_branch_status= {'labels': [r['label'] for r in chart_branch_status],
                               'ok':    [r['ok_cnt']    for r in chart_branch_status],
                               'low':   [r['low_cnt']   for r in chart_branch_status],
                               'empty': [r['empty_cnt'] for r in chart_branch_status]},
        chart_top_out      = _rows_to_chart(chart_top_out),
        chart_form_qty     = _rows_to_chart(chart_form_qty),
        chart_flight       = chart_flight,
    )


# ── 분기 신청 추정 ────────────────────────────────────────────────────────────

@app.route('/report/forecast')
@login_required
def forecast():
    import math
    from datetime import date
    from collections import defaultdict

    conn  = get_db()
    today = date.today()
    role  = session.get('role')
    bid   = session.get('branch_id')
    bf    = request.args.get('branch_id', '')

    cur_q        = (today.month - 1) // 3 + 1
    next_q       = cur_q % 4 + 1
    nq_y         = today.year + (1 if next_q == 1 else 0)
    nq_start_m   = (cur_q * 3) % 12 + 1
    nq_end_m     = nq_start_m + 2
    is_qend      = today.month in (3, 6, 9, 12)

    branches = conn.execute('SELECT * FROM branches ORDER BY type, code').fetchall()

    if role != 'admin':
        target_ids = [bid] if bid else []
    elif bf:
        target_ids = [int(bf)]
    else:
        target_ids = [b['id'] for b in branches]

    if not target_ids:
        conn.close()
        return render_template('forecast.html',
            forecast_sections=[], month_labels=[], branches=branches,
            today=today, cur_q=cur_q, next_q=next_q,
            nq_y=nq_y, nq_start_m=nq_start_m, nq_end_m=nq_end_m,
            is_qend=is_qend, bf=bf)

    id_str = ','.join(str(x) for x in target_ids)

    month_labels = []
    y, m = today.year, today.month - 1
    if m == 0: m, y = 12, y - 1
    for _ in range(6):
        month_labels.insert(0, f'{y}-{m:02d}')
        m -= 1
        if m == 0: m, y = 12, y - 1

    raw = conn.execute(f'''
        SELECT t.from_branch_id bid, f.id fid,
               f.name form_name, f.unit, f.unit_detail,
               COALESCE(t.period_month, TO_CHAR(t.created_at, 'YYYY-MM')) mo,
               SUM(t.quantity) qty
        FROM transactions t
        JOIN form_types f ON t.form_type_id = f.id
        WHERE t.type = 'OUT'
          AND t.from_branch_id IN ({id_str})
          AND COALESCE(t.period_month, TO_CHAR(t.created_at, 'YYYY-MM'))
              BETWEEN TO_CHAR(CURRENT_DATE - INTERVAL '6 months', 'YYYY-MM')
                  AND TO_CHAR(CURRENT_DATE, 'YYYY-MM')
        GROUP BY t.from_branch_id, f.id, f.name, f.unit, f.unit_detail, COALESCE(t.period_month, TO_CHAR(t.created_at, 'YYYY-MM'))
    ''').fetchall()

    inv_map = {(r['branch_id'], r['form_type_id']): r['quantity']
               for r in conn.execute(
                   f'SELECT branch_id, form_type_id, quantity FROM inventory WHERE branch_id IN ({id_str})'
               ).fetchall()}

    # ── 운항편수 보정 인자 ────────────────────────────────────────────
    nq_months = []
    for i in range(3):
        m_i, y_i = nq_start_m + i, nq_y
        if m_i > 12: m_i -= 12; y_i += 1
        nq_months.append(f'{y_i}-{m_i:02d}')

    hist_ms = "'" + "','".join(month_labels) + "'"
    next_ms  = "'" + "','".join(nq_months) + "'"
    fh_rows = conn.execute(f'''
        SELECT branch_id, flight_count FROM flight_schedule
        WHERE branch_id IN ({id_str}) AND year_month IN ({hist_ms})
    ''').fetchall()
    fn_rows = conn.execute(f'''
        SELECT branch_id, flight_count FROM flight_schedule
        WHERE branch_id IN ({id_str}) AND year_month IN ({next_ms})
    ''').fetchall()

    flight_factor_map = {}
    flight_info_map   = {}
    for b_id in target_ids:
        hc = [r['flight_count'] for r in fh_rows if r['branch_id'] == b_id and r['flight_count'] > 0]
        nc = [r['flight_count'] for r in fn_rows if r['branch_id'] == b_id and r['flight_count'] > 0]
        if hc and nc:
            ah, an = sum(hc) / len(hc), sum(nc) / len(nc)
            f = an / ah
            flight_factor_map[b_id] = f
            flight_info_map[b_id]   = dict(avg_hist=round(ah, 1), avg_next=round(an, 1), factor=round(f, 2))
        else:
            flight_factor_map[b_id] = 1.0
            flight_info_map[b_id]   = None

    branch_map = {b['id']: dict(b) for b in branches}

    bucket = defaultdict(lambda: defaultdict(int))
    meta   = {}
    for r in raw:
        key = (r['bid'], r['fid'])
        bucket[key][r['mo']] += r['qty']
        meta[key] = (r['form_name'], r['unit'], r['unit_detail'])

    forecast_by_branch = defaultdict(list)
    for (b_id, f_id), monthly in bucket.items():
        vals     = [monthly.get(mo, 0) for mo in month_labels]
        total_6m = sum(vals)
        nonzero  = [v for v in vals if v > 0]
        avg_m    = total_6m / len(nonzero) if nonzero else 0

        r3 = sum(vals[3:]) / 3
        o3 = sum(vals[:3]) / 3
        if o3 > 0:
            trend_pct = (r3 - o3) / o3
        elif r3 > 0:
            trend_pct = 1.0
        else:
            trend_pct = 0.0

        if trend_pct > 0.15:
            trend_label, trend_icon, trend_factor = '증가', '↑', 1.15
        elif trend_pct < -0.15:
            trend_label, trend_icon, trend_factor = '감소', '↓', 0.9
        else:
            trend_label, trend_icon, trend_factor = '안정', '→', 1.0

        flight_factor = flight_factor_map.get(b_id, 1.0)
        cur_stock   = inv_map.get((b_id, f_id), 0)
        nq_est      = math.ceil(avg_m * 3 * trend_factor * flight_factor)
        recommended = max(0, nq_est - cur_stock)

        fname, unit, udesc = meta.get((b_id, f_id), ('', '', ''))
        forecast_by_branch[b_id].append(dict(
            form_name=fname, unit=unit, unit_detail=udesc,
            vals=vals, total_6m=total_6m,
            avg_monthly=round(avg_m, 1),
            trend_pct=round(trend_pct * 100),
            trend_label=trend_label, trend_icon=trend_icon,
            cur_stock=cur_stock,
            nq_est=nq_est,
            recommended=recommended,
        ))

    for b_id in forecast_by_branch:
        forecast_by_branch[b_id].sort(key=lambda x: x['recommended'], reverse=True)

    forecast_sections = [
        dict(branch_id=b_id,
             branch_code=branch_map[b_id]['code'],
             branch_name=branch_map[b_id]['name'],
             branch_type=branch_map[b_id]['type'],
             form_items=forecast_by_branch[b_id],
             col_sums=[sum(it['vals'][i] for it in forecast_by_branch[b_id]) for i in range(6)])
        for b_id in sorted(forecast_by_branch)
        if forecast_by_branch[b_id]
    ]

    conn.close()
    return render_template('forecast.html',
        forecast_sections=forecast_sections,
        month_labels=month_labels,
        branches=branches, bf=bf,
        today=today, cur_q=cur_q, next_q=next_q,
        nq_y=nq_y, nq_start_m=nq_start_m, nq_end_m=nq_end_m,
        is_qend=is_qend, flight_info_map=flight_info_map)


@app.route('/report/forecast/download')
@login_required
def forecast_download():
    import math
    from datetime import date
    from collections import defaultdict

    conn  = get_db()
    today = date.today()
    role  = session.get('role')
    bid   = session.get('branch_id')
    bf    = request.args.get('branch_id', '')

    cur_q      = (today.month - 1) // 3 + 1
    next_q     = cur_q % 4 + 1
    nq_y       = today.year + (1 if next_q == 1 else 0)
    nq_start_m = (cur_q * 3) % 12 + 1

    branches = conn.execute('SELECT * FROM branches ORDER BY type, code').fetchall()
    if role != 'admin':
        target_ids = [bid] if bid else []
    elif bf:
        target_ids = [int(bf)]
    else:
        target_ids = [b['id'] for b in branches]

    if not target_ids:
        conn.close()
        flash('데이터 없음', 'warning')
        return redirect(url_for('forecast'))

    id_str = ','.join(str(x) for x in target_ids)

    month_labels = []
    y, m = today.year, today.month - 1
    if m == 0: m, y = 12, y - 1
    for _ in range(6):
        month_labels.insert(0, f'{y}-{m:02d}')
        m -= 1
        if m == 0: m, y = 12, y - 1

    raw = conn.execute(f'''
        SELECT t.from_branch_id bid, f.id fid, f.name form_name, f.unit,
               COALESCE(t.period_month, TO_CHAR(t.created_at, 'YYYY-MM')) mo,
               SUM(t.quantity) qty
        FROM transactions t JOIN form_types f ON t.form_type_id=f.id
        WHERE t.type='OUT' AND t.from_branch_id IN ({id_str})
          AND COALESCE(t.period_month, TO_CHAR(t.created_at, 'YYYY-MM'))
              BETWEEN TO_CHAR(CURRENT_DATE - INTERVAL '6 months', 'YYYY-MM')
                  AND TO_CHAR(CURRENT_DATE, 'YYYY-MM')
        GROUP BY t.from_branch_id, f.id, f.name, f.unit, COALESCE(t.period_month, TO_CHAR(t.created_at, 'YYYY-MM'))
    ''').fetchall()

    inv_map = {(r['branch_id'], r['form_type_id']): r['quantity']
               for r in conn.execute(
                   f'SELECT branch_id, form_type_id, quantity FROM inventory WHERE branch_id IN ({id_str})'
               ).fetchall()}

    # ── 운항편수 보정 인자 ────────────────────────────────────────────
    dl_nq_months = []
    for i in range(3):
        m_i, y_i = nq_start_m + i, nq_y
        if m_i > 12: m_i -= 12; y_i += 1
        dl_nq_months.append(f'{y_i}-{m_i:02d}')

    dl_hist_ms = "'" + "','".join(month_labels) + "'"
    dl_next_ms  = "'" + "','".join(dl_nq_months) + "'"
    dl_fh = conn.execute(f'''
        SELECT branch_id, flight_count FROM flight_schedule
        WHERE branch_id IN ({id_str}) AND year_month IN ({dl_hist_ms})
    ''').fetchall()
    dl_fn = conn.execute(f'''
        SELECT branch_id, flight_count FROM flight_schedule
        WHERE branch_id IN ({id_str}) AND year_month IN ({dl_next_ms})
    ''').fetchall()

    dl_flight_factor = {}
    for b_id in target_ids:
        hc = [r['flight_count'] for r in dl_fh if r['branch_id'] == b_id and r['flight_count'] > 0]
        nc = [r['flight_count'] for r in dl_fn if r['branch_id'] == b_id and r['flight_count'] > 0]
        if hc and nc:
            dl_flight_factor[b_id] = (sum(nc) / len(nc)) / (sum(hc) / len(hc))
        else:
            dl_flight_factor[b_id] = 1.0

    branch_map = {b['id']: dict(b) for b in branches}

    bucket = defaultdict(lambda: defaultdict(int))
    meta   = {}
    for r in raw:
        key = (r['bid'], r['fid'])
        bucket[key][r['mo']] += r['qty']
        meta[key] = (r['form_name'], r['unit'])

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    is_admin = role == 'admin'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{nq_y}년 Q{next_q} 신청 추정'
    hfill, hfont, halign = _make_header_style()
    bd = _border()

    headers = (['지점코드','지점명'] if is_admin else []) + \
              ['양식명','단위'] + [f'{mo}출고' for mo in month_labels] + \
              ['월평균','추세','현재재고','다음분기예상','권장신청수량']
    col_widths = ([10,14] if is_admin else []) + \
                 [34,8] + [10]*6 + [10,8,10,12,12]
    for col,(h,w) in enumerate(zip(headers,col_widths),1):
        c = ws.cell(row=1,column=col,value=h)
        c.fill,c.font,c.alignment = hfill,hfont,halign
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    green_f = PatternFill('solid',fgColor='D1FAE5')
    yellow_f= PatternFill('solid',fgColor='FEF9C3')
    red_f   = PatternFill('solid',fgColor='FEE2E2')

    row_num = 2
    for b_id in sorted(bucket, key=lambda x: branch_map.get(x, {}).get('code', '')):
        b = branch_map.get(b_id, {})
        for (bb,fid),monthly in {k:v for k,v in bucket.items() if k[0]==b_id}.items():
            fname,unit = meta.get((bb,fid),('',''))
            vals   = [monthly.get(mo,0) for mo in month_labels]
            nonzero= [v for v in vals if v > 0]
            avg_m  = sum(vals)/len(nonzero) if nonzero else 0
            r3,o3  = sum(vals[3:])/3, sum(vals[:3])/3
            tp     = (r3-o3)/o3 if o3>0 else (1.0 if r3>0 else 0.0)
            tf     = 1.15 if tp>0.15 else (0.9 if tp<-0.15 else 1.0)
            tl     = '증가' if tp>0.15 else ('감소' if tp<-0.15 else '안정')
            cur_s  = inv_map.get((bb,fid),0)
            nq_est = math.ceil(avg_m * 3 * tf * dl_flight_factor.get(bb, 1.0))
            rec    = max(0, nq_est-cur_s)

            row_vals = ([b.get('code',''), b.get('name','')] if is_admin else []) + \
                       [fname, unit] + vals + \
                       [round(avg_m,1), tl, cur_s, nq_est, rec]
            fill = red_f if rec>5 else (yellow_f if rec>0 else green_f)
            for col,v in enumerate(row_vals,1):
                c = ws.cell(row=row_num,column=col,value=v)
                c.border = bd
                c.alignment = Alignment(vertical='center',
                    horizontal='center' if col>=(3+(2 if is_admin else 0)) else 'left')
            last_col = len(row_vals)
            ws.cell(row=row_num,column=last_col).fill = fill
            ws.row_dimensions[row_num].height = 16
            row_num += 1

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    conn.close()

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname_dl = f'{nq_y}년_Q{next_q}_신청추정_{today.strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname_dl,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 운항편수 관리 ──────────────────────────────────────────────────────────────

@app.route('/flight-schedule', methods=['GET', 'POST'])
@login_required
def flight_schedule_view():
    from datetime import date
    conn  = get_db()
    role  = session.get('role')
    bid   = session.get('branch_id')
    today = date.today()
    bf    = request.args.get('branch_id', '')

    branches = conn.execute('SELECT * FROM branches ORDER BY type, code').fetchall()

    if role == 'admin':
        view_bid = int(bf) if bf else (branches[0]['id'] if branches else None)
    else:
        view_bid = bid

    if request.method == 'POST':
        post_bid = int(request.form.get('branch_id', view_bid or 0))
        if role != 'admin' and post_bid != bid:
            flash('권한이 없습니다.', 'danger')
            conn.close()
            return redirect(url_for('flight_schedule_view'))

        months = request.form.getlist('month')
        counts = request.form.getlist('count')
        for mo, cnt in zip(months, counts):
            try:
                c = max(0, int(cnt)) if cnt.strip() else 0
                conn.execute('''
                    INSERT INTO flight_schedule (branch_id, year_month, flight_count, updated_at)
                    VALUES (%s, %s, %s, NOW())
                    ON CONFLICT (branch_id, year_month)
                    DO UPDATE SET flight_count=EXCLUDED.flight_count, updated_at=NOW()
                ''', (post_bid, mo, c))
            except (ValueError, Exception):
                pass
        conn.commit()
        flash('운항편수가 저장되었습니다.', 'success')
        conn.close()
        qs = f'?branch_id={post_bid}' if role == 'admin' else ''
        return redirect(url_for('flight_schedule_view') + qs)

    # 과거 3개월 + 현재월 + 미래 8개월 = 12개월 표시
    month_list = []
    y, m = today.year, today.month - 3
    for _ in range(12):
        if m <= 0: m += 12; y -= 1
        if m > 12: m -= 12; y += 1
        month_list.append(f'{y}-{m:02d}')
        m += 1

    existing = {}
    if view_bid:
        rows = conn.execute(
            'SELECT year_month, flight_count FROM flight_schedule WHERE branch_id=%s',
            (view_bid,)
        ).fetchall()
        existing = {r['year_month']: r['flight_count'] for r in rows}

    conn.close()
    return render_template('flight_schedule.html',
        branches=branches, view_bid=view_bid,
        month_list=month_list, existing=existing,
        today_str=today.strftime('%Y-%m'), bf=bf)


@app.route('/api/stock')
@login_required
def api_stock():
    bid = request.args.get('branch_id')
    fid = request.args.get('form_type_id')
    conn = get_db()
    row = conn.execute(
        'SELECT quantity FROM inventory WHERE branch_id=%s AND form_type_id=%s',
        (bid, fid)
    ).fetchone()
    conn.close()
    return jsonify({'quantity': row['quantity'] if row else 0})


@app.route('/api/notifications/read', methods=['POST'])
@login_required
def mark_notifications_read():
    bid = session.get('branch_id')
    if bid:
        conn = get_db()
        conn.execute(
            'UPDATE notifications SET is_read=1 WHERE branch_id=%s AND is_read=0', (bid,)
        )
        conn.commit()
        conn.close()
    return jsonify({'ok': True})


# ── 카탈로그 관련 헬퍼 ────────────────────────────────────────────────────────

def _seed_catalog_defs(conn):
    """CATALOG_ITEMS를 catalog_defs에 시딩.
    name/cat/sub_desc/sort_order는 코드 기준으로 갱신.
    img는 관리자가 이미지를 업로드한 경우(img_data != '')엔 유지,
    업로드 이미지가 없는 경우에만 코드 기준값으로 갱신."""
    # user_deleted 컬럼이 존재하는지 확인 후 ON CONFLICT 조건 결정
    try:
        conn.execute('SELECT user_deleted FROM catalog_defs LIMIT 1')
        has_user_deleted = True
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        has_user_deleted = False

    for item in CATALOG_ITEMS:
        if has_user_deleted:
            sql = (
                'INSERT INTO catalog_defs (code, img, name, cat, sub_desc, sort_order) '
                'VALUES (%s, %s, %s, %s, %s, %s) '
                'ON CONFLICT (code) DO UPDATE SET '
                '  img=CASE WHEN COALESCE(catalog_defs.img_data,\'\')=\'\' THEN EXCLUDED.img ELSE catalog_defs.img END, '
                '  sort_order=EXCLUDED.sort_order '
                'WHERE catalog_defs.user_deleted IS NOT TRUE'
            )
        else:
            sql = (
                'INSERT INTO catalog_defs (code, img, name, cat, sub_desc, sort_order) '
                'VALUES (%s, %s, %s, %s, %s, %s) '
                'ON CONFLICT (code) DO UPDATE SET '
                '  img=CASE WHEN COALESCE(catalog_defs.img_data,\'\')=\'\' THEN EXCLUDED.img ELSE catalog_defs.img END, '
                '  sort_order=EXCLUDED.sort_order'
            )
        conn.execute(sql, (item['code'], item['img'], item['name'], item['cat'],
                           item.get('sub_desc', ''), item.get('sort', 0)))
    conn.commit()


def _warm_img_tmp_cache_bg():
    """백그라운드 스레드에서 DB img_data를 /tmp에 일괄 저장.
    페이지 응답을 블로킹하지 않고, 이미지 요청이 들어오기 전에 /tmp를 미리 채움."""
    global _img_tmp_warmed
    if _img_tmp_warmed or USE_SQLITE:
        return
    import base64 as _b64, psycopg2 as _pg2
    try:
        _conn = _DB(_pg2.connect(**_PG))
        os.makedirs(_TMP_IMG_DIR, exist_ok=True)
        rows = _conn.execute(
            "SELECT img, img_data FROM catalog_defs "
            "WHERE img_data IS NOT NULL AND img_data != ''"
        ).fetchall()
        for row in rows:
            tmp_path = os.path.join(_TMP_IMG_DIR, row['img'] + '.png')
            if not os.path.isfile(tmp_path):
                try:
                    with open(tmp_path, 'wb') as _f:
                        _f.write(_b64.b64decode(row['img_data']))
                except Exception:
                    pass
        _conn.close()
        _img_tmp_warmed = True
    except Exception as _e:
        print(f'[warm_img_cache_bg] {_e}')


def _catalog_img_ver():
    """img_data 변경 시 localStorage 캐시를 깨기 위한 버전 해시."""
    try:
        conn = get_db()
        r = conn.execute(
            "SELECT COUNT(*) AS cnt, "
            "SUM(CASE WHEN img_data IS NOT NULL AND img_data != '' THEN 1 ELSE 0 END) AS custom_cnt, "
            "MAX(LENGTH(COALESCE(img_data,''))) AS max_len "
            "FROM catalog_defs"
        ).fetchone()
        conn.close()
        if r:
            return abs(hash((r['cnt'], r['custom_cnt'], r['max_len']))) % 10000000
    except Exception:
        pass
    return 0


def _invalidate_catalog_cache():
    pass  # 캐시 제거됨 — 매 요청 DB 직접 조회로 다중 인스턴스 정합성 보장


def _fast_schema_check(conn):
    """cold start에서 DDL 쿼리 5개 대신 SELECT 1개로 스키마 확인 (~150ms 절약).
    테이블/컬럼 존재 시 즉시 반환; 없을 때만 전체 DDL 실행 (첫 배포 시 1회)."""
    global _catalog_table_ready, _user_deleted_migrated
    if _catalog_table_ready and _user_deleted_migrated:
        return
    if USE_SQLITE:
        _catalog_table_ready = True
        _user_deleted_migrated = True
        return
    try:
        conn.execute("SELECT code, user_deleted FROM catalog_defs LIMIT 0")
        _catalog_table_ready  = True
        _user_deleted_migrated = True
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        _ensure_catalog_table()
        _ensure_user_deleted_col()


def _get_catalog_items(conn):
    """DB에서 카탈로그 아이템 목록 로드 (CAT_ORDER + sort_order 기준 정렬)."""
    try:
        rows = conn.execute(
            "SELECT code, img, name, cat, sub_desc, sort_order, "
            "CASE WHEN COALESCE(img_data,'')!='' THEN 1 ELSE 0 END AS is_custom "
            'FROM catalog_defs WHERE user_deleted IS NOT TRUE'
        ).fetchall()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        rows = conn.execute(
            "SELECT code, img, name, cat, sub_desc, sort_order, "
            "CASE WHEN COALESCE(img_data,'')!='' THEN 1 ELSE 0 END AS is_custom "
            'FROM catalog_defs'
        ).fetchall()
    items = [dict(r) for r in rows]
    cat_idx = {cat: i for i, cat in enumerate(CAT_ORDER)}
    items.sort(key=lambda x: (cat_idx.get(x['cat'], 999), x.get('sort_order', 0)))
    return items


def _get_cat_groups(conn, items=None):
    """카탈로그 아이템을 cat 별로 그룹화 (CAT_ORDER 정렬). {cat: [items]} OrderedDict 반환."""
    from collections import OrderedDict
    if items is None:
        items = _get_catalog_items(conn)
    groups = OrderedDict()
    for cat in CAT_ORDER:
        groups[cat] = []
    for item in items:
        cat = item['cat']
        if cat not in groups:
            groups[cat] = []
        groups[cat].append(item)
    return OrderedDict((k, v) for k, v in groups.items() if v)


# ── 사용자 관리 (관리자 전용) ─────────────────────────────────────────────────

def _ensure_user_deleted_col():
    """catalog_defs.user_deleted 컬럼 마이그레이션. _catalog_table_ready와 독립적으로 실행."""
    global _user_deleted_migrated
    if _user_deleted_migrated or USE_SQLITE:
        return
    _conn = None
    try:
        import psycopg2 as _pg2
        _conn = _pg2.connect(**_PG)
        _conn.autocommit = True
        cur = _conn.cursor()
        cur.execute(
            "ALTER TABLE catalog_defs ADD COLUMN IF NOT EXISTS user_deleted BOOLEAN NOT NULL DEFAULT FALSE"
        )
    except Exception as _e:
        print(f'[ensure_user_deleted_col] {_e}')
    finally:
        # 성공/실패 무관하게 반드시 닫음 (연결 풀 누수 방지)
        if _conn is not None:
            try:
                _conn.close()
            except Exception:
                pass
        # 재시도 루프 방지 — 실패해도 폴백 쿼리가 처리함
        _user_deleted_migrated = True


def _ensure_catalog_table():
    """catalog_branch_items + catalog_defs 테이블 보장 및 시딩.
    warm 인스턴스에서는 플래그로 DDL 재실행을 건너뜀."""
    global _catalog_table_ready
    if _catalog_table_ready:
        return
    if USE_SQLITE:
        # SQLite는 init_db에서 이미 생성됨 — 시딩만 확인
        try:
            conn = get_db()
            cnt = conn.execute('SELECT COUNT(*) AS cnt FROM catalog_defs').fetchone()['cnt']
            if cnt == 0:
                _seed_catalog_defs(conn)
            _catalog_table_ready = True
        except Exception as _e:
            print(f'[ensure_catalog_table/sqlite] {_e}')
        return
    try:
        conn = get_db()
        conn.execute('''
            CREATE TABLE IF NOT EXISTS catalog_branch_items (
                id         SERIAL PRIMARY KEY,
                branch_id  INTEGER NOT NULL REFERENCES branches(id),
                item_code  TEXT NOT NULL,
                quantity   INTEGER DEFAULT 1,
                updated_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(branch_id, item_code)
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS catalog_defs (
                code       TEXT PRIMARY KEY,
                img        TEXT NOT NULL,
                name       TEXT NOT NULL,
                cat        TEXT NOT NULL,
                sub_desc   TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                img_data   TEXT DEFAULT ''
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS catalog_requests (
                id            SERIAL PRIMARY KEY,
                branch_id     INTEGER NOT NULL REFERENCES branches(id),
                status        TEXT NOT NULL DEFAULT 'draft',
                notes         TEXT DEFAULT '',
                reject_reason TEXT DEFAULT '',
                created_at    TIMESTAMP DEFAULT NOW(),
                updated_at    TIMESTAMP DEFAULT NOW()
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS catalog_request_items (
                id              SERIAL PRIMARY KEY,
                request_id      INTEGER NOT NULL REFERENCES catalog_requests(id),
                item_code       TEXT NOT NULL,
                quantity        INTEGER DEFAULT 1,
                custom_img_data TEXT DEFAULT '',
                custom_text     TEXT DEFAULT '',
                item_name       TEXT DEFAULT '',
                item_cat        TEXT DEFAULT '',
                UNIQUE(request_id, item_code)
            )
        ''')
        conn.execute('ALTER TABLE catalog_request_items ADD COLUMN IF NOT EXISTS item_name TEXT DEFAULT \'\'')
        conn.execute('ALTER TABLE catalog_request_items ADD COLUMN IF NOT EXISTS item_cat  TEXT DEFAULT \'\'')
        conn.execute('ALTER TABLE catalog_defs ADD COLUMN IF NOT EXISTS user_deleted BOOLEAN NOT NULL DEFAULT FALSE')
        conn.commit()
        cnt = conn.execute('SELECT COUNT(*) AS cnt FROM catalog_defs').fetchone()['cnt']
        if cnt == 0:
            _seed_catalog_defs(conn)
        _catalog_table_ready = True
    except Exception as _e:
        print(f'[ensure_catalog_table] {_e}')
        try:
            get_db().rollback()
        except Exception:
            pass


@app.route('/catalog')
@login_required
def catalog():
    bid  = session.get('branch_id')
    role = session.get('role')

    conn = get_db()

    # cold start 스키마 확인 — DDL 5개 대신 SELECT 1개
    if not (_catalog_table_ready and _user_deleted_migrated):
        _fast_schema_check(conn)

    # 관리자만 지점 목록 필요
    branches = []
    if role == 'admin':
        branches = conn.execute(
            'SELECT id, code, name, type FROM branches ORDER BY type, code'
        ).fetchall()

    my_items = {}
    cart_cnt = 0
    if bid:
        # cart_cnt + my_items 단일 쿼리 — 2 round-trip → 1 round-trip
        combined = conn.execute(
            'SELECT cbi.item_code, cbi.quantity, cc.cart_cnt'
            ' FROM (SELECT COUNT(*) AS cart_cnt'
            '       FROM catalog_request_items ri'
            '       JOIN catalog_requests r ON r.id=ri.request_id'
            '       WHERE r.branch_id=%s AND r.status=%s) cc'
            ' LEFT JOIN catalog_branch_items cbi ON cbi.branch_id=%s',
            (bid, 'draft', bid)
        ).fetchall()
        cart_cnt = combined[0]['cart_cnt'] if combined else 0
        my_items = {r['item_code']: r['quantity'] for r in combined if r['item_code'] is not None}

    catalog_items = _get_catalog_items(conn)
    cat_groups    = _get_cat_groups(None, items=catalog_items)
    conn.close()

    if not _img_tmp_warmed and not USE_SQLITE:
        import threading
        threading.Thread(target=_warm_img_tmp_cache_bg, daemon=True).start()

    return render_template('catalog.html',
        catalog_items=catalog_items,
        cat_groups=cat_groups,
        my_items=my_items,
        branches=branches,
        cart_cnt=cart_cnt,
        custom_cats=['X-Banner', '스탠션'],
    )


@app.route('/catalog/imgs')
@login_required
def catalog_imgs():
    """카탈로그 이미지 URL 맵 반환. 커스텀 업로드만 base64, 나머지는 정적 URL."""
    import base64 as _b64
    _ensure_catalog_table()
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT img, img_data FROM catalog_defs WHERE user_deleted IS NOT TRUE"
        ).fetchall()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        rows = conn.execute("SELECT img, img_data FROM catalog_defs").fetchall()
    conn.close()
    data = {}
    for r in rows:
        if r['img_data']:
            data[r['img']] = 'data:image/png;base64,' + r['img_data']
        else:
            data[r['img']] = '/static/item_images/' + r['img'] + '.png'
    resp = jsonify(data)
    resp.headers['Cache-Control'] = 'public, max-age=3600'
    return resp


@app.route('/catalog/banner-editor')
@login_required
def catalog_banner_editor():
    """배너/스탠션 디자인 편집기 — 새 탭에서 열리고 결과를 localStorage로 전달."""
    item_code = request.args.get('item_code', '')
    item_name = request.args.get('item_name', '')
    item_cat  = request.args.get('item_cat', '')
    return render_template('catalog_banner_editor.html',
        item_code=item_code, item_name=item_name, item_cat=item_cat)


# ── 카탈로그 신청 헬퍼 ────────────────────────────────────────────────────────

def _get_or_create_draft(conn, branch_id):
    """branch_id의 현재 draft 요청을 반환하거나 새로 생성."""
    ph = '%s' if not USE_SQLITE else '?'
    row = conn.execute(
        f'SELECT id FROM catalog_requests WHERE branch_id={ph} AND status={ph}',
        (branch_id, 'draft')
    ).fetchone()
    if row:
        return row['id']
    conn.execute(
        f'INSERT INTO catalog_requests (branch_id, status) VALUES ({ph},{ph})',
        (branch_id, 'draft')
    )
    conn.commit()
    return conn.execute(
        f'SELECT id FROM catalog_requests WHERE branch_id={ph} AND status={ph}',
        (branch_id, 'draft')
    ).fetchone()['id']


def _cart_count(conn, branch_id):
    """현재 draft 장바구니 아이템 수."""
    ph = '%s' if not USE_SQLITE else '?'
    row = conn.execute(
        f'SELECT COUNT(*) AS cnt FROM catalog_request_items ri '
        f'JOIN catalog_requests r ON r.id=ri.request_id '
        f'WHERE r.branch_id={ph} AND r.status={ph}',
        (branch_id, 'draft')
    ).fetchone()
    return row['cnt'] if row else 0


# ── 카탈로그 신청: 장바구니 업데이트 (AJAX) ──────────────────────────────────

@app.route('/catalog/cart/update', methods=['POST'])
@login_required
def catalog_cart_update():
    _ensure_catalog_table()
    bid  = session.get('branch_id')
    role = session.get('role')
    data = request.get_json(silent=True) or {}

    target_bid = int(data.get('branch_id', bid or 0)) if role == 'admin' else bid
    if not target_bid:
        return jsonify({'ok': False, 'msg': '지점 정보 없음'}), 400

    item_code   = (data.get('item_code') or '').strip()
    qty         = max(0, int(data.get('quantity', 1) or 1))
    custom_img  = (data.get('custom_img_data') or '').strip()
    custom_text = (data.get('custom_text') or '').strip()
    item_name   = (data.get('item_name') or '').strip()
    item_cat    = (data.get('item_cat') or '').strip()
    action      = data.get('action', 'add')   # add | remove | clear

    conn = get_db()
    ph   = '%s' if not USE_SQLITE else '?'

    if action == 'clear':
        req_row = conn.execute(
            f'SELECT id FROM catalog_requests WHERE branch_id={ph} AND status={ph}',
            (target_bid, 'draft')
        ).fetchone()
        if req_row:
            conn.execute(f'DELETE FROM catalog_request_items WHERE request_id={ph}', (req_row['id'],))
            conn.execute(f'DELETE FROM catalog_requests WHERE id={ph}', (req_row['id'],))
            conn.commit()
        conn.close()
        return jsonify({'ok': True, 'cart_count': 0})

    if not item_code:
        conn.close()
        return jsonify({'ok': False, 'msg': '아이템 코드 없음'}), 400

    if action == 'remove' or qty == 0:
        req_row = conn.execute(
            f'SELECT id FROM catalog_requests WHERE branch_id={ph} AND status={ph}',
            (target_bid, 'draft')
        ).fetchone()
        if req_row:
            conn.execute(
                f'DELETE FROM catalog_request_items WHERE request_id={ph} AND item_code={ph}',
                (req_row['id'], item_code)
            )
            conn.commit()
    else:
        req_id = _get_or_create_draft(conn, target_bid)
        if USE_SQLITE:
            conn.execute('''
                INSERT INTO catalog_request_items
                    (request_id,item_code,quantity,custom_img_data,custom_text,item_name,item_cat)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(request_id,item_code) DO UPDATE
                SET quantity=excluded.quantity, custom_img_data=excluded.custom_img_data,
                    custom_text=excluded.custom_text, item_name=excluded.item_name,
                    item_cat=excluded.item_cat
            ''', (req_id, item_code, qty, custom_img, custom_text, item_name, item_cat))
        else:
            conn.execute('''
                INSERT INTO catalog_request_items
                    (request_id,item_code,quantity,custom_img_data,custom_text,item_name,item_cat)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT(request_id,item_code) DO UPDATE
                SET quantity=EXCLUDED.quantity, custom_img_data=EXCLUDED.custom_img_data,
                    custom_text=EXCLUDED.custom_text, item_name=EXCLUDED.item_name,
                    item_cat=EXCLUDED.item_cat
            ''', (req_id, item_code, qty, custom_img, custom_text, item_name, item_cat))
        conn.commit()

    cnt = _cart_count(conn, target_bid)
    conn.close()
    return jsonify({'ok': True, 'cart_count': cnt})


# ── 카탈로그 신청: 장바구니 조회 ──────────────────────────────────────────────

@app.route('/catalog/cart')
@login_required
def catalog_cart():
    _ensure_catalog_table()
    bid  = session.get('branch_id')
    role = session.get('role')
    target_bid = int(request.args.get('branch_id', bid or 0)) if role == 'admin' else bid
    if not target_bid:
        return render_template('catalog_request.html',
            cart_items=[], branches=[], target_bid=None,
            catalog_items=[], custom_cats=['X-Banner','스탠션'])

    conn = get_db()
    ph   = '%s' if not USE_SQLITE else '?'
    req_row = conn.execute(
        f'SELECT id FROM catalog_requests WHERE branch_id={ph} AND status={ph}',
        (target_bid, 'draft')
    ).fetchone()

    cart_items = []
    if req_row:
        rows = conn.execute(
            f'SELECT ri.*, COALESCE(cd.name, ri.item_name) AS name, '
            f'COALESCE(cd.cat, ri.item_cat) AS cat, cd.img '
            f'FROM catalog_request_items ri '
            f'LEFT JOIN catalog_defs cd ON cd.code=ri.item_code '
            f'WHERE ri.request_id={ph}',
            (req_row['id'],)
        ).fetchall()
        cart_items = [dict(r) for r in rows]

    branches = conn.execute('SELECT id,code,name FROM branches ORDER BY type,code').fetchall()
    catalog_items = _get_catalog_items(conn)
    conn.close()

    return render_template('catalog_request.html',
        cart_items=cart_items,
        branches=branches,
        target_bid=target_bid,
        catalog_items=catalog_items,
        custom_cats=['X-Banner', '스탠션'],
    )


# ── 카탈로그 신청: 제출 ────────────────────────────────────────────────────────

@app.route('/catalog/request/submit', methods=['POST'])
@login_required
def catalog_request_submit():
    _ensure_catalog_table()
    bid  = session.get('branch_id')
    role = session.get('role')
    data = request.get_json(silent=True) or {}
    target_bid = int(data.get('branch_id', bid or 0)) if role == 'admin' else bid
    notes = (data.get('notes') or '').strip()

    if not target_bid:
        return jsonify({'ok': False, 'msg': '지점 정보 없음'}), 400

    conn = get_db()
    ph   = '%s' if not USE_SQLITE else '?'
    req_row = conn.execute(
        f'SELECT id FROM catalog_requests WHERE branch_id={ph} AND status={ph}',
        (target_bid, 'draft')
    ).fetchone()
    if not req_row:
        conn.close()
        return jsonify({'ok': False, 'msg': '장바구니가 비어있습니다.'}), 400

    cnt = conn.execute(
        f'SELECT COUNT(*) AS cnt FROM catalog_request_items WHERE request_id={ph}',
        (req_row['id'],)
    ).fetchone()['cnt']
    if cnt == 0:
        conn.close()
        return jsonify({'ok': False, 'msg': '장바구니에 아이템이 없습니다.'}), 400

    now_sql = 'NOW()' if not USE_SQLITE else "datetime('now')"
    conn.execute(
        f"UPDATE catalog_requests SET status={ph}, notes={ph}, updated_at={now_sql} WHERE id={ph}",
        ('pending', notes, req_row['id'])
    )
    # 관리자에게 알림
    admin_rows = conn.execute(
        f"SELECT id FROM branches WHERE code='관리자' UNION "
        f"SELECT DISTINCT branch_id FROM users WHERE role='admin'"
    ).fetchall() if not USE_SQLITE else conn.execute(
        "SELECT DISTINCT branch_id id FROM users WHERE role='admin'"
    ).fetchall()
    branch_name_row = conn.execute(
        f'SELECT name FROM branches WHERE id={ph}', (target_bid,)
    ).fetchone()
    branch_name = branch_name_row['name'] if branch_name_row else str(target_bid)
    for ar in admin_rows:
        try:
            conn.execute(
                f'INSERT INTO notifications (branch_id, message) VALUES ({ph},{ph})',
                (ar['id'], f'[카탈로그 신청] {branch_name} 지점에서 신청서가 접수되었습니다.')
            )
        except Exception:
            pass
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'msg': '신청이 접수되었습니다.'})


# ── 카탈로그 신청: 내역 조회 (지점) ──────────────────────────────────────────

@app.route('/catalog/requests')
@login_required
def catalog_requests_branch():
    _ensure_catalog_table()
    bid  = session.get('branch_id')
    role = session.get('role')
    if not bid and role != 'admin':
        return redirect(url_for('catalog'))

    conn = get_db()
    ph   = '%s' if not USE_SQLITE else '?'
    if role == 'admin':
        reqs = conn.execute(
            'SELECT r.*, b.name branch_name, b.code branch_code '
            'FROM catalog_requests r JOIN branches b ON b.id=r.branch_id '
            "WHERE r.status != 'draft' ORDER BY r.updated_at DESC"
        ).fetchall()
    else:
        reqs = conn.execute(
            f'SELECT r.*, b.name branch_name, b.code branch_code '
            f'FROM catalog_requests r JOIN branches b ON b.id=r.branch_id '
            f"WHERE r.branch_id={ph} AND r.status!='draft' ORDER BY r.updated_at DESC",
            (bid,)
        ).fetchall()
    conn.close()
    return render_template('catalog_requests_list.html', requests=reqs)


# ── 카탈로그 신청: 상세 조회 ──────────────────────────────────────────────────

@app.route('/catalog/requests/<int:req_id>')
@login_required
def catalog_request_detail(req_id):
    _ensure_catalog_table()
    bid  = session.get('branch_id')
    role = session.get('role')
    conn = get_db()
    ph   = '%s' if not USE_SQLITE else '?'
    req  = conn.execute(
        f'SELECT r.*, b.name branch_name, b.code branch_code '
        f'FROM catalog_requests r JOIN branches b ON b.id=r.branch_id '
        f'WHERE r.id={ph}', (req_id,)
    ).fetchone()
    if not req:
        conn.close()
        flash('신청서를 찾을 수 없습니다.', 'danger')
        return redirect(url_for('catalog_requests_branch'))
    if role != 'admin' and req['branch_id'] != bid:
        conn.close()
        flash('권한이 없습니다.', 'danger')
        return redirect(url_for('catalog_requests_branch'))

    items = conn.execute(
        f'SELECT ri.*, COALESCE(cd.name, ri.item_name) item_name, '
        f'COALESCE(cd.cat, ri.item_cat) cat, cd.img '
        f'FROM catalog_request_items ri '
        f'LEFT JOIN catalog_defs cd ON cd.code=ri.item_code '
        f'WHERE ri.request_id={ph}', (req_id,)
    ).fetchall()
    conn.close()
    return render_template('catalog_request_detail.html',
        req=dict(req), items=[dict(i) for i in items])


# ── 카탈로그 신청: 관리자 액션 (승인/반려/대기) ───────────────────────────────

@app.route('/catalog/requests/<int:req_id>/action', methods=['POST'])
@login_required
def catalog_request_action(req_id):
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'msg': '권한 없음'}), 403
    _ensure_catalog_table()
    data   = request.get_json(silent=True) or {}
    action = data.get('action', '')           # approved | rejected | on_hold | pending
    reason = (data.get('reason') or '').strip()

    if action not in ('approved', 'rejected', 'on_hold', 'pending'):
        return jsonify({'ok': False, 'msg': '유효하지 않은 액션'}), 400
    if action == 'rejected' and not reason:
        return jsonify({'ok': False, 'msg': '반려 사유를 입력해 주세요.'}), 400

    conn = get_db()
    ph   = '%s' if not USE_SQLITE else '?'
    req  = conn.execute(
        f'SELECT * FROM catalog_requests WHERE id={ph}', (req_id,)
    ).fetchone()
    if not req:
        conn.close()
        return jsonify({'ok': False, 'msg': '신청서 없음'}), 404

    now_sql = 'NOW()' if not USE_SQLITE else "datetime('now')"
    conn.execute(
        f"UPDATE catalog_requests SET status={ph}, reject_reason={ph}, updated_at={now_sql} WHERE id={ph}",
        (action, reason if action == 'rejected' else '', req_id)
    )

    # 지점에 알림
    status_kr = {'approved':'승인','rejected':'반려','on_hold':'보류','pending':'검토중'}
    msg = f'[카탈로그 신청] 신청서가 {status_kr.get(action, action)} 처리되었습니다.'
    if action == 'rejected' and reason:
        msg += f' 사유: {reason}'
    try:
        conn.execute(
            f'INSERT INTO notifications (branch_id, message) VALUES ({ph},{ph})',
            (req['branch_id'], msg)
        )
    except Exception:
        pass
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'status': action})


@app.route('/catalog/save', methods=['POST'])
@login_required
def catalog_save():
    _ensure_catalog_table()
    data     = request.get_json(silent=True) or {}
    bid      = session.get('branch_id')
    role     = session.get('role')
    # JSON의 null/NaN은 Python None → or 연산으로 세션 branch_id fallback
    _raw     = data.get('branch_id')
    target   = int(_raw) if _raw is not None else bid

    if role != 'admin' and target != bid:
        return jsonify({'ok': False, 'msg': '권한 없음'}), 403
    if not target:
        return jsonify({'ok': False, 'msg': '지점이 배정되지 않았습니다. 관리자에게 문의하세요.'}), 400

    items = data.get('items', {})   # {item_code: quantity}

    conn  = get_db()
    for code, qty in items.items():
        qty = max(0, int(qty))
        if qty == 0:
            conn.execute(
                'DELETE FROM catalog_branch_items WHERE branch_id=%s AND item_code=%s',
                (target, code)
            )
        else:
            conn.execute('''
                INSERT INTO catalog_branch_items (branch_id, item_code, quantity, updated_at)
                VALUES (%s, %s, %s, NOW())
                ON CONFLICT (branch_id, item_code)
                DO UPDATE SET quantity=%s, updated_at=NOW()
            ''', (target, code, qty, qty))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/catalog/inventory')
@login_required
def catalog_inventory():
    _ensure_catalog_table()
    conn = get_db()
    bid  = session.get('branch_id')
    role = session.get('role')

    if role == 'admin':
        rows = conn.execute('''
            SELECT b.id, b.code, b.name, b.type,
                   cbi.item_code, cbi.quantity
            FROM branches b
            LEFT JOIN catalog_branch_items cbi ON b.id = cbi.branch_id
            ORDER BY b.type, b.code, cbi.item_code
        ''').fetchall()
    else:
        rows = conn.execute('''
            SELECT b.id, b.code, b.name, b.type,
                   cbi.item_code, cbi.quantity
            FROM branches b
            LEFT JOIN catalog_branch_items cbi ON b.id = cbi.branch_id
            WHERE b.id=%s
            ORDER BY cbi.item_code
        ''', (bid,)).fetchall()

    from collections import defaultdict
    branch_map  = {}
    branch_order = []
    for r in rows:
        bc = r['code']
        if bc not in branch_map:
            branch_map[bc] = {'name': r['name'], 'type': r['type'], 'items': {}}
            branch_order.append(bc)
        if r['item_code']:
            branch_map[bc]['items'][r['item_code']] = r['quantity']

    catalog_items = _get_catalog_items(conn)
    conn.close()

    _CATALOG_ORDER = [
        'GMP','CJU','CJJ','PUS','ICN',
        'NRT','KIX','FUK','CTS','OKA','KMJ','TKS',
        'BKK','CNX','DAD','CXR','PQC','MDC',
        'TSA','TPE','PVG','YNJ','CGO','YNT','HKG','ALA',
    ]
    _exclude = {'광명역', '서울역', '이지드랍', 'CARGO'}
    branch_order = [bc for bc in branch_order if bc not in _exclude]
    branch_order.sort(key=lambda bc: _CATALOG_ORDER.index(bc) if bc in _CATALOG_ORDER else 999)

    return render_template('catalog_inventory.html',
        catalog_items=catalog_items,
        branch_map=branch_map,
        branch_order=branch_order,
    )


# ── 관리자 카탈로그 편집 ───────────────────────────────────────────────────────

_STATIC_IMG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'item_images')
_TMP_IMG_DIR    = '/tmp/item_images'


@app.route('/ci/<filename>')
def catalog_img(filename):
    """카탈로그 이미지 서빙: /tmp → static → DB(base64) 순 탐색.
    DB에서 읽은 이미지는 /tmp에 캐시해 다음 요청부터 DB 조회를 건너뜀."""
    import re as _re, base64 as _b64
    if not _re.match(r'^[\w\-]+\.(?:png|jpg|jpeg|webp)$', filename, _re.IGNORECASE):
        return '', 404

    _CC = 'public, max-age=86400, stale-while-revalidate=604800'

    # 1) /tmp 확인
    tmp_path = os.path.join(_TMP_IMG_DIR, filename)
    if os.path.isfile(tmp_path):
        resp = send_file(tmp_path)
        resp.headers['Cache-Control'] = _CC
        return resp
    # 2) static 확인
    static_path = os.path.join(_STATIC_IMG_DIR, filename)
    if os.path.isfile(static_path):
        from flask import send_from_directory
        resp = send_from_directory(_STATIC_IMG_DIR, filename)
        resp.headers['Cache-Control'] = _CC
        return resp
    # 3) DB img_data (base64) 폴백 — Vercel cold start 후에도 서빙 가능
    try:
        stem = os.path.splitext(filename)[0]
        conn = get_db()
        row  = conn.execute(
            'SELECT img_data FROM catalog_defs WHERE img=%s', (stem,)
        ).fetchone()
        conn.close()
        if row and row['img_data']:
            img_bytes = _b64.b64decode(row['img_data'])
            # warm 인스턴스 내 /tmp 캐시 — 다음 요청부터 DB 조회 생략
            try:
                os.makedirs(_TMP_IMG_DIR, exist_ok=True)
                with open(tmp_path, 'wb') as _f:
                    _f.write(img_bytes)
            except Exception:
                pass
            resp = send_file(BytesIO(img_bytes), mimetype='image/png')
            resp.headers['Cache-Control'] = _CC
            return resp
    except Exception:
        pass
    return '', 404


@app.template_global()
def cat_img_url(img_stem):
    """카탈로그 이미지 URL — /ci/ 라우트 경유 (tmp 업로드 이미지 포함 모두 처리)"""
    return url_for('catalog_img', filename=img_stem + '.png')


@app.route('/admin/catalog/edit')
@login_required
def catalog_edit():
    """카탈로그 편집 페이지 제거 — 편집 기능은 /catalog 에서 직접 제공."""
    return redirect(url_for('catalog'))


@app.route('/admin/catalog/add', methods=['POST'])
@login_required
def catalog_add():
    is_ajax = request.form.get('_ajax') == '1'
    def _fail(msg):
        if is_ajax:
            return jsonify({'ok': False, 'msg': msg})
        flash(msg, 'danger')
        return redirect(url_for('catalog'))

    if session.get('role') != 'admin':
        return _fail('관리자 권한이 필요합니다.')
    _ensure_catalog_table()

    name     = request.form.get('name', '').strip()
    cat      = request.form.get('cat', '').strip()
    sub_desc = request.form.get('sub_desc', '').strip()
    custom_code = request.form.get('code', '').strip()

    if not name or not cat:
        return _fail('아이템명과 카테고리는 필수입니다.')

    f = request.files.get('image')
    if not f or not f.filename:
        return _fail('이미지 파일을 선택해 주세요.')

    # 확장자는 원본 파일명에서 직접 추출 (secure_filename은 한글 등 비ASCII 제거로 확장자 오파싱 가능)
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.png', '.jpg', '.jpeg', '.webp'):
        return _fail('PNG / JPG / WEBP 형식만 허용됩니다.')

    import uuid
    uid       = uuid.uuid4().hex[:10]
    img_stem  = f'adm_{uid}'
    save_name = img_stem + '.png'

    # 저장 폴더: static/item_images가 쓰기 가능하면 우선, 아니면 /tmp/item_images (Vercel 등)
    if os.path.isdir(_STATIC_IMG_DIR) and os.access(_STATIC_IMG_DIR, os.W_OK):
        save_dir = _STATIC_IMG_DIR
    else:
        os.makedirs(_TMP_IMG_DIR, exist_ok=True)
        save_dir = _TMP_IMG_DIR

    save_path = os.path.join(save_dir, save_name)
    try:
        from PIL import Image as _PILImage
        _img = _PILImage.open(f.stream)
        _bg  = _PILImage.new('RGB', _img.size, (255, 255, 255))
        if _img.mode in ('RGBA', 'LA', 'P'):
            _img = _img.convert('RGBA')
            _bg.paste(_img, mask=_img.split()[3])
        else:
            _bg.paste(_img.convert('RGB'))
        _bg.save(save_path, 'PNG')
    except Exception:
        f.stream.seek(0)
        f.save(save_path)

    # 이미지를 base64로 DB에 저장 — Vercel cold start 후 /tmp가 날아가도 서빙 가능
    import base64 as _b64
    try:
        with open(save_path, 'rb') as _fh:
            img_data_b64 = _b64.b64encode(_fh.read()).decode()
    except Exception:
        img_data_b64 = ''

    code = custom_code if custom_code else f'adm_{uid}'
    # Validate code uniqueness
    conn = get_db()
    existing = conn.execute('SELECT code FROM catalog_defs WHERE code=%s', (code,)).fetchone()
    if existing:
        conn.close()
        return _fail(f'코드 "{code}" 가 이미 사용 중입니다.')

    max_sort = conn.execute(
        'SELECT COALESCE(MAX(sort_order), 0) AS m FROM catalog_defs WHERE cat=%s', (cat,)
    ).fetchone()['m']

    conn.execute(
        'INSERT INTO catalog_defs (code, img, name, cat, sub_desc, sort_order, img_data) '
        'VALUES (%s, %s, %s, %s, %s, %s, %s)',
        (code, img_stem, name, cat, sub_desc, max_sort + 10, img_data_b64)
    )
    conn.commit()
    conn.close()
    _invalidate_catalog_cache()
    if is_ajax:
        return jsonify({'ok': True, 'msg': f'"{name}" 아이템이 추가되었습니다.'})
    flash(f'"{name}" 아이템이 추가되었습니다.', 'success')
    return redirect(url_for('catalog'))


@app.route('/admin/catalog/update', methods=['POST'])
@login_required
def catalog_update():
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'msg': '권한 없음'}), 403
    _ensure_catalog_table()

    code     = request.form.get('code', '').strip()
    new_code = request.form.get('new_code', '').strip() or code
    name     = request.form.get('name', '').strip()
    cat      = request.form.get('cat', '').strip()
    sub_desc = request.form.get('sub_desc', '').strip()
    try:
        sort_order = int(request.form.get('sort_order', 0) or 0)
    except (ValueError, TypeError):
        sort_order = 0

    if not code or not new_code or not name or not cat:
        return jsonify({'ok': False, 'msg': '필수 항목 누락'}), 400

    conn = get_db()
    existing = conn.execute(
        'SELECT img, img_data, sort_order FROM catalog_defs WHERE code=%s', (code,)
    ).fetchone()
    if not existing:
        conn.close()
        return jsonify({'ok': False, 'msg': '아이템을 찾을 수 없습니다.'}), 404

    # 코드 변경 시 중복 확인 + branch_items 연동
    if new_code != code:
        dup = conn.execute('SELECT code FROM catalog_defs WHERE code=%s', (new_code,)).fetchone()
        if dup:
            conn.close()
            return jsonify({'ok': False, 'msg': f'코드 "{new_code}" 가 이미 사용 중입니다.'})
        conn.execute(
            'UPDATE catalog_branch_items SET item_code=%s WHERE item_code=%s', (new_code, code)
        )

    img_stem     = existing['img']
    img_data_b64 = existing['img_data'] or ''
    new_img_url  = None

    f = request.files.get('image')
    if f and f.filename:
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.png', '.jpg', '.jpeg', '.webp'):
            conn.close()
            return jsonify({'ok': False, 'msg': 'PNG / JPG / WEBP 형식만 허용됩니다.'}), 400

        import uuid as _uuid
        uid       = _uuid.uuid4().hex[:10]
        img_stem  = f'adm_{uid}'
        save_name = img_stem + '.png'

        if os.path.isdir(_STATIC_IMG_DIR) and os.access(_STATIC_IMG_DIR, os.W_OK):
            save_dir = _STATIC_IMG_DIR
        else:
            os.makedirs(_TMP_IMG_DIR, exist_ok=True)
            save_dir = _TMP_IMG_DIR

        save_path = os.path.join(save_dir, save_name)
        try:
            from PIL import Image as _PILImage
            _img = _PILImage.open(f.stream)
            _bg  = _PILImage.new('RGB', _img.size, (255, 255, 255))
            if _img.mode in ('RGBA', 'LA', 'P'):
                _img = _img.convert('RGBA')
                _bg.paste(_img, mask=_img.split()[3])
            else:
                _bg.paste(_img.convert('RGB'))
            _bg.save(save_path, 'PNG')
        except Exception:
            f.stream.seek(0)
            f.save(save_path)

        import base64 as _b64
        try:
            with open(save_path, 'rb') as _fh:
                img_data_b64 = _b64.b64encode(_fh.read()).decode()
        except Exception:
            img_data_b64 = ''

        new_img_url = url_for('catalog_img', filename=img_stem + '.png')

    conn.execute(
        'UPDATE catalog_defs SET code=%s, name=%s, cat=%s, sub_desc=%s, '
        'img=%s, img_data=%s, sort_order=%s WHERE code=%s',
        (new_code, name, cat, sub_desc, img_stem, img_data_b64, sort_order, code)
    )
    conn.commit()
    conn.close()
    _invalidate_catalog_cache()
    resp = {'ok': True, 'sub_desc': sub_desc, 'new_code': new_code}
    if new_img_url:
        resp['img_url'] = new_img_url
    return jsonify(resp)


@app.route('/admin/catalog/delete', methods=['POST'])
@login_required
def catalog_delete():
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'msg': '권한 없음'}), 403
    _ensure_catalog_table()
    code = request.get_json(silent=True, force=True).get('code', '') if request.is_json else request.form.get('code', '')
    if not code:
        return jsonify({'ok': False, 'msg': '코드 없음'}), 400
    conn = get_db()
    conn.execute('UPDATE catalog_defs SET user_deleted=TRUE WHERE code=%s', (code,))
    conn.execute('DELETE FROM catalog_branch_items WHERE item_code=%s', (code,))
    conn.commit()
    conn.close()
    _invalidate_catalog_cache()
    return jsonify({'ok': True})


@app.route('/admin/users')
@login_required
def manage_users():
    if session.get('role') != 'admin':
        flash('관리자 권한이 필요합니다.', 'danger')
        return redirect(url_for('dashboard'))
    conn = get_db()
    users    = conn.execute('''
        SELECT u.*, b.name branch_name, b.code branch_code
        FROM users u LEFT JOIN branches b ON u.branch_id = b.id
        ORDER BY u.role DESC, u.username
    ''').fetchall()
    branches = conn.execute('SELECT * FROM branches ORDER BY type, code').fetchall()
    conn.close()
    return render_template('users.html', users=users, branches=branches)


@app.route('/admin/users/create', methods=['POST'])
@login_required
def create_user():
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'error': '권한 없음'}), 403

    username  = request.form.get('username', '').strip()
    password  = request.form.get('password', '').strip()
    branch_id = request.form.get('branch_id') or None
    role      = request.form.get('role', 'staff')

    if not username or not password:
        flash('아이디와 비밀번호를 입력하세요.', 'danger')
        return redirect(url_for('manage_users'))
    pw_err = validate_password(password)
    if pw_err:
        flash(pw_err, 'danger')
        return redirect(url_for('manage_users'))

    conn = get_db()
    try:
        conn.execute(
            'INSERT INTO users (username, password, branch_id, role) VALUES (%s,%s,%s,%s)',
            (username, hash_pw(password), branch_id, role)
        )
        conn.commit()
        flash(f'계정 [{username}] 이(가) 생성되었습니다.', 'success')
    except (psycopg2.errors.UniqueViolation, sqlite3.IntegrityError):
        conn.rollback()
        flash(f'이미 존재하는 아이디입니다: {username}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('manage_users'))


@app.route('/admin/users/<int:uid>/delete', methods=['POST'])
@login_required
def delete_user(uid):
    if session.get('role') != 'admin':
        return jsonify({'ok': False}), 403
    if uid == session.get('user_id'):
        flash('본인 계정은 삭제할 수 없습니다.', 'danger')
        return redirect(url_for('manage_users'))
    conn = get_db()
    conn.execute('DELETE FROM users WHERE id=%s', (uid,))
    conn.commit()
    conn.close()
    flash('계정이 삭제되었습니다.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/admin/users/<int:uid>/reset-password', methods=['POST'])
@login_required
def reset_user_password(uid):
    if session.get('role') != 'admin':
        return jsonify({'ok': False}), 403
    new_pw = request.form.get('new_password', '').strip()
    pw_err = validate_password(new_pw)
    if pw_err:
        flash(pw_err, 'danger')
        return redirect(url_for('manage_users'))
    conn = get_db()
    conn.execute('UPDATE users SET password=%s, password_changed_at=NOW() WHERE id=%s', (hash_pw(new_pw), uid))
    conn.commit()
    conn.close()
    flash('비밀번호가 초기화되었습니다.', 'success')
    return redirect(url_for('manage_users'))


# ── 지점 이메일 관리 ──────────────────────────────────────────────────────────

@app.route('/admin/branches/<int:bid>/update-email', methods=['POST'])
@login_required
def update_branch_email(bid):
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'error': '권한 없음'}), 403
    email = request.form.get('email', '').strip()
    conn = get_db()
    conn.execute('UPDATE branches SET email=%s WHERE id=%s', (email or None, bid))
    conn.commit()
    conn.close()
    flash('이메일이 저장되었습니다.', 'success')
    return redirect(url_for('manage_users') + '#branch-emails')


@app.route('/admin/test-email')
@login_required
def test_email():
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'error': '권한 없음'}), 403
    to = request.args.get('to', '').strip()
    if not to:
        return jsonify({'ok': False, 'error': 'to 파라미터 필요. 예: /admin/test-email?to=xxx@xxx.com'})
    if not MAIL_HOST:
        return jsonify({'ok': False, 'error': 'MAIL_HOST 환경변수 미설정'})
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = '[ZEGO] 이메일 발송 테스트'
        msg['From']    = MAIL_FROM
        msg['To']      = to
        msg.attach(MIMEText('ZEGO 이메일 발송 테스트입니다.', 'plain', 'utf-8'))
        with smtplib.SMTP(MAIL_HOST, MAIL_PORT, timeout=10) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(MAIL_USER, MAIL_PASS)
            server.sendmail(MAIL_FROM, [to], msg.as_string())
        return jsonify({'ok': True, 'message': f'{to} 로 발송 성공'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ── 비밀번호 변경 (본인) ───────────────────────────────────────────────────────

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        cur_pw  = request.form.get('current_password', '').strip()
        new_pw  = request.form.get('new_password', '').strip()
        conf_pw = request.form.get('confirm_password', '').strip()

        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE id=%s', (session['user_id'],)).fetchone()

        if not verify_pw(cur_pw, user['password']):
            flash('현재 비밀번호가 올바르지 않습니다.', 'danger')
            conn.close()
            return redirect(url_for('change_password'))
        pw_err = validate_password(new_pw)
        if pw_err:
            flash(pw_err, 'danger')
            conn.close()
            return redirect(url_for('change_password'))
        if new_pw != conf_pw:
            flash('새 비밀번호가 일치하지 않습니다.', 'danger')
            conn.close()
            return redirect(url_for('change_password'))

        conn.execute(
            'UPDATE users SET password=%s, password_changed_at=NOW() WHERE id=%s',
            (hash_pw(new_pw), session['user_id'])
        )
        conn.commit()
        conn.close()
        session['_pwd_changed_at'] = datetime.now(timezone.utc).timestamp()
        log_action('비밀번호변경')
        flash('비밀번호가 변경되었습니다.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html')


# ── 입고 엑셀 템플릿 다운로드 ────────────────────────────────────────────────

@app.route('/inbound/template')
@login_required
def inbound_template():
    conn = get_db()
    branches   = conn.execute('SELECT code, name, type FROM branches ORDER BY type, code').fetchall()
    form_types = conn.execute('SELECT name, unit, unit_detail FROM form_types WHERE is_active ORDER BY sort_order').fetchall()

    my_branch_code = ''
    if session.get('role') != 'admin' and session.get('branch_id'):
        row = conn.execute('SELECT code FROM branches WHERE id=%s', (session['branch_id'],)).fetchone()
        my_branch_code = row['code'] if row else ''
    conn.close()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = '입고업로드'
    hfill, hfont, halign = _make_header_style()
    bd  = _border()

    headers    = ['지점코드', '양식명', '수량', '비고']
    col_widths = [14, 38, 10, 26]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill, cell.font, cell.alignment = hfill, hfont, halign
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    guide = '※ 지점코드·양식명은 드롭다운 또는 "참고" 시트를 확인하세요. 수량은 BOX/포대/권 단위입니다.'
    ws.merge_cells('A2:D2')
    c = ws.cell(row=2, column=1, value=guide)
    c.font      = Font(color='5A6A8A', size=9, italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill      = PatternFill('solid', fgColor='EEF2FF')
    ws.row_dimensions[2].height = 18

    is_staff   = session.get('role') != 'admin'
    gray_fill  = PatternFill('solid', fgColor='F1F5F9')
    data_start = 3

    for r in range(data_start, data_start + 50):
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.border    = bd
            cell.alignment = Alignment(vertical='center',
                                       horizontal='center' if col == 3 else 'left')
        if is_staff and my_branch_code:
            c = ws.cell(row=r, column=1, value=my_branch_code)
            c.fill      = gray_fill
            c.font      = Font(color='374151', bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 17

    last_branch = len(branches) + 1
    last_form   = len(form_types) + 1
    sqref       = f'A{data_start}:A{data_start+49}'
    sqref_form  = f'B{data_start}:B{data_start+49}'

    if not is_staff:
        dv_b = DataValidation(type='list',
                              formula1=f"'참고'!$A$2:$A${last_branch}",
                              allow_blank=True,
                              showErrorMessage=True,
                              error='참고 시트의 지점코드 목록에서 선택하세요.',
                              errorTitle='잘못된 지점코드')
        dv_b.sqref = sqref
        ws.add_data_validation(dv_b)

    dv_f = DataValidation(type='list',
                          formula1=f"'참고'!$C$2:$C${last_form}",
                          allow_blank=True,
                          showErrorMessage=True,
                          error='참고 시트의 양식명 목록에서 선택하세요.',
                          errorTitle='잘못된 양식명')
    dv_f.sqref = sqref_form
    ws.add_data_validation(dv_f)

    ws.freeze_panes = 'A3'

    ws2 = wb.create_sheet('참고')
    ref_headers = [('A', '지점코드', 12), ('B', '지점명', 16), ('C', '양식명', 38),
                   ('D', '단위', 10), ('E', '단위상세', 14)]
    for col_letter, h, w in ref_headers:
        c = ws2[f'{col_letter}1']
        c.value, c.fill, c.font, c.alignment = h, hfill, hfont, halign
        ws2.column_dimensions[col_letter].width = w
    ws2.row_dimensions[1].height = 22

    type_kr = {'DOM': '국내', 'INTL': '국제', 'CARGO': '화물'}
    for i, b in enumerate(branches, 2):
        ws2.cell(row=i, column=1, value=b['code']).alignment = Alignment(horizontal='center')
        type_label = type_kr.get(b['type'], b['type'])
        c = ws2.cell(row=i, column=2, value=f"[{type_label}] {b['name']}")
        c.alignment = Alignment(horizontal='left')

    for i, f in enumerate(form_types, 2):
        ws2.cell(row=i, column=3, value=f['name'])
        ws2.cell(row=i, column=4, value=f['unit']).alignment = Alignment(horizontal='center')
        ws2.cell(row=i, column=5, value=f['unit_detail']).alignment = Alignment(horizontal='center')

    ws2.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import date
    suffix = f'_{my_branch_code}' if my_branch_code else ''
    fname  = f'입고업로드_템플릿{suffix}_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 입고 엑셀 업로드 처리 ──────────────────────────────────────────────────────

@app.route('/inbound/upload', methods=['POST'])
@login_required
def inbound_upload():
    f = request.files.get('file')
    if not f or not f.filename:
        flash('파일을 선택해주세요.', 'danger')
        return redirect(url_for('inbound'))
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        flash('xlsx 또는 xls 파일만 업로드 가능합니다.', 'danger')
        return redirect(url_for('inbound'))

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'파일을 읽을 수 없습니다: {e}', 'danger')
        return redirect(url_for('inbound'))

    conn = get_db()
    branch_map = {r['code'].upper(): r['id']
                  for r in conn.execute('SELECT code, id FROM branches').fetchall()}
    form_map   = {r['name']: r['id']
                  for r in conn.execute('SELECT name, id FROM form_types').fetchall()}

    is_staff    = session.get('role') != 'admin'
    forced_bid  = None
    forced_code = ''
    if is_staff:
        if not session.get('branch_id'):
            flash('소속 지점이 없습니다. 관리자에게 문의하세요.', 'danger')
            conn.close()
            return redirect(url_for('inbound'))
        forced_bid  = session['branch_id']
        row = conn.execute('SELECT code FROM branches WHERE id=%s', (forced_bid,)).fetchone()
        forced_code = row['code'] if row else ''

    results = []
    ok_cnt  = 0

    all_rows = list(ws.iter_rows(values_only=True))
    data_rows = []
    for i, row in enumerate(all_rows):
        first = str(row[0]).strip() if row[0] is not None else ''
        if first in ('지점코드', '※ 지점코드·양식명은 드롭다운') or first.startswith('※'):
            continue
        if not any(row[:3]):
            continue
        data_rows.append((i + 1, row))

    for row_num, row in data_rows:
        raw_code = str(row[0]).strip().upper() if row[0] is not None else ''
        raw_form = str(row[1]).strip()         if row[1] is not None else ''
        raw_qty  = row[2]
        notes    = str(row[3]).strip()         if row[3] is not None else ''

        entry = {'row': row_num, 'branch': raw_code or forced_code,
                 'form': raw_form, 'qty': raw_qty}

        if is_staff:
            bid        = forced_bid
            entry['branch'] = forced_code
        else:
            if not raw_code:
                entry.update(status='skip', msg='지점코드 없음'); results.append(entry); continue
            bid = branch_map.get(raw_code)
            if not bid:
                entry.update(status='error', msg=f'지점코드 없음: {raw_code}')
                results.append(entry); continue

        if not raw_form:
            entry.update(status='skip', msg='양식명 없음'); results.append(entry); continue
        fid = form_map.get(raw_form)
        if not fid:
            entry.update(status='error', msg=f'양식명 불일치: {raw_form}')
            results.append(entry); continue

        try:
            qty = int(float(str(raw_qty)))
        except (ValueError, TypeError):
            entry.update(status='error', msg=f'수량 오류: {raw_qty}')
            results.append(entry); continue
        if qty <= 0:
            entry.update(status='skip', msg=f'수량 0 이하 ({qty})'); results.append(entry); continue

        conn.execute('''
            INSERT INTO inventory (branch_id, form_type_id, quantity, last_updated)
            VALUES (%s,%s,%s,NOW())
            ON CONFLICT(branch_id, form_type_id) DO UPDATE SET
              quantity     = inventory.quantity + EXCLUDED.quantity,
              last_updated = NOW()
        ''', (bid, fid, qty))
        conn.execute(
            "INSERT INTO transactions (type, form_type_id, to_branch_id, quantity, notes, created_by) "
            "VALUES ('IN',%s,%s,%s,%s,%s)",
            (fid, bid, qty,
             f'[엑셀업로드] {notes}' if notes else '[엑셀업로드]',
             session['username'])
        )

        entry.update(status='ok', qty=qty)
        results.append(entry)
        ok_cnt += 1

    conn.commit()
    conn.close()

    err_cnt  = sum(1 for r in results if r['status'] == 'error')
    skip_cnt = sum(1 for r in results if r['status'] == 'skip')
    return render_template('upload_result.html',
                           results=results, ok_cnt=ok_cnt,
                           err_cnt=err_cnt, skip_cnt=skip_cnt)


# ── 보유 지점 상세 API ────────────────────────────────────────────────────────

@app.route('/api/inventory/branches')
@login_required
def api_inventory_branches():
    fid = request.args.get('form_type_id')
    bf  = request.args.get('branch_id', '')
    if not fid:
        return jsonify([])

    conn = get_db()
    conditions, params = ['i.form_type_id=%s'], [fid]

    if session.get('role') != 'admin' and session.get('branch_id'):
        conditions.append('i.branch_id=%s')
        params.append(session['branch_id'])
    elif bf:
        conditions.append('i.branch_id=%s')
        params.append(bf)

    rows = conn.execute(f'''
        SELECT b.code branch_code, b.name branch_name, b.type branch_type,
               i.quantity, i.last_updated, f.min_threshold, f.name form_name
        FROM inventory i
        JOIN branches b ON i.branch_id = b.id
        JOIN form_types f ON i.form_type_id = f.id
        WHERE {" AND ".join(conditions)}
        ORDER BY b.type, b.code
    ''', params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# ── 지점 이메일 맵 API ────────────────────────────────────────────────────────

@app.route('/api/branches/emails')
@login_required
def api_branch_emails():
    """전체 지점 이메일 맵 반환: {branch_id: email}"""
    conn = get_db()
    rows = conn.execute('SELECT id, email FROM branches').fetchall()
    conn.close()
    return jsonify({str(r['id']): (r['email'] or '') for r in rows})


# ── 지점 이메일 관리 (관리자) ─────────────────────────────────────────────────

@app.route('/admin/branches/emails', methods=['GET', 'POST'])
@login_required
def admin_branch_emails():
    if session.get('role') != 'admin':
        flash('관리자 권한이 필요합니다.', 'danger')
        return redirect(url_for('dashboard'))
    conn = get_db()
    ph   = '%s' if not USE_SQLITE else '?'
    if request.method == 'POST':
        data = request.form.to_dict()
        for key, val in data.items():
            if key.startswith('email_'):
                bid = key[6:]  # 'email_123' → '123'
                email_val = val.strip()
                conn.execute(
                    f'UPDATE branches SET email={ph} WHERE id={ph}',
                    (email_val, bid)
                )
        conn.commit()
        conn.close()
        flash('지점 이메일이 저장되었습니다.', 'success')
        return redirect(url_for('admin_branch_emails'))
    branches = conn.execute(
        'SELECT id, code, name, type, COALESCE(email, \'\') AS email FROM branches ORDER BY type, code'
    ).fetchall()
    conn.close()
    return render_template('branch_emails.html', branches=branches)


# ── 최소기준 수정 ─────────────────────────────────────────────────────────────

@app.route('/api/update_threshold', methods=['POST'])
@login_required
def update_threshold():
    data = request.get_json()
    form_type_id = data.get('form_type_id')
    threshold    = data.get('threshold')
    if form_type_id is None or threshold is None or int(threshold) < 0:
        return jsonify({'ok': False, 'error': '잘못된 값'}), 400
    conn = get_db()
    conn.execute('UPDATE form_types SET min_threshold=%s WHERE id=%s', (int(threshold), form_type_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ── Excel 헬퍼 ────────────────────────────────────────────────────────────────

def _make_header_style():
    fill   = PatternFill('solid', fgColor='1A2340')
    font   = Font(color='FFFFFF', bold=True, size=10)
    align  = Alignment(horizontal='center', vertical='center')
    return fill, font, align

def _border():
    s = Side(style='thin', color='D0D7E3')
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_header(ws, headers, col_widths):
    fill, font, align = _make_header_style()
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill, cell.font, cell.alignment = fill, font, align
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

def _cell_border(ws, row, col):
    ws.cell(row=row, column=col).border = _border()


# ── 재고현황 Excel 다운로드 ───────────────────────────────────────────────────

@app.route('/inventory/download')
@login_required
def inventory_download():
    conn = get_db()
    bf, ff = request.args.get('branch_id',''), request.args.get('form_type_id','')
    conditions, params = ['1=1'], []

    if session.get('role') != 'admin' and session.get('branch_id'):
        conditions.append('i.branch_id=%s'); params.append(session['branch_id'])
    elif bf:
        conditions.append('i.branch_id=%s'); params.append(bf)
    if ff:
        conditions.append('i.form_type_id=%s'); params.append(ff)

    rows = conn.execute(f'''
        SELECT i.*, f.name form_name, f.unit, f.unit_detail, f.unit_price, f.min_threshold,
               b.name branch_name, b.code branch_code, b.type branch_type
        FROM inventory i
        JOIN form_types f ON i.form_type_id = f.id
        JOIN branches b ON i.branch_id = b.id
        WHERE {" AND ".join(conditions)}
        ORDER BY b.type, b.code, f.name
    ''', params).fetchall()
    conn.close()

    is_admin = session.get('role') == 'admin'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '재고현황'

    if is_admin:
        headers    = ['구분','지점코드','지점명','양식명','단위','단위상세','현재수량','최소기준','상태','단가(원)','재고금액(원)','최종업데이트']
        col_widths = [8, 10, 14, 32, 8, 12, 10, 10, 8, 12, 14, 20]
    else:
        headers    = ['구분','지점코드','지점명','양식명','단위','단위상세','현재수량','최소기준','상태','최종업데이트']
        col_widths = [8, 10, 14, 32, 8, 12, 10, 10, 8, 20]
    _apply_header(ws, headers, col_widths)

    type_map = {'DOM':'국내','INTL':'국제','CARGO':'화물'}
    green  = PatternFill('solid', fgColor='D1FAE5')
    yellow = PatternFill('solid', fgColor='FEF9C3')
    red    = PatternFill('solid', fgColor='FEE2E2')

    for r, row in enumerate(rows, 2):
        qty = row['quantity']
        thr = row['min_threshold']
        status = '소진' if qty == 0 else ('부족' if qty <= thr else '정상')
        color  = red    if qty == 0 else (yellow if qty <= thr else green)
        if is_admin:
            vals = [
                type_map.get(row['branch_type'], row['branch_type']),
                row['branch_code'], row['branch_name'], row['form_name'],
                row['unit'], row['unit_detail'], qty, thr, status,
                row['unit_price'], qty * (row['unit_price'] or 0),
                str(row['last_updated'] or '')[:16],
            ]
        else:
            vals = [
                type_map.get(row['branch_type'], row['branch_type']),
                row['branch_code'], row['branch_name'], row['form_name'],
                row['unit'], row['unit_detail'], qty, thr, status,
                str(row['last_updated'] or '')[:16],
            ]
        status_col = 9
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border    = _border()
            cell.alignment = Alignment(vertical='center')
            if c in (7, 8):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if c == status_col:
                cell.fill = color
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if is_admin and c in (10, 11):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0'
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import date
    fname = f'재고현황_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 입출고이력 Excel 다운로드 ─────────────────────────────────────────────────

@app.route('/transactions/download')
@login_required
def transactions_download():
    conn = get_db()
    bf        = request.args.get('branch_id', '')
    ff        = request.args.get('form_type_id', '')
    tf        = request.args.get('type', '')
    month_f   = request.args.get('month_f', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    conditions, params = ['1=1'], []
    role, bid = session.get('role'), session.get('branch_id')

    if role != 'admin' and bid:
        conditions.append('(t.from_branch_id=%s OR t.to_branch_id=%s)'); params.extend([bid, bid])
    elif bf:
        conditions.append('(t.from_branch_id=%s OR t.to_branch_id=%s)'); params.extend([bf, bf])
    if ff:
        conditions.append('t.form_type_id=%s'); params.append(ff)
    if tf:
        conditions.append('t.type=%s'); params.append(tf)
    if month_f:
        conditions.append("TO_CHAR(t.transaction_date, 'YYYY-MM') = %s"); params.append(month_f)
    if date_from:
        conditions.append('t.transaction_date >= %s'); params.append(date_from)
    if date_to:
        conditions.append('t.transaction_date <= %s'); params.append(date_to)

    rows = conn.execute(f'''
        SELECT t.*, f.name form_name, f.unit,
               fb.name from_branch_name, fb.code from_branch_code,
               tb.name to_branch_name,   tb.code to_branch_code
        FROM transactions t
        JOIN form_types f ON t.form_type_id = f.id
        LEFT JOIN branches fb ON t.from_branch_id = fb.id
        LEFT JOIN branches tb ON t.to_branch_id   = tb.id
        WHERE {" AND ".join(conditions)}
        ORDER BY t.transaction_date DESC, t.created_at DESC
    ''', params).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '입출고이력'

    headers    = ['거래일자','등록일시','구분','양식명','단위','출발지점코드','출발지점','도착지점코드','도착지점','수량','비고','처리자']
    col_widths = [12, 16, 8, 32, 8, 12, 14, 12, 14, 10, 20, 12]
    _apply_header(ws, headers, col_widths)

    type_kr = {'IN':'입고', 'OUT':'출고', 'TRANSFER':'이전'}
    in_fill  = PatternFill('solid', fgColor='D1FAE5')
    out_fill = PatternFill('solid', fgColor='FEE2E2')
    tr_fill  = PatternFill('solid', fgColor='DBEAFE')

    for r, row in enumerate(rows, 2):
        t_kr = type_kr.get(row['type'], row['type'])
        fill = in_fill if row['type']=='IN' else (out_fill if row['type']=='OUT' else tr_fill)
        vals = [
            str(row['transaction_date'] or '')[:10],
            str(row['created_at'] or '')[:16], t_kr, row['form_name'], row['unit'],
            row['from_branch_code'] or '', row['from_branch_name'] or '',
            row['to_branch_code']   or '', row['to_branch_name']   or '',
            row['quantity'], row['notes'] or '', row['created_by'] or '',
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border    = _border()
            cell.alignment = Alignment(vertical='center')
            if c == 3:
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if c == 10:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import date
    fname = f'입출고이력_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


with app.app_context():
    try:
        init_db()
    except Exception as _e:
        print(f"[init_db] {_e}")


# ── 5년 경과 데이터 자동 파기 (관리자 전용) ──────────────────────────────────

@app.route('/bulk-outbound')
@login_required
def bulk_outbound_page():
    conn = get_db()
    role = session.get('role')
    bid  = session.get('branch_id')
    branches = conn.execute('SELECT * FROM branches ORDER BY type, code').fetchall()

    if role == 'admin':
        bid_filter = request.args.get('branch_id', '')
        if bid_filter:
            rows = conn.execute('''
                SELECT i.branch_id, i.form_type_id, i.quantity,
                       f.name form_name, f.unit, f.min_threshold,
                       b.name branch_name, b.code branch_code
                FROM inventory i
                JOIN form_types f ON i.form_type_id=f.id
                JOIN branches  b ON i.branch_id=b.id
                WHERE i.branch_id=%s
                ORDER BY f.name
            ''', (bid_filter,)).fetchall()
        else:
            rows = conn.execute('''
                SELECT i.branch_id, i.form_type_id, i.quantity,
                       f.name form_name, f.unit, f.min_threshold,
                       b.name branch_name, b.code branch_code
                FROM inventory i
                JOIN form_types f ON i.form_type_id=f.id
                JOIN branches  b ON i.branch_id=b.id
                ORDER BY b.code, f.name
            ''').fetchall()
    else:
        rows = conn.execute('''
            SELECT i.branch_id, i.form_type_id, i.quantity,
                   f.name form_name, f.unit, f.min_threshold,
                   b.name branch_name, b.code branch_code
            FROM inventory i
            JOIN form_types f ON i.form_type_id=f.id
            JOIN branches  b ON i.branch_id=b.id
            WHERE i.branch_id=%s
            ORDER BY f.name
        ''', (bid,)).fetchall()
    conn.close()
    from datetime import date
    return render_template('bulk_outbound.html', inventory_rows=rows, branches=branches,
                           today=date.today().isoformat())


@app.route('/transfer/my-requests')
@login_required
def my_transfer_requests():
    conn = get_db()
    role = session.get('role')
    bid  = session.get('branch_id')
    status_f   = request.args.get('status', '')
    date_from  = request.args.get('date_from', '')
    date_to    = request.args.get('date_to', '')
    page       = max(1, int(request.args.get('page', 1)))
    per_page   = 20

    conditions = ['1=1']
    params     = []
    if role != 'admin' and bid:
        conditions.append('(tr.from_branch_id=%s OR tr.to_branch_id=%s)')
        params.extend([bid, bid])
    if status_f:
        conditions.append('tr.status=%s'); params.append(status_f)
    if date_from:
        conditions.append('tr.created_at >= %s'); params.append(date_from)
    if date_to:
        conditions.append('tr.created_at <= %s'); params.append(date_to + ' 23:59:59')

    total = conn.execute(
        f'SELECT COUNT(*) AS cnt FROM transfer_requests tr WHERE {" AND ".join(conditions)}',
        params
    ).fetchone()['cnt']

    rows = conn.execute(
        _TR_SELECT + f' WHERE {" AND ".join(conditions)} ORDER BY tr.created_at DESC LIMIT {per_page} OFFSET {(page-1)*per_page}',
        params
    ).fetchall()
    conn.close()

    total_pages = max(1, (total + per_page - 1) // per_page)
    return render_template('my_transfers.html', rows=rows,
                           status_f=status_f, date_from=date_from, date_to=date_to,
                           page=page, total_pages=total_pages, total=total)


@app.route('/admin/access-logs')
@login_required
def access_logs_view():
    if session.get('role') != 'admin':
        flash('관리자 권한이 필요합니다.', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db()
    username_f = request.args.get('username', '')
    action_f   = request.args.get('action', '')
    date_from  = request.args.get('date_from', '')
    date_to    = request.args.get('date_to', '')
    page       = max(1, int(request.args.get('page', 1)))
    per_page   = 50

    conditions, params = ['1=1'], []
    if username_f:
        conditions.append('username LIKE %s'); params.append(f'%{username_f}%')
    if action_f:
        conditions.append('action LIKE %s'); params.append(f'%{action_f}%')
    if date_from:
        conditions.append('created_at >= %s'); params.append(date_from)
    if date_to:
        conditions.append('created_at <= %s'); params.append(date_to + ' 23:59:59')

    total = conn.execute(
        f'SELECT COUNT(*) AS cnt FROM access_logs WHERE {" AND ".join(conditions)}', params
    ).fetchone()['cnt']
    logs = conn.execute(
        f'SELECT * FROM access_logs WHERE {" AND ".join(conditions)} '
        f'ORDER BY created_at DESC LIMIT {per_page} OFFSET {(page-1)*per_page}',
        params
    ).fetchall()
    conn.close()

    total_pages = max(1, (total + per_page - 1) // per_page)
    return render_template('access_logs.html', logs=logs,
                           username_f=username_f, action_f=action_f,
                           date_from=date_from, date_to=date_to,
                           page=page, total_pages=total_pages, total=total)


@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def notifications_read_all():
    bid = session.get('branch_id')
    if not bid and session.get('role') != 'admin':
        return jsonify({'ok': False}), 403
    conn = get_db()
    if session.get('role') == 'admin':
        conn.execute('UPDATE notifications SET is_read=1 WHERE is_read=0')
    else:
        conn.execute('UPDATE notifications SET is_read=1 WHERE branch_id=%s AND is_read=0', (bid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/bulk-outbound', methods=['POST'])
@login_required
def bulk_outbound():
    data  = request.get_json(silent=True) or {}
    items = data.get('items', [])
    if not items:
        return jsonify({'ok': False, 'msg': '출고 항목이 없습니다.'}), 400

    conn    = get_db()
    role    = session.get('role')
    my_bid  = session.get('branch_id')
    success = []
    failed  = []

    for item in items:
        try:
            bid  = int(item['branch_id'])
            fid  = int(item['form_type_id'])
            qty  = int(item['quantity'])
            note = item.get('notes', '')
            period_month = item.get('period_month', datetime.now(timezone.utc).strftime('%Y-%m'))
            tx_date = item.get('transaction_date', datetime.now(timezone.utc).strftime('%Y-%m-%d'))

            if role != 'admin' and bid != my_bid:
                failed.append({'branch_id': bid, 'form_type_id': fid, 'reason': '권한 없음'})
                continue
            if qty <= 0:
                failed.append({'branch_id': bid, 'form_type_id': fid, 'reason': '수량 오류'})
                continue

            cur = conn.execute(
                'SELECT quantity FROM inventory WHERE branch_id=%s AND form_type_id=%s', (bid, fid)
            ).fetchone()
            if not cur or cur['quantity'] < qty:
                failed.append({'branch_id': bid, 'form_type_id': fid,
                               'reason': f'재고 부족 (현재 {cur["quantity"] if cur else 0}개)'})
                continue

            conn.execute(
                'UPDATE inventory SET quantity=quantity-%s, last_updated=NOW() '
                'WHERE branch_id=%s AND form_type_id=%s', (qty, bid, fid)
            )
            conn.execute(
                "INSERT INTO transactions "
                "(type, form_type_id, from_branch_id, quantity, notes, created_by, period_month, transaction_date) "
                "VALUES ('OUT',%s,%s,%s,%s,%s,%s,%s)",
                (fid, bid, qty, note, session['username'], period_month, tx_date)
            )
            # 재고 부족 알림
            inv_after = conn.execute(
                'SELECT i.quantity, f.min_threshold, f.name fn, b.name bn '
                'FROM inventory i JOIN form_types f ON i.form_type_id=f.id '
                'JOIN branches b ON i.branch_id=b.id '
                'WHERE i.branch_id=%s AND i.form_type_id=%s', (bid, fid)
            ).fetchone()
            if inv_after and inv_after['quantity'] <= inv_after['min_threshold']:
                dup = conn.execute(
                    "SELECT id FROM notifications WHERE branch_id=%s AND is_read=0 AND message LIKE %s LIMIT 1",
                    (bid, f'%{inv_after["fn"]}%')
                ).fetchone()
                if not dup:
                    sl = '소진' if inv_after['quantity'] == 0 else '부족'
                    conn.execute(
                        "INSERT INTO notifications (branch_id, message) VALUES (%s,%s)",
                        (bid, f'[재고{sl}] {inv_after["bn"]} — {inv_after["fn"]} 잔여 {inv_after["quantity"]}개')
                    )
            success.append({'branch_id': bid, 'form_type_id': fid, 'quantity': qty})
        except Exception as e:
            failed.append({'branch_id': item.get('branch_id'), 'form_type_id': item.get('form_type_id'),
                           'reason': str(e)})

    if success:
        conn.commit()
    conn.close()
    return jsonify({'ok': True, 'processed': len(success), 'failed': failed})


@app.route('/admin/access-logs/download')
@login_required
def download_access_logs():
    if session.get('role') != 'admin':
        flash('관리자 권한이 필요합니다.', 'danger')
        return redirect(url_for('dashboard'))

    import csv
    from io import StringIO

    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    table     = request.args.get('table', 'access_logs')

    if table not in ('access_logs', 'access_logs_archive'):
        table = 'access_logs'

    conn = get_db()
    params = []
    where  = []
    if date_from:
        where.append('created_at >= %s'); params.append(date_from)
    if date_to:
        where.append('created_at <= %s'); params.append(date_to + ' 23:59:59')

    sql = f'SELECT id, user_id, username, action, target_info, ip_address, created_at FROM {table}'
    if where:
        sql += ' WHERE ' + ' AND '.join(where)
    sql += ' ORDER BY created_at DESC'

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    si = StringIO()
    w = csv.writer(si)
    w.writerow(['ID', '사용자ID', '사용자명', '행위', '대상정보', 'IP주소', '일시'])
    for r in rows:
        w.writerow([r['id'], r['user_id'], r['username'], r['action'],
                    r['target_info'], r['ip_address'], r['created_at']])

    log_action('접속로그_다운로드', f'{table}, {len(rows)}건')

    fname = f'access_logs_{datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")}.csv'
    return send_file(
        BytesIO(si.getvalue().encode('utf-8-sig')),
        as_attachment=True, download_name=fname,
        mimetype='text/csv'
    )


@app.route('/admin/purge-old-data', methods=['POST'])
@login_required
def purge_old_data():
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'error': '관리자 권한 필요'}), 403

    conn = get_db()
    results = {}

    # 5년(1826일) 이전 거래 이력 삭제
    r = conn.execute(
        "DELETE FROM transactions WHERE created_at < NOW() - INTERVAL '5 years' RETURNING id"
    ).fetchall()
    results['transactions'] = len(r)

    # 5년 이전 접속 로그 삭제 (1년 보관 의무이므로 5년은 충분히 보관 후 파기)
    r = conn.execute(
        "DELETE FROM access_logs WHERE created_at < NOW() - INTERVAL '5 years' RETURNING id"
    ).fetchall()
    results['access_logs'] = len(r)

    conn.commit()

    log_action('5년경과데이터파기', f"transactions:{results['transactions']}건, logs:{results['access_logs']}건")
    conn.close()

    return jsonify({'ok': True, 'purged': results})


# ── 운송양식 공급 신청 시스템 ──────────────────────────────────────────────────

@app.route('/admin/form-supply/settings', methods=['GET', 'POST'])
@login_required
def form_supply_settings():
    """관리자 — 운송양식 신청 기간 설정"""
    if session.get('role') != 'admin':
        flash('관리자만 접근 가능합니다.', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db()
    ph = '%s' if not USE_SQLITE else '?'
    now_sql = 'NOW()' if not USE_SQLITE else "datetime('now')"

    # lazy migration — 컬럼이 없으면 즉시 추가
    try:
        if USE_SQLITE:
            _cols = [r[1] for r in conn.execute('PRAGMA table_info(form_supply_settings)').fetchall()]
            if 'title' not in _cols:
                conn.execute("ALTER TABLE form_supply_settings ADD COLUMN title TEXT NOT NULL DEFAULT ''")
            _cols2 = [r[1] for r in conn.execute('PRAGMA table_info(form_supply_requests)').fetchall()]
            if 'period_title' not in _cols2:
                conn.execute("ALTER TABLE form_supply_requests ADD COLUMN period_title TEXT NOT NULL DEFAULT ''")
        else:
            conn.execute("ALTER TABLE form_supply_settings ADD COLUMN IF NOT EXISTS title TEXT NOT NULL DEFAULT ''")
            conn.execute("ALTER TABLE form_supply_requests ADD COLUMN IF NOT EXISTS period_title TEXT NOT NULL DEFAULT ''")
        conn.commit()
    except Exception as _me:
        app.logger.warning(f'[lazy-migrate] {_me}')
        try: conn.rollback()
        except Exception: pass

    if request.method == 'POST':
        period_title = request.form.get('period_title', '').strip()
        period_start = request.form.get('period_start', '').strip()
        period_end   = request.form.get('period_end', '').strip()
        is_enabled   = 1 if request.form.get('is_enabled') else 0

        if not period_start or not period_end:
            flash('시작일과 종료일을 모두 입력해주세요.', 'danger')
            conn.close()
            return redirect(url_for('form_supply_settings'))
        if period_start > period_end:
            flash('시작일은 종료일보다 빠르거나 같아야 합니다.', 'danger')
            conn.close()
            return redirect(url_for('form_supply_settings'))

        conn.execute(
            f"INSERT INTO form_supply_settings (title, period_start, period_end, is_enabled, created_by, updated_at) "
            f"VALUES ({ph},{ph},{ph},{ph},{ph},{now_sql})",
            (period_title, period_start, period_end, is_enabled, session['username'])
        )
        conn.commit()
        log_action('운송양식_신청기간_설정', f'[{period_title}] {period_start}~{period_end} (활성:{is_enabled})')
        flash('신청 기간 설정이 저장되었습니다.', 'success')
        conn.close()
        return redirect(url_for('form_supply_settings'))

    # GET — 현재 최신 설정 + 이력
    current = conn.execute(
        'SELECT * FROM form_supply_settings ORDER BY id DESC LIMIT 1'
    ).fetchone()
    settings_history = conn.execute(
        'SELECT * FROM form_supply_settings ORDER BY id DESC LIMIT 50'
    ).fetchall()
    # SQLite Row를 dict처럼 다루기 위해 created_at 대체 (테이블에는 updated_at만 있음)
    history_list = []
    for s in settings_history:
        d = dict(s)
        d['created_at'] = d.get('updated_at')
        history_list.append(d)
    form_types = conn.execute('SELECT * FROM form_types WHERE is_active ORDER BY sort_order').fetchall()
    conn.close()
    return render_template('form_supply_settings.html',
                           current=current,
                           settings_history=history_list,
                           form_types=form_types)


@app.route('/admin/form-supply/settings/<int:setting_id>/edit', methods=['POST'])
@login_required
def form_supply_setting_edit(setting_id):
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'error': '권한 없음'}), 403
    data = request.get_json(force=True)
    title        = data.get('title', '').strip()
    period_start = data.get('period_start', '').strip()
    period_end   = data.get('period_end', '').strip()
    is_enabled   = 1 if data.get('is_enabled') else 0
    if not period_start or not period_end or period_start > period_end:
        return jsonify({'ok': False, 'error': '날짜가 올바르지 않습니다.'}), 400
    ph = '%s' if not USE_SQLITE else '?'
    now_sql = 'NOW()' if not USE_SQLITE else "datetime('now')"
    try:
        conn = get_db()
        conn.execute(
            f'UPDATE form_supply_settings SET title={ph}, period_start={ph}, period_end={ph}, is_enabled={ph}, updated_at={now_sql} WHERE id={ph}',
            (title, period_start, period_end, is_enabled, setting_id)
        )
        conn.commit()
        conn.close()
        log_action('운송양식_신청기간_수정', f'#{setting_id} [{title}] {period_start}~{period_end}')
        return jsonify({'ok': True})
    except Exception as e:
        app.logger.error(f'[setting_edit] {e}')
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/form-supply/settings/<int:setting_id>/delete', methods=['POST'])
@login_required
def form_supply_setting_delete(setting_id):
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'error': '권한 없음'}), 403
    ph = '%s' if not USE_SQLITE else '?'
    try:
        conn = get_db()
        conn.execute(f'DELETE FROM form_supply_settings WHERE id={ph}', (setting_id,))
        conn.commit()
        conn.close()
        log_action('운송양식_신청기간_삭제', f'#{setting_id}')
        return jsonify({'ok': True})
    except Exception as e:
        app.logger.error(f'[setting_delete] {e}')
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/form-type/<int:form_id>/memo', methods=['POST'])
@login_required
def admin_form_type_memo(form_id):
    if session.get('role') != 'admin':
        return {'ok': False, 'error': '권한 없음'}, 403
    memo = request.get_json(force=True).get('memo', '').strip()
    ph = '%s' if not USE_SQLITE else '?'
    conn = get_db()
    conn.execute(f'UPDATE form_types SET memo={ph} WHERE id={ph}', (memo, form_id))
    conn.commit()
    conn.close()
    return {'ok': True}


@app.route('/api/form-supply/period')
@login_required
def api_form_supply_period():
    """현재 신청 기간 상태 API"""
    period = _get_active_supply_period()
    if period:
        return jsonify({
            'active': True,
            'id': period['id'],
            'period_start': period['period_start'],
            'period_end':   period['period_end'],
        })
    # 활성 기간이 없을 때 가장 최근 설정 정보 반환
    conn = get_db()
    latest = conn.execute(
        'SELECT * FROM form_supply_settings ORDER BY id DESC LIMIT 1'
    ).fetchone()
    conn.close()
    if latest:
        return jsonify({
            'active': False,
            'id': latest['id'],
            'period_start': latest['period_start'],
            'period_end':   latest['period_end'],
        })
    return jsonify({'active': False})


@app.route('/form-supply/request', methods=['GET', 'POST'])
@login_required
def form_supply_request():
    """직원 — 운송양식 신청"""
    role = session.get('role')
    bid  = session.get('branch_id')

    if role == 'admin':
        flash('관리자는 신청 페이지를 사용할 수 없습니다. 기간 설정 화면으로 이동합니다.', 'info')
        return redirect(url_for('form_supply_settings'))
    if not bid:
        flash('소속 지점이 없습니다. 관리자에게 문의하세요.', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db()
    ph = '%s' if not USE_SQLITE else '?'
    now_sql = 'NOW()' if not USE_SQLITE else "datetime('now')"

    # 활성 기간 / 최근 설정 조회 (템플릿용)
    today = datetime.now().strftime('%Y-%m-%d')
    latest = conn.execute(
        'SELECT * FROM form_supply_settings ORDER BY id DESC LIMIT 1'
    ).fetchone()
    period_ctx = None
    if latest:
        d = dict(latest)
        d['in_range'] = (
            str(d.get('period_start', '')) <= today <= str(d.get('period_end', ''))
        )
        period_ctx = d

    if request.method == 'POST':
        # 서버사이드 기간 검증
        active = _get_active_supply_period()
        if not active:
            flash('현재 신청 기간이 아닙니다.', 'danger')
            conn.close()
            return redirect(url_for('form_supply_request'))

        notes = request.form.get('notes', '').strip()

        # qty_<form_type_id> 필드 수집
        items = []
        for ft in conn.execute('SELECT id FROM form_types').fetchall():
            fid = ft['id']
            raw = request.form.get(f'qty_{fid}', '').strip()
            if not raw:
                continue
            try:
                q = int(raw)
            except ValueError:
                continue
            if q >= 1:
                items.append((fid, min(q, 9999)))

        if not items:
            flash('최소 1개 이상의 양식과 수량을 선택해주세요.', 'danger')
            conn.close()
            return redirect(url_for('form_supply_request'))

        # 현재 기간 제목 가져오기
        p_title = (active.get('title') or '') if isinstance(active, dict) else (active['title'] if 'title' in active.keys() else '')

        # 신청서 INSERT
        if USE_SQLITE:
            cur = conn.execute(
                f"INSERT INTO form_supply_requests (branch_id, status, notes, period_title, requested_by, created_at, updated_at) "
                f"VALUES ({ph},'pending',{ph},{ph},{ph},{now_sql},{now_sql})",
                (bid, notes, p_title, session['username'])
            )
            req_id = cur.lastrowid
        else:
            row = conn.execute(
                "INSERT INTO form_supply_requests (branch_id, status, notes, period_title, requested_by) "
                "VALUES (%s,'pending',%s,%s,%s) RETURNING id",
                (bid, notes, p_title, session['username'])
            ).fetchone()
            req_id = row['id']

        # 항목 INSERT
        for fid, q in items:
            conn.execute(
                f"INSERT INTO form_supply_request_items (request_id, form_type_id, quantity) "
                f"VALUES ({ph},{ph},{ph})",
                (req_id, fid, q)
            )

        # 관리자 알림 — admin 사용자의 branch_id가 NULL일 수 있어 신청 지점에 사본 알림만 남김
        b_name_row = conn.execute('SELECT name FROM branches WHERE id=%s', (bid,)).fetchone()
        b_name = b_name_row['name'] if b_name_row else f'지점#{bid}'
        msg = f"[운송양식 신청] {b_name} — {len(items)}종 신청 (#{req_id})"
        conn.execute(
            "INSERT INTO notifications (branch_id, message) VALUES (%s,%s)",
            (bid, msg)
        )
        conn.commit()
        log_action('운송양식_신청', f'#{req_id} {len(items)}종')
        flash(f'신청 완료. ({len(items)}종 / 신청번호 #{req_id})', 'success')
        conn.close()
        return redirect(url_for('form_supply_my_requests'))

    # GET — 폼 렌더
    form_types = conn.execute('SELECT * FROM form_types WHERE is_active ORDER BY sort_order').fetchall()
    conn.close()
    return render_template('form_supply_request.html',
                           period=period_ctx,
                           form_types=form_types)


@app.route('/form-supply/my-requests')
@login_required
def form_supply_my_requests():
    """직원 — 내 신청 현황"""
    role = session.get('role')
    bid  = session.get('branch_id')

    if role == 'admin':
        return redirect(url_for('form_supply_admin_requests'))
    if not bid:
        flash('소속 지점이 없습니다.', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db()
    reqs = conn.execute(
        'SELECT * FROM form_supply_requests WHERE branch_id=%s ORDER BY id DESC',
        (bid,)
    ).fetchall()

    result = []
    for r in reqs:
        items = conn.execute(
            'SELECT i.form_type_id, i.quantity, f.name AS form_name, f.unit, f.unit_detail '
            'FROM form_supply_request_items i '
            'JOIN form_types f ON f.id = i.form_type_id '
            'WHERE i.request_id=%s ORDER BY f.name',
            (r['id'],)
        ).fetchall()
        d = dict(r)
        d['items'] = [dict(x) for x in items]
        result.append(d)
    conn.close()
    return render_template('form_supply_my_requests.html', requests=result)


@app.route('/admin/form-supply/requests')
@login_required
def form_supply_admin_requests():
    """관리자 — 전체 신청 목록"""
    if session.get('role') != 'admin':
        flash('관리자만 접근 가능합니다.', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db()
    reqs = conn.execute(
        'SELECT r.*, b.code AS branch_code, b.name AS branch_name '
        'FROM form_supply_requests r '
        'JOIN branches b ON b.id = r.branch_id '
        'ORDER BY r.id DESC'
    ).fetchall()

    result = []
    for r in reqs:
        items = conn.execute(
            'SELECT i.form_type_id, i.quantity, f.name AS form_name, f.unit, f.unit_detail '
            'FROM form_supply_request_items i '
            'JOIN form_types f ON f.id = i.form_type_id '
            'WHERE i.request_id=%s ORDER BY f.name',
            (r['id'],)
        ).fetchall()
        d = dict(r)
        d['items'] = [dict(x) for x in items]
        result.append(d)

    # 기간별 제목 목록 (중복 제거, 최신순)
    seen = set()
    period_titles = []
    for r in result:
        t = r.get('period_title') or ''
        if t and t not in seen:
            seen.add(t)
            period_titles.append(t)

    conn.close()
    return render_template('form_supply_admin_requests.html', requests=result, period_titles=period_titles)


@app.route('/admin/form-supply/requests/<int:req_id>/action', methods=['POST'])
@login_required
def form_supply_admin_action(req_id):
    """관리자 — 승인/반려 처리"""
    if session.get('role') != 'admin':
        flash('관리자만 접근 가능합니다.', 'danger')
        return redirect(url_for('dashboard'))

    action = request.form.get('action', '').strip()
    reject_reason  = request.form.get('reject_reason', '').strip()
    approve_reason = request.form.get('approve_reason', '').strip()

    if action not in ('approved', 'rejected'):
        flash('잘못된 요청입니다.', 'danger')
        return redirect(url_for('form_supply_admin_requests'))
    if action == 'rejected' and not reject_reason:
        flash('반려 사유를 입력해주세요.', 'danger')
        return redirect(url_for('form_supply_admin_requests'))

    conn = get_db()
    ph = '%s' if not USE_SQLITE else '?'
    now_sql = 'NOW()' if not USE_SQLITE else "datetime('now')"

    req = conn.execute(
        'SELECT r.*, b.name AS branch_name, b.email AS branch_email FROM form_supply_requests r '
        'JOIN branches b ON b.id = r.branch_id WHERE r.id=%s',
        (req_id,)
    ).fetchone()
    if not req:
        flash('신청을 찾을 수 없습니다.', 'danger')
        conn.close()
        return redirect(url_for('form_supply_admin_requests'))
    if req['status'] != 'pending':
        flash('이미 처리된 신청입니다.', 'warning')
        conn.close()
        return redirect(url_for('form_supply_admin_requests'))

    conn.execute(
        f"UPDATE form_supply_requests "
        f"SET status={ph}, reject_reason={ph}, approve_reason={ph}, processed_by={ph}, processed_at={now_sql}, updated_at={now_sql} "
        f"WHERE id={ph}",
        (action,
         reject_reason if action == 'rejected' else '',
         approve_reason if action == 'approved' else '',
         session['username'], req_id)
    )

    if action == 'approved':
        notif = f"[운송양식 신청 승인] 신청 #{req_id}가 승인되었습니다."
    else:
        notif = f"[운송양식 신청 반려] 신청 #{req_id}가 반려되었습니다. 사유: {reject_reason}"
    conn.execute(
        "INSERT INTO notifications (branch_id, message) VALUES (%s,%s)",
        (req['branch_id'], notif)
    )
    conn.commit()

    # ── 이메일 발송 ──────────────────────────────────────────────────
    branch_email = (req.get('branch_email') or '').strip()
    if branch_email:
        items = conn.execute(
            'SELECT i.quantity, ft.name AS form_name, ft.unit, ft.unit_detail '
            'FROM form_supply_request_items i '
            'JOIN form_types ft ON ft.id = i.form_type_id '
            'WHERE i.request_id=%s ORDER BY ft.sort_order',
            (req_id,)
        ).fetchall()

        if action == 'approved':
            mail_subject = f'[이스타항공] 운송양식 신청 #{req_id} 승인 안내'
        else:
            mail_subject = f'[이스타항공] 운송양식 신청 #{req_id} 반려 안내'

        item_lines = '\n'.join(
            f"  - {it['form_name']} / {it['unit'] or '—'} / {it['unit_detail'] or '—'} / {it['quantity']}개"
            for it in items
        )
        if action == 'approved':
            mail_body = (
                f"{req['branch_name']} 지점의 운송양식 신청 #{req_id}이 승인되었습니다.\n\n"
                + (f"전달사항: {approve_reason}\n\n" if approve_reason else '')
                + f"[신청 항목]\n{item_lines}\n\n"
                f"ZEGO에 로그인하여 내 신청 내역에서 확인해 주세요."
            )
        else:
            mail_body = (
                f"{req['branch_name']} 지점의 운송양식 신청 #{req_id}이 반려되었습니다.\n\n"
                f"반려 사유: {reject_reason}\n\n"
                f"[신청 항목]\n{item_lines}\n\n"
                f"ZEGO에 로그인하여 내 신청 내역에서 확인해 주세요."
            )
        send_mail([branch_email], mail_subject, mail_body)
    # ─────────────────────────────────────────────────────────────────

    log_action(f'운송양식_{action}', f'#{req_id} {req["branch_name"]}')
    flash(f'신청 #{req_id} 처리 완료 ({action}).', 'success')
    conn.close()
    return redirect(url_for('form_supply_admin_requests'))


@app.route('/admin/form-supply/matrix')
@login_required
def form_supply_matrix():
    """관리자 — 전 지점 양식 신청 현황 매트릭스"""
    if session.get('role') != 'admin':
        flash('관리자만 접근 가능합니다.', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db()

    form_types = conn.execute(
        'SELECT id, name FROM form_types WHERE is_active ORDER BY sort_order'
    ).fetchall()

    branches_dom  = conn.execute(
        "SELECT id, code, name FROM branches WHERE type='DOM' ORDER BY code"
    ).fetchall()
    branches_intl = conn.execute(
        "SELECT id, code, name FROM branches WHERE type='INTL' ORDER BY code"
    ).fetchall()

    rows = conn.execute(
        'SELECT i.form_type_id, r.branch_id, i.quantity, r.created_at, r.status, r.period_title '
        'FROM form_supply_request_items i '
        'JOIN form_supply_requests r ON r.id = i.request_id '
        'ORDER BY r.created_at'
    ).fetchall()
    conn.close()

    # 기간 제목 목록 (중복 제거)
    seen_t = set()
    period_titles = []
    for row in rows:
        t = row['period_title'] or ''
        if t and t not in seen_t:
            seen_t.add(t)
            period_titles.append(t)

    # pivot[period_title or ''][form_type_id][branch_id] = [{'qty','date','status'}]
    # 전체용 pivot도 별도 구성
    def build_pivot(filtered_rows):
        p = {ft['id']: {} for ft in form_types}
        for row in filtered_rows:
            fid = row['form_type_id']
            bid = row['branch_id']
            if fid not in p:
                continue
            if bid not in p[fid]:
                p[fid][bid] = []
            raw = str(row['created_at'] or '')
            date_label = raw[5:10].replace('-', '.')
            p[fid][bid].append({'qty': row['quantity'], 'date': date_label, 'status': row['status']})
        return p

    pivot_all = build_pivot(rows)
    pivot_by_title = {t: build_pivot([r for r in rows if (r['period_title'] or '') == t]) for t in period_titles}

    return render_template('form_supply_matrix.html',
                           form_types=form_types,
                           branches_dom=branches_dom,
                           branches_intl=branches_intl,
                           pivot_all=pivot_all,
                           pivot_by_title=pivot_by_title,
                           period_titles=period_titles)


if __name__ == '__main__':
    pass
    _port = int(os.environ.get('PORT', '5000'))
    _host = os.environ.get('FLASK_RUN_HOST', '127.0.0.1')
    print(f'\n  로컬에서 브라우저 주소창에 입력: http://127.0.0.1:{_port}/\n'
          f'  (Windows에서 5000번이 안 열리면 설정 → AirPlay 수신 끄기, 또는 PORT=5001 로 실행)\n')
    app.run(debug=True, host=_host, port=_port)
